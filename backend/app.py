"""
Flask backend for the Aapryl Clone Tool
Results are cached to disk so browser refreshes don't lose data.
"""
import os, json, threading, pickle, math, sys
import numpy as np
import pandas as pd
from flask import Flask, request, jsonify
from werkzeug.utils import secure_filename
from data_loader import (load_factor_returns, load_manager_returns, load_weights,
                         run_cloning, build_portfolio_view, fuzzy_match,
                         diagnose_clone_determinism, CANONICAL_BENCHMARKS)
from s3_storage import save_uploaded_file, resolve_path, is_enabled as s3_enabled

# Optional Postgres-backed client management. When the DB is reachable it is
# the source of truth for the client roster (names, benchmarks, managers,
# weights); everything else (clone results, risk, exposures) still comes from
# the file/pickle pipeline. When the DB is unreachable the app falls back to
# the Excel/cache path so it still boots. Import defensively so a missing DB
# dependency never stops the backend from starting.
try:
    import db as client_db
except Exception as _db_import_err:  # noqa: BLE001
    client_db = None
    print(f"[db] client-management DB unavailable at import: {_db_import_err}")

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['CACHE_FILE']    = os.path.join(os.path.dirname(__file__), 'cache', 'results.pkl')
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024

# Allow frontend origin — set FRONTEND_URL env var in production
# e.g. FRONTEND_URL=https://main.d1234.amplifyapp.com
_frontend_url = os.environ.get('FRONTEND_URL', '')
if _frontend_url:
    try:
        from flask_cors import CORS
        CORS(app, origins=[_frontend_url])
        print(f"CORS enabled for {_frontend_url}")
    except ImportError:
        print("flask-cors not installed — skipping CORS setup")

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(os.path.dirname(app.config['CACHE_FILE']), exist_ok=True)

state = {
    'clone_results': None, 'manager_dfs': None, 'weights': None,
    'client_benchmarks': {},
    'risk_data': None,
    'universe_clone_results': None, 'universe_dfs': None,
    'exposures_data': None,
    'norm_skill_by_tab': {},
    'progress': [], 'running': False, 'error': None, 'files': {},
    'progress_current': 0, 'progress_total': 0, 'progress_phase': '',
    # Basenames of the files used for the most recent buy-list clone run.
    # Compared against currently-uploaded filenames to surface a staleness
    # warning when the user has uploaded new files but not yet re-run clones.
    'clone_run_files': {},   # {'manager_returns': basename, 'factor_returns': basename}
    # Stock-level risk data from FactSet Security-Level Risk DNA export.
    # {managers: {mgr: [stocks]}, benchmarks: {bmk: {factor: val}}, ...}
    'security_risk_data': None,
    # User-edited style buckets for placeholder managers (those in the
    # weights file but missing clone data). Keyed by weight-file manager
    # name. Empty dict means "use default Core=1.0". Persisted across runs.
    'placeholder_buckets': {},
    # Parsed firm/strategy qualitative data (firm AUM, ownership description,
    # diverse/woman-owned %, per-strategy AUM). Populated by /upload_qualitative,
    # read by _qual_lookup for diverse-ownership rollups and manager-row
    # enrichment. Persisted so it survives restarts.
    'qualitative_data': None,
    # Per-tab market-cycle universe classification state (heavy O(N) downside-
    # capture computation). Precomputed at /run_universe, consumed by
    # /market_cycle, invalidated when a new factor_returns file is uploaded.
    'mc_universe_cache': {},
}

# ── Cache helpers ──────────────────────────────────────────────────────────
def save_cache():
    try:
        with open(app.config['CACHE_FILE'], 'wb') as f:
            pickle.dump({
                'clone_results':           state['clone_results'],
                'weights':                 state['weights'],
                'client_benchmarks':       state['client_benchmarks'],
                'risk_data':               state['risk_data'],
                'universe_clone_results':  state['universe_clone_results'],
                'exposures_data':          state['exposures_data'],
                'norm_skill_by_tab':       state['norm_skill_by_tab'],
                'files':                   state['files'],
                'clone_run_files':         state['clone_run_files'],
                'security_risk_data':      state['security_risk_data'],
                'placeholder_buckets':     state.get('placeholder_buckets') or {},
                'qualitative_data':        state.get('qualitative_data'),
                'mc_universe_cache':       state.get('mc_universe_cache') or {},
            }, f)
        print("Cache saved.")
        # Also push the cache to S3 so it survives container/instance
        # replacement (Fargate tasks have ephemeral disks). On next boot,
        # load_cache() pulls it back if there's no local copy.
        _push_cache_to_s3()
    except Exception as e:
        print(f"Cache save error: {e}")


# S3 object key for the persisted cache pickle.
_CACHE_S3_KEY = 'state/results.pkl'


def _push_cache_to_s3():
    if not s3_enabled():
        return
    try:
        from s3_storage import S3_BUCKET
        import boto3, sys
        # Store at the bucket ROOT key state/results.pkl (NOT under the uploads/
        # prefix) — matches where the cache is seeded and kept.
        boto3.client('s3').upload_file(app.config['CACHE_FILE'], S3_BUCKET, _CACHE_S3_KEY)
        print(f"[s3] cache pushed to s3://{S3_BUCKET}/{_CACHE_S3_KEY}", flush=True)
    except Exception as e:  # noqa: BLE001
        print(f"[s3] cache upload skipped: {e}", flush=True)


def _pull_cache_from_s3():
    """If there's no local cache but one exists in S3, download it. Lets a
    fresh Fargate task recover the last saved analytical state (clone results,
    uploaded-file references) without a manual re-run."""
    if not s3_enabled():
        return
    cf = app.config['CACHE_FILE']
    if os.path.exists(cf):
        print(f"[s3] local cache present at {cf}; not pulling from S3.", flush=True)
        return
    try:
        from s3_storage import S3_BUCKET
        import boto3
        # Cache lives at the bucket ROOT (state/results.pkl), no uploads/ prefix.
        boto3.client('s3').download_file(S3_BUCKET, _CACHE_S3_KEY, cf)
        print(f"[s3] cache restored from s3://{S3_BUCKET}/{_CACHE_S3_KEY}", flush=True)
    except Exception as e:  # noqa: BLE001
        print(f"[s3] no cache in S3 to restore ({e}); starting fresh.", flush=True)


# ── Client-DB helpers ───────────────────────────────────────────────────────
def db_enabled():
    """True iff the Postgres client DB is importable AND reachable."""
    return bool(client_db) and client_db.is_enabled()


def refresh_clients_from_db():
    """Load the client roster from Postgres into app state.

    Overrides state['weights'] and state['client_benchmarks'] with the DB
    values so the DB is the single source of truth for clients. All other
    state (clone_results, risk, exposures, …) is untouched. No-op when the DB
    is unavailable, so the Excel/cache path remains in effect.

    Returns True if the roster was loaded from the DB, False otherwise.
    """
    if not db_enabled():
        return False
    try:
        weights, benchmarks = client_db.load_weights_from_db()
        # Guard: never wipe a populated roster with an empty DB. If Postgres is
        # reachable but has no clients (seed skipped/failed, or a transient
        # empty read), keep whatever the cache/Excel already loaded rather than
        # blanking the app. Only overwrite state when the DB actually has clients.
        if not weights:
            existing = len(state.get('weights') or {})
            if existing:
                print(f"[db] DB roster is empty — keeping {existing} cached clients "
                      "(not overwriting with an empty roster).")
                return False
            # Both empty: nothing to preserve, fall through to set the empty roster.
        state['weights'] = weights
        state['client_benchmarks'] = benchmarks
        print(f"[db] client roster loaded from Postgres — {len(weights)} clients.")
        return True
    except Exception as e:  # noqa: BLE001
        print(f"[db] failed to load clients from DB (keeping existing state): {e}")
        return False


def sync_weights_to_state(weights=None, benchmarks=None):
    """Reconcile a freshly-parsed Excel weights book with app state.

    When the DB is authoritative, an uploaded/reloaded workbook is IMPORTED
    into Postgres (upsert) and then state is reloaded from the DB — so the DB
    stays the single source of truth and the workbook can still be used to bulk-
    update. When the DB is off, this just sets state from the parsed workbook
    (the legacy behavior). Pass already-parsed (weights, benchmarks) or let the
    caller set them first; returns True when the DB import path ran.

    IMPORTANT: preserves user-entered proposed weights / style buckets on
    existing managers — upsert_from_weight_maps only refreshes current weights
    and benchmarks, leaving drafts intact for managers that still exist.
    """
    if not db_enabled():
        return False
    if weights is None:
        weights = state.get('weights') or {}
        benchmarks = state.get('client_benchmarks') or {}
    try:
        client_db.repository.upsert_from_weight_maps(weights, benchmarks)
        refresh_clients_from_db()
        print(f"[db] imported {len(weights)} clients from workbook into Postgres.")
        return True
    except Exception as e:  # noqa: BLE001
        print(f"[db] workbook→DB import failed (keeping DB roster): {e}")
        # Re-assert the DB roster so a failed import never leaves Excel data live.
        refresh_clients_from_db()
        return False


# Managers excluded from norm-skill computation due to known data errors.
NORM_SKILL_EXCLUDE = set()   # Henry James SC returns have been corrected — no longer excluded


# ── Style bucket post-processing ──────────────────────────────────────────
_STYLE_BUCKET_KEYS = ['Core','Value','Growth','Yield','Quality',
                      'Dynamic','Defensive','Low Vol']

def absorb_regional_into_core(clone_results):
    """For each manager, add any non-style weight (residual from summing the
    8 style buckets) into the Core bucket so style_buckets always sums to
    100%. This makes 'Core' include regional and country factors (UK, Japan,
    Euro ex UK, etc.) that don't carry a style label.

    Idempotent: re-running is a no-op because the residual is already zero.
    Mutates the dicts in place so every consumer (peer table, portfolio
    summary, market cycle, etc.) sees the same adjusted buckets."""
    if not clone_results:
        return
    for tab, td in clone_results.items():
        if not isinstance(td, dict):
            continue
        for _, d in td.items():
            sb = d.get('style_buckets')
            if not isinstance(sb, dict):
                continue
            style_sum = sum((sb.get(k, 0) or 0) for k in _STYLE_BUCKET_KEYS)
            residual  = 1.0 - style_sum
            if residual > 0.0005:
                sb['Core'] = (sb.get('Core', 0) or 0) + residual


def recompute_norm_skill(tabs=None):
    """Recompute normalized-skill Z-scores for the given tabs (or all tabs
    with clone results). Safe to call whenever clone_results or
    universe_clone_results change. Silent no-op if there is nothing to do.

    Uses skill_engine's default target_managers=None (filters to buy-list
    managers only) — the post-clone snapshot is fast enough now (~5 sec on
    a US tab with ~3K universe managers) that no progress UI is needed.
    """
    try:
        from skill_engine import compute_norm_skill_latest
    except Exception as e:
        print(f"skill_engine import failed: {e}")
        return
    cr  = state.get('clone_results') or {}
    ucr = state.get('universe_clone_results') or {}
    if tabs is None:
        tabs = sorted(set(cr.keys()) | set(ucr.keys()))
    if 'norm_skill_by_tab' not in state or state['norm_skill_by_tab'] is None:
        state['norm_skill_by_tab'] = {}
    for t in tabs:
        try:
            state['norm_skill_by_tab'][t] = compute_norm_skill_latest(
                cr, ucr, t, exclude=NORM_SKILL_EXCLUDE,
            )
            n = len(state['norm_skill_by_tab'][t])
            print(f"Normalized skill recomputed for {t}: {n} managers.")
        except Exception as e:
            import traceback
            print(f"Norm-skill compute failed for {t}: {e}\n{traceback.format_exc()}")

def _resolve_cached_path(p):
    """Cached file paths may have been saved as absolute paths on a different
    machine, or as bare filenames after a zip round-trip. Try both: the path
    as-is, then as a filename inside uploads/. Returns a resolved absolute
    path if the file exists, otherwise None.
    """
    if not p:
        return None
    # S3 mode: the cached path may be a stale absolute path from another
    # machine (e.g. a Windows dev box: c:\...\uploads\file.xlsx). The S3 object
    # is stored under the prefix by BASENAME, so strip to the basename and let
    # resolve_path download it from S3. This lets a fresh cloud task rebind
    # every cached file to its S3 copy instead of failing on the dead path.
    if s3_enabled():
        basename = os.path.basename(str(p).replace('\\', '/'))
        resolved = resolve_path(basename, app.config['UPLOAD_FOLDER'])
        if resolved and os.path.exists(resolved):
            return resolved
        # Fall back to trying the path as given (in case it was already a key).
        resolved = resolve_path(p, app.config['UPLOAD_FOLDER'])
        return resolved if resolved and os.path.exists(resolved) else None
    if os.path.isabs(p) and os.path.exists(p):
        return p
    # Relative path, try as-is (relative to cwd)
    if os.path.exists(p):
        return os.path.abspath(p)
    # Fall back: interpret as a filename in the uploads folder. Cached paths
    # may have Windows backslashes even when the app is running in Linux.
    basename = os.path.basename(str(p).replace('\\', '/'))
    alt = os.path.join(app.config['UPLOAD_FOLDER'], basename)
    if os.path.exists(alt):
        return os.path.abspath(alt)
    return None


def _restore_cached_upload(resolved, key):
    """Backfilled deploy caches may have clone_run_files but an empty files map.
    Reconnect those file names to the baked uploads folder so endpoints that
    need the original workbooks can run without a manual re-upload.
    """
    if key in resolved:
        return

    cached_name = (state.get('clone_run_files') or {}).get(key)
    restored = _resolve_cached_path(cached_name)
    if restored:
        resolved[key] = restored


def load_cache():
    cf = app.config['CACHE_FILE']
    if not os.path.exists(cf):
        return
    try:
        with open(cf, 'rb') as f:
            data = pickle.load(f)
        state['clone_results']          = data.get('clone_results')
        state['weights']                = data.get('weights')
        state['client_benchmarks']      = data.get('client_benchmarks') or {}
        state['risk_data']              = data.get('risk_data')
        state['clone_run_files']        = data.get('clone_run_files') or {}
        state['security_risk_data']     = data.get('security_risk_data')
        state['universe_clone_results'] = data.get('universe_clone_results')
        state['exposures_data']         = data.get('exposures_data')
        state['norm_skill_by_tab']      = data.get('norm_skill_by_tab', {})
        state['placeholder_buckets']    = data.get('placeholder_buckets') or {}
        state['qualitative_data']       = data.get('qualitative_data')
        state['mc_universe_cache']      = data.get('mc_universe_cache') or {}
        cached_files                    = data.get('files', {})

        # Resolve each cached path. Drop entries whose files can no longer be
        # located; rewrite survivors to their new absolute paths. This keeps
        # /run from trying to open a ghost path and failing mysteriously.
        resolved = {}
        for key, val in (cached_files or {}).items():
            if isinstance(val, dict):
                # universe_returns is a {peer_tab: path} sub-dict
                sub = {}
                for pt, pp in val.items():
                    rp = _resolve_cached_path(pp)
                    if rp:
                        sub[pt] = rp
                if sub:
                    resolved[key] = sub
            else:
                rp = _resolve_cached_path(val)
                if rp:
                    resolved[key] = rp

        _restore_cached_upload(resolved, 'manager_returns')
        _restore_cached_upload(resolved, 'factor_returns')
        state['files'] = resolved

        # Reload manager_dfs if the manager returns file resolved successfully
        mgr_path = state['files'].get('manager_returns')
        if mgr_path and os.path.exists(mgr_path):
            state['manager_dfs'] = load_manager_returns(mgr_path)

        factor_path = state['files'].get('factor_returns')
        if factor_path and os.path.exists(factor_path):
            state['factor_df'] = load_factor_returns(factor_path)

        n_mgrs = sum(len(v) for v in (state['clone_results'] or {}).values())
        missing = [k for k in (cached_files or {}) if k not in resolved]
        tail = f" (dropped unresolved paths: {missing})" if missing else ""
        print(f"Cache loaded — {n_mgrs} managers.{tail}")
        # Roll any non-style residual into Core so buckets sum to 100%
        absorb_regional_into_core(state.get('clone_results'))
        absorb_regional_into_core(state.get('universe_clone_results'))
    except Exception as e:
        print(f"Cache load error (will recompute): {e}")

# Load cache on startup. First try to restore it from S3 (Fargate disks are
# ephemeral, so a fresh task has no local cache) — then load whatever is on disk.
_pull_cache_from_s3()
load_cache()

# The client roster (names/benchmarks/managers/weights) comes from Postgres
# when available, overriding whatever the pickle cache held. Everything else
# (clone results, risk, exposures) still comes from load_cache() above.
refresh_clients_from_db()

# Recomputing normalized skill during module import can keep Flask from
# binding its port for a long time when the cached clone panel is large.
# Use the cached values at startup and refresh them after clone runs, or
# lazily for a specific tab if a request needs a missing cache entry.
if state.get('clone_results'):
    cached_ns = state.get('norm_skill_by_tab') or {}
    cached_ns_count = sum(len(v or {}) for v in cached_ns.values())
    if cached_ns_count:
        print(f"Using cached normalized skill on startup — {cached_ns_count} managers.")
    else:
        print("Normalized skill cache empty on startup — it will refresh on demand.")

def cb(msg): state['progress'].append(msg)

# Progress bookkeeping helpers ────────────────────────────────────────────
import re as _re
_PROGRESS_LINE_RE = _re.compile(r'^\[(?:universe\s+)?(\d+)\s*/\s*(\d+)\]')

def reset_progress(phase='', total=0):
    """Clear the progress log and counters at the start of a new run."""
    state['progress']          = []
    state['progress_current']  = 0
    state['progress_total']    = int(total or 0)
    state['progress_phase']    = phase or ''
    state['progress_sub_current'] = 0
    state['progress_sub_total']   = 0

def set_progress_phase(phase, total=None, reset_current=False):
    """Update the phase label (and optionally total / current counter)."""
    state['progress_phase'] = phase or ''
    if total is not None:
        state['progress_total'] = int(total)
    if reset_current:
        state['progress_current'] = 0
    # Whenever the phase changes, clear sub-progress; the new phase is
    # responsible for repopulating it if it wants to advertise per-step
    # ticks (e.g. norm-skill snapshots).
    state['progress_sub_current'] = 0
    state['progress_sub_total']   = 0

def cb_counting(msg):
    """Like cb(), but also bumps progress_current whenever it sees a
    `[done/total]` or `[universe done/total]` marker produced by the clone
    engine. We parse `done` directly rather than simply incrementing on
    every call because the total is authoritative and we don't want
    ancillary "Loading..." messages to skew the counter.
    """
    state['progress'].append(msg)
    if isinstance(msg, str):
        m = _PROGRESS_LINE_RE.match(msg)
        if m:
            try:
                done_here = int(m.group(1))
                # The engine's `done` resets per tab in the consolidated
                # universe path, so never let progress_current go backwards.
                # We carry a per-phase offset so multi-tab runs aggregate.
                base = state.get('_progress_phase_base', 0)
                state['progress_current'] = max(
                    state.get('progress_current', 0),
                    base + done_here,
                )
            except Exception:
                pass

def advance_progress_base():
    """Snapshot the current counter as the new base so that a subsequent
    phase (e.g. the next peer tab) starts counting from where we left off
    instead of resetting to zero when the engine emits `[1/N]` again."""
    state['_progress_phase_base'] = state.get('progress_current', 0)


# ── Routes ─────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    """Service identity. The user-facing UI lives on the separate Next.js
    frontend (typically at :3000) — this Flask server only exposes JSON."""
    return jsonify({
        'service': 'aapryl-clone-tool-backend',
        'ui': 'see Next.js frontend (default http://localhost:3000)',
    })

@app.route('/upload', methods=['POST'])
def upload_files():
    saved = []
    warnings = {}   # {key: warning_dict} — surfaced to the UI
    for key in ['manager_returns', 'factor_returns', 'weights']:
        if key in request.files:
            f = request.files[key]
            if f.filename:
                path = save_uploaded_file(f, secure_filename(f.filename), app.config['UPLOAD_FOLDER'])
                state['files'][key] = path
                saved.append(key)
                # Sniff the factor-returns file for missing expected factors.
                # FACTOR_CATEGORIES enumerates every factor that any peer-tab
                # clone can pick from. Anything in that union but absent from
                # the uploaded file is silently dropped at clone time, which
                # changes which factors LASSO can select and shifts both
                # betas and R² versus the historical R workflow. Surfacing
                # this at upload time lets the user fix the FactSet export
                # before running a multi-minute clone.
                if key == 'factor_returns':
                    try:
                        from clone_engine import FACTOR_CATEGORIES
                        from data_loader import _resolve_factors_ci
                        fdf = load_factor_returns(path)
                        # Build a single union of all expected factors, then
                        # check which ones the file actually provides using
                        # the same case-insensitive matcher the clone engine
                        # uses. Anything not resolvable is genuinely missing.
                        expected = set()
                        for cat_list in FACTOR_CATEGORIES.values():
                            expected.update(cat_list)
                        resolved = _resolve_factors_ci(sorted(expected), fdf.columns)
                        # Map back: which expected names didn't resolve?
                        def _k(s): return ' '.join(str(s).split()).lower()
                        resolved_keys = {_k(c) for c in resolved}
                        missing = sorted(f for f in expected if _k(f) not in resolved_keys)
                        if missing:
                            # Group by which peer tabs need each missing factor.
                            # A factor can appear in both _full and _3factor for
                            # the same tab — de-dup so the UI doesn't show the
                            # same name twice.
                            tab_impact = {}
                            for cat_name, cat_list in FACTOR_CATEGORIES.items():
                                tab = cat_name.split('_')[0]   # 'ISC_full' → 'ISC'
                                for f_name in missing:
                                    if f_name in cat_list:
                                        tab_impact.setdefault(tab, set()).add(f_name)
                            tab_impact = {t: sorted(v) for t, v in tab_impact.items()}
                            warnings[key] = {
                                'type':         'missing_factors',
                                'missing':      missing,
                                'missing_count': len(missing),
                                'expected_count': len(expected),
                                'tab_impact':   tab_impact,
                            }
                    except Exception as e:
                        # A parse error here shouldn't block the upload; just
                        # log it and let the clone run surface the problem.
                        warnings[key] = {'type': 'sniff_error', 'message': str(e)}
                    # A new factor file invalidates the cached market-cycle
                    # universe state (built from factor exposures) and the
                    # loaded factor_df — both must rebuild on demand rather
                    # than serving stale results.
                    state['mc_universe_cache'] = {}
                    state['factor_df'] = None
    return jsonify({'status': 'ok', 'saved': saved, 'warnings': warnings})


@app.route('/upload_universe', methods=['POST'])
def upload_universe():
    """Stage a peer-universe return file on the server. Each file maps to
    ONE peer tab via the 'peer_tab' form field (EAFE, ACWI, ISC, EM, US,
    USSC). Cloning does NOT run automatically — the user must click
    "Run Universe Clones" to trigger it via /run_universe.
    """
    if 'universe_returns' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file provided'})
    f = request.files['universe_returns']
    if not f.filename:
        return jsonify({'status': 'error', 'message': 'Empty filename'})
    peer_tab = request.form.get('peer_tab', '').strip().upper()
    if peer_tab not in ['EAFE', 'ACWI', 'ISC', 'EM', 'US', 'USSC']:
        return jsonify({'status': 'error',
                        'message': f"peer_tab must be one of EAFE/ACWI/ISC/EM/US/USSC, got '{peer_tab}'"})
    if 'factor_returns' not in state['files']:
        return jsonify({'status': 'error',
                        'message': 'Upload factor_returns first — universe cloning needs it.'})

    safe_fn = secure_filename(f"universe_{peer_tab}_{f.filename}")
    path = save_uploaded_file(f, safe_fn, app.config['UPLOAD_FOLDER'])
    # Keep track of universe files per peer tab
    uni_map = state['files'].get('universe_returns', {})
    if not isinstance(uni_map, dict):
        uni_map = {}
    uni_map[peer_tab] = path
    state['files']['universe_returns'] = uni_map
    save_cache()
    return jsonify({'status': 'staged', 'peer_tab': peer_tab,
                    'message': f'{peer_tab} universe file staged — click Run Universe Clones to compute.'})


@app.route('/upload_universe_consolidated', methods=['POST'])
def upload_universe_consolidated():
    """Upload ONE consolidated universe returns file with all peer groups on
    separate sheets (US LC, US SC, ISC, EAFE, Global, EM). The file is
    staged only — cloning is NOT auto-triggered. The user must click
    "Run Universe Clones" to execute it via /run_universe.
    """
    if 'universe_returns' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file provided'})
    f = request.files['universe_returns']
    if not f.filename:
        return jsonify({'status': 'error', 'message': 'Empty filename'})
    if 'factor_returns' not in state['files']:
        return jsonify({'status': 'error',
                        'message': 'Upload factor_returns first — universe cloning needs it.'})

    safe_fn = secure_filename(f"universe_consolidated_{f.filename}")
    path = save_uploaded_file(f, safe_fn, app.config['UPLOAD_FOLDER'])

    # Store under a reserved __all__ key so we remember the source file
    uni_map = state['files'].get('universe_returns', {})
    if not isinstance(uni_map, dict):
        uni_map = {}
    uni_map['__all__'] = path
    state['files']['universe_returns'] = uni_map

    # Peek at the sheets so we can tell the user what we recognised, without
    # running any cloning yet. Cheap — just inspects sheet names.
    try:
        import pandas as pd
        xl = pd.ExcelFile(path, engine='openpyxl')
        from data_loader import _UNIVERSE_SHEET_TO_TAB
        recognised = sorted({_UNIVERSE_SHEET_TO_TAB[s.strip()]
                             for s in xl.sheet_names
                             if s.strip() in _UNIVERSE_SHEET_TO_TAB})
    except Exception:
        recognised = []

    save_cache()
    return jsonify({
        'status':    'staged',
        'recognised_tabs': recognised,
        'message':   ('Consolidated universe file staged — click '
                      'Run Universe Clones to compute.')
    })

def _clone_results_stale():
    """True if the current manager_returns or factor_returns file has a
    different basename to what was used for the last buy-list clone run."""
    if not state.get('clone_results'):
        return False   # no clones yet — nothing to be stale
    run_files = state.get('clone_run_files') or {}
    if not run_files:
        return False   # older cache without metadata — can't tell
    for key in ('manager_returns', 'factor_returns'):
        current = os.path.basename(state['files'].get(key, ''))
        used    = run_files.get(key, '')
        if current and used and current != used:
            return True
    return False


@app.route('/status')
def status():
    """Tell the UI whether cached results are already available."""
    ucr = state.get('universe_clone_results') or {}
    uni_tabs_cached = sorted([t for t, m in ucr.items() if m])
    uni_files = state.get('files', {}).get('universe_returns', {})
    if not isinstance(uni_files, dict):
        uni_files = {}
    uni_files_staged = sorted(uni_files.keys())
    return jsonify({
        'has_results':   state['clone_results'] is not None,
        'has_weights':   state['weights'] is not None,
        'has_risk':      state['risk_data'] is not None,
        'has_security_risk': state.get('security_risk_data') is not None,
        'has_universe':  len(uni_tabs_cached) > 0,
        'universe_tabs': uni_tabs_cached,
        'universe_files_staged': uni_files_staged,
        'has_exposures': state['exposures_data'] is not None,
        'exposures_benchmark': (state['exposures_data'] or {}).get('benchmark_name', ''),
        'exposures_managers':  (state['exposures_data'] or {}).get('manager_names', []),
        'has_qualitative': state.get('qualitative_data') is not None,
        'qualitative_firms':      (state.get('qualitative_data') or {}).get('n_firms', 0),
        'qualitative_strategies': (state.get('qualitative_data') or {}).get('n_strategies', 0),
        'files': {k: os.path.basename(v) if isinstance(v, str) else v
                  for k, v in state['files'].items()},
        # Staleness signal: True when the currently-uploaded manager_returns
        # or factor_returns file differs from what was used for the last clone
        # run. Prompts the user to re-run buy-list clones.
        'clone_stale': _clone_results_stale(),
        'clone_run_files': state.get('clone_run_files') or {},
    })

@app.route('/run', methods=['POST'])
def run():
    if state['running']:
        return jsonify({'status': 'already_running'})
    missing = [k for k in ['manager_returns', 'factor_returns'] if k not in state['files']]
    if missing:
        return jsonify({'status': 'error', 'message': f'Missing: {missing}'})
    reset_progress(phase='Loading input files', total=0)
    # Snapshot the prior run's results before we nuke state so the
    # determinism diagnostic can diff old-vs-new at the end. Cheap copy:
    # the diagnostic only inspects each manager's betas + _input_hash, so
    # we keep a reference rather than deep-copying the whole structure.
    prev_clone_results = state.get('clone_results') or {}
    state.update({'running': True, 'error': None, 'clone_results': None})

    def worker():
        try:
            # Pre-count managers so the progress bar is meaningful from the
            # first tick. Cheap: just opens the workbook and counts columns.
            try:
                from data_loader import load_manager_returns as _lmr
                _pre_dfs = _lmr(state['files']['manager_returns'])
                total_mgrs = sum(df.shape[1] for df in _pre_dfs.values())
                set_progress_phase('Cloning managers', total=total_mgrs, reset_current=True)
                state['_progress_phase_base'] = 0
            except Exception:
                # Non-fatal: bar will just run indeterminate if this fails.
                pass
            results = run_cloning(state['files']['manager_returns'],
                                  state['files']['factor_returns'],
                                  progress_callback=cb_counting)
            state['clone_results'] = results
            # Run-vs-run determinism diagnostic. Splits material beta
            # changes into 'inputs SAME' (algorithmic non-determinism, e.g.
            # threaded BLAS reordering FP ops and flipping a near-zero
            # LASSO coefficient in/out of the active set) vs 'inputs
            # CHANGED' (file drift from re-saved manager_returns or
            # factor_returns). Output goes to both the progress log (so
            # the user sees it in the UI) and stdout for grep-ability.
            try:
                diag_lines = diagnose_clone_determinism(prev_clone_results, results)
                for line in diag_lines:
                    state['progress'].append(line)
                    print(line)
            except Exception as e:
                print(f"Determinism diagnostic skipped: {e}")
            state['manager_dfs']   = load_manager_returns(state['files']['manager_returns'])
            if 'weights' in state['files']:
                w, b = load_weights(state['files']['weights'])
                # When the DB is authoritative, import the workbook into Postgres
                # (preserving drafts) and reload state from the DB instead of
                # letting Excel silently overwrite the DB-backed roster.
                if db_enabled():
                    sync_weights_to_state(w, b)
                else:
                    state['weights'], state['client_benchmarks'] = w, b
            # Cache factor returns for benchmark lookups
            from data_loader import load_factor_returns as lf
            state['factor_df'] = lf(state['files']['factor_returns'])
            set_progress_phase('Computing normalized skill')
            absorb_regional_into_core(state['clone_results'])
            recompute_norm_skill()
            set_progress_phase('Saving cache')
            save_cache()
            # Record which files were used so the UI can warn if newer files
            # have been uploaded without re-running clones.
            state['clone_run_files'] = {
                'manager_returns': os.path.basename(state['files'].get('manager_returns', '')),
                'factor_returns':  os.path.basename(state['files'].get('factor_returns', '')),
            }
            set_progress_phase('Done')
            if state.get('progress_total'):
                state['progress_current'] = state['progress_total']
            cb("DONE")
        except Exception as e:
            import traceback
            state['error'] = str(e) + "\n" + traceback.format_exc()
            cb(f"ERROR: {e}")
        finally:
            state['running'] = False

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({'status': 'started'})

@app.route('/reload_weights', methods=['POST'])
def reload_weights():
    """Re-read just the weights file without rerunning clones."""
    if 'weights' not in state['files']:
        return jsonify({'status': 'error', 'message': 'No weights file uploaded yet'})
    try:
        w, b = load_weights(state['files']['weights'])
        # DB authoritative → import workbook into Postgres (drafts preserved)
        # and reload from DB; otherwise fall back to the legacy in-memory set.
        if db_enabled():
            sync_weights_to_state(w, b)
        else:
            state['weights'], state['client_benchmarks'] = w, b
        save_cache()
        return jsonify({'status': 'ok', 'clients': list(state['weights'].keys()),
                        'client_benchmarks': state['client_benchmarks']})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/reload_inputs', methods=['POST'])
def reload_inputs():
    """Re-read the weights file, the FactSet Risk Summary, and the FactSet
    Exposures file from their cached paths without rerunning clones. Useful
    while iterating on any of those input files — edit the xlsx, hit
    'Reload Inputs' in the UI, and the tool refreshes without the 5-15 min
    universe clone run.

    Returns a status dict indicating which inputs were refreshed.
    """
    status = {'weights': 'skipped', 'risk': 'skipped', 'exposures': 'skipped',
              'qualitative': 'skipped'}
    errors = {}

    # Weights
    if 'weights' in state['files'] and os.path.exists(state['files']['weights']):
        try:
            w, b = load_weights(state['files']['weights'])
            if db_enabled():
                sync_weights_to_state(w, b)   # import into DB, drafts preserved
            else:
                state['weights'], state['client_benchmarks'] = w, b
            status['weights'] = 'ok'
        except Exception as e:
            status['weights'] = 'error'
            errors['weights'] = str(e)

    # Risk summary
    if 'risk_summary' in state['files'] and os.path.exists(state['files']['risk_summary']):
        try:
            state['risk_data'] = parse_risk_summary(state['files']['risk_summary'])
            status['risk'] = 'ok'
        except Exception as e:
            status['risk'] = 'error'
            errors['risk'] = str(e)

    # Exposures
    if 'exposures' in state['files'] and os.path.exists(state['files']['exposures']):
        try:
            from exposures_engine import parse_exposures_file
            state['exposures_data'] = parse_exposures_file(state['files']['exposures'])
            status['exposures'] = 'ok'
        except Exception as e:
            status['exposures'] = 'error'
            errors['exposures'] = str(e)

    # Qualitative firm/strategy data
    if 'qualitative' in state['files'] and os.path.exists(state['files']['qualitative']):
        try:
            from qualitative_loader import parse_qualitative_file
            state['qualitative_data'] = parse_qualitative_file(state['files']['qualitative'])
            _QUAL_MATCH_CACHE.clear()  # manager names may have changed
            status['qualitative'] = 'ok'
        except Exception as e:
            status['qualitative'] = 'error'
            errors['qualitative'] = str(e)

    save_cache()
    return jsonify({
        'status':     'ok',
        'refreshed':  status,
        'errors':     errors,
        'clients':    list((state.get('weights') or {}).keys()),
    })

@app.route('/clear_cache', methods=['POST'])
def clear_cache():
    """Wipe saved results so next run starts fresh."""
    cf = app.config['CACHE_FILE']
    if os.path.exists(cf):
        os.remove(cf)
    state.update({'clone_results': None, 'manager_dfs': None, 'weights': None, 'files': {}})
    return jsonify({'status': 'ok'})

@app.route('/progress')
def progress():
    # `done` means "the in-flight run finished" — i.e. we're no longer
    # running AND we have some clone_results on hand. The previous
    # definition (`clone_results is not None`) made the poller stop
    # immediately on universe runs, because clone_results is already
    # populated before a universe run begins.
    is_done = (not state['running']) and (state['clone_results'] is not None)
    total   = int(state.get('progress_total', 0) or 0)
    current = int(state.get('progress_current', 0) or 0)
    if total and current > total:
        current = total
    pct = (100.0 * current / total) if total > 0 else None
    sub_total   = int(state.get('progress_sub_total', 0) or 0)
    sub_current = int(state.get('progress_sub_current', 0) or 0)
    sub_pct = (100.0 * sub_current / sub_total) if sub_total > 0 else None
    return jsonify({
        'messages': state['progress'][-30:],
        'running':  state['running'],
        'done':     is_done,
        'error':    state['error'],
        'progress_current': current,
        'progress_total':   total,
        'progress_phase':   state.get('progress_phase', ''),
        'progress_pct':     pct,   # None when total unknown; UI falls back to indeterminate animation
        # Per-phase sub-progress (e.g. norm-skill snapshots). When
        # populated, the UI can show a more responsive percentage during
        # phases that don't tick the global counter.
        'progress_sub_current': sub_current,
        'progress_sub_total':   sub_total,
        'progress_sub_pct':     sub_pct,
    })

@app.route('/clients')
def clients():
    # Always report `editable` (even for an empty roster) so a reachable-but-
    # empty DB still shows the "+ Add Client" control and the first client can
    # be created.
    return jsonify({
        'clients':    list((state.get('weights') or {}).keys()),
        'benchmarks': state.get('client_benchmarks') or {},
        'editable':   db_enabled(),
    })


@app.route('/benchmarks')
def benchmarks_catalog():
    """Canonical benchmark labels for the client add/rename dropdown.

    A strict, backend-owned list so the UI can't submit a typo'd benchmark —
    every value here resolves through resolve_peer_group() for the risk math.
    """
    return jsonify({'benchmarks': CANONICAL_BENCHMARKS})


# ── Client management (CRUD) ─────────────────────────────────────────────────
# These require the Postgres client DB. Each mutates the DB, then reloads the
# roster into state so every other endpoint immediately sees the change.
def _require_db():
    if not db_enabled():
        return jsonify({
            'error': 'Client management requires the database. Start it with '
                     '`docker compose up -d`.'
        }), 503
    return None


def _clients_payload():
    return {
        'clients':    list((state.get('weights') or {}).keys()),
        'benchmarks': state.get('client_benchmarks') or {},
        'editable':   True,
    }


def _validate_benchmark(bench):
    """Return a cleaned benchmark, or raise ValueError only for values the risk
    engine genuinely can't use. None/empty is allowed (benchmark is optional).

    Accepts a value if EITHER it's one of the canonical dropdown labels OR it
    resolves to a real peer group via resolve_peer_group(). This is deliberately
    more permissive than the dropdown: the weights workbook carries real spelling
    variants ('MSCI EAFE+CANADA', 'MSCI ACWI ex US') that aren't canonical
    strings but DO resolve to a peer group, so renaming such a client in the UI
    must not reject its existing benchmark. Only truly unrecognizable strings
    (that resolve to nothing) are rejected."""
    if bench is None:
        return None
    b = str(bench).strip()
    if not b or b.lower() == 'none':
        return None
    if b in CANONICAL_BENCHMARKS:
        return b
    from data_loader import resolve_peer_group
    if resolve_peer_group(b) is not None:
        return b   # non-canonical spelling, but resolvable — keep it verbatim
    raise ValueError(
        f"Unrecognized benchmark {b!r}. Pick one of the supported benchmarks."
    )


def _db_error_response(action):
    """Translate the current exception into a client-safe response.

    Connection-level failures (DB went down mid-request) → 503 with the
    'database required' message. Everything else → a generic 500 with the full
    traceback logged server-side (never leak SQL/schema to the client)."""
    from sqlalchemy.exc import OperationalError, InterfaceError, DBAPIError
    exc = sys.exc_info()[1]
    if isinstance(exc, (OperationalError, InterfaceError)) or (
        isinstance(exc, DBAPIError) and getattr(exc, 'connection_invalidated', False)
    ):
        # Force a re-probe next call so a recovered DB is picked up.
        try:
            client_db.session._enabled_positive = False  # type: ignore[attr-defined]
        except Exception:  # noqa: BLE001
            pass
        return jsonify({
            'error': 'The database is temporarily unavailable. Please try again.'
        }), 503
    app.logger.exception("%s failed", action)
    return jsonify({'error': f'{action} failed. Please try again.'}), 500


@app.route('/clients', methods=['POST'])
def create_client():
    guard = _require_db()
    if guard:
        return guard
    body = request.get_json(silent=True) or {}
    name = (body.get('name') or '').strip()
    managers = body.get('managers') or []   # optional [{manager_name, weight}]
    try:
        benchmark = _validate_benchmark(body.get('benchmark'))
        client_db.repository.create_client(name, benchmark, managers)
        refresh_clients_from_db()
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception:  # noqa: BLE001
        return _db_error_response('Create client')
    # Echo the normalized stored name so the UI selects the right client.
    created = next((c for c in (state.get('weights') or {}) if c.lower() == name.lower()), name)
    return jsonify({'ok': True, 'created': created, **_clients_payload()})


@app.route('/clients/<path:client_name>', methods=['PATCH'])
def update_client(client_name):
    guard = _require_db()
    if guard:
        return guard
    body = request.get_json(silent=True) or {}
    new_name = body.get('name')
    benchmark_provided = 'benchmark' in body
    try:
        benchmark = _validate_benchmark(body.get('benchmark')) if benchmark_provided else None
        client_db.repository.update_client(
            client_name,
            new_name=new_name,
            benchmark=benchmark,
            benchmark_provided=benchmark_provided,
        )
        refresh_clients_from_db()
    except ValueError as e:
        msg = str(e)
        code = 404 if 'not found' in msg.lower() else 400
        return jsonify({'error': msg}), code
    except Exception:  # noqa: BLE001
        return _db_error_response('Update client')
    # Report the normalized new name if renamed, else the original.
    renamed = client_name
    if new_name:
        norm = ' '.join(new_name.split())
        renamed = next((c for c in (state.get('weights') or {}) if c.lower() == norm.lower()), norm)
    return jsonify({'ok': True, 'renamed_to': renamed, **_clients_payload()})


@app.route('/clients/<path:client_name>', methods=['DELETE'])
def delete_client(client_name):
    guard = _require_db()
    if guard:
        return guard
    try:
        client_db.repository.delete_client(client_name)
        refresh_clients_from_db()
    except ValueError as e:
        return jsonify({'error': str(e)}), 404
    except Exception:  # noqa: BLE001
        return _db_error_response('Delete client')
    return jsonify({'ok': True, 'deleted': client_name, **_clients_payload()})


@app.route('/clients/<path:client_name>/portfolio', methods=['PUT'])
def save_client_portfolio(client_name):
    """Persist a client's full edited portfolio (the 'Save' button).

    Body JSON: { managers: [ { matched_name|manager_name, tab?, current_weight,
                               proposed_weight, style_buckets? }, ... ] }

    Saves the manager list, both weight books, and any placeholder style
    buckets, then refreshes app state so the change is immediately live.
    """
    guard = _require_db()
    if guard:
        return guard
    body = request.get_json(silent=True) or {}
    incoming = body.get('managers')
    if not isinstance(incoming, list):
        return jsonify({'error': 'Body must include a "managers" list.'}), 400

    # Normalize each manager to the repository's expected shape.
    #
    # DB row identity is the STABLE weight-file name (the original workbook
    # name), NOT matched_name. matched_name is the fuzzy-matched clone key used
    # only for analytics lookups and can differ from the workbook name by
    # whitespace/casing/token-sort — persisting it would silently rename the
    # manager on every save. Fall back to matched_name only when no separate
    # weight_file_name was supplied (e.g. a manager added mid-session).
    rows = []
    seen_names = set()
    for m in incoming:
        name = (m.get('weight_file_name')
                or m.get('manager_name')
                or m.get('matched_name')
                or '').strip()
        if not name:
            continue
        if name in seen_names:
            return jsonify({
                'error': f'Duplicate manager in save payload: {name!r}.'
            }), 400
        seen_names.add(name)
        try:
            cw = _validate_weight(m.get('current_weight'), 'current_weight', name)
            pw = _validate_weight(m.get('proposed_weight'), 'proposed_weight', name)
            buckets = _validate_buckets(m.get('style_buckets'), name)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        # Only persist style buckets for PLACEHOLDER managers (no clone data).
        # Real clone managers get their buckets recomputed each run; storing a
        # snapshot here would later shadow fresh clone output as a stale
        # "placeholder override". Frontend flags placeholders via is_placeholder.
        if not m.get('is_placeholder'):
            buckets = None
        rows.append({
            'manager_name':    name,
            'current_weight':  cw,
            'proposed_weight': pw,
            'style_buckets':   buckets,
        })

    try:
        client_db.repository.save_client_portfolio(client_name, rows)
        refreshed = refresh_clients_from_db()   # reloads state['weights']
        _sync_placeholder_buckets_from_db()      # reloads buckets into state
    except ValueError as e:
        msg = str(e)
        code = 404 if 'not found' in msg.lower() else 400
        return jsonify({'error': msg}), code
    except Exception:  # noqa: BLE001
        return _db_error_response('Save portfolio')

    # If the DB write committed but the in-memory refresh failed, don't persist
    # the (now stale) cache — surface a soft warning instead of silently
    # baking stale weights to disk.
    if not refreshed:
        return jsonify({
            'ok': True, 'saved': client_name,
            'warning': 'Saved to the database, but the in-memory view may be '
                       'stale until reload.',
        })

    save_cache()  # persist the placeholder-bucket change across restarts too
    return jsonify({'ok': True, 'saved': client_name})


# ── Save-payload validation helpers ─────────────────────────────────────────
def _validate_weight(value, field, manager):
    """Coerce a weight to a finite float >= 0, or raise ValueError."""
    if value is None:
        return 0.0
    try:
        f = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field} for {manager!r} is not a number.")
    if not math.isfinite(f):
        raise ValueError(f"{field} for {manager!r} must be a finite number.")
    if f < 0:
        raise ValueError(f"{field} for {manager!r} cannot be negative.")
    return f


def _validate_buckets(buckets, manager):
    """Coerce style buckets to {str: finite float > 0} or None; raise on junk."""
    if not buckets:
        return None
    if not isinstance(buckets, dict):
        raise ValueError(f"style_buckets for {manager!r} must be an object.")
    clean = {}
    for k, v in buckets.items():
        try:
            f = float(v)
        except (TypeError, ValueError):
            raise ValueError(f"style bucket {k!r} for {manager!r} is not a number.")
        if not math.isfinite(f):
            raise ValueError(f"style bucket {k!r} for {manager!r} must be finite.")
        if f > 0:
            clean[str(k)] = f
    return clean or None


def _normalize_buckets(buckets):
    """Normalize a {bucket: weight} dict to sum=1.0 (drop non-positive).

    Single source of truth for bucket normalization, used by BOTH the DB sync
    and the /placeholder_buckets endpoint so the same manager's buckets are
    identical regardless of which path last wrote them.
    """
    if not buckets:
        return None
    clean = {}
    for k, v in buckets.items():
        try:
            f = float(v)
        except (TypeError, ValueError):
            continue
        if f > 0:
            clean[str(k)] = f
    total = sum(clean.values())
    if total <= 0:
        return None
    return {k: v / total for k, v in clean.items()}


def _sync_placeholder_buckets_from_db():
    """Rebuild state['placeholder_buckets'] from every client's saved buckets.

    Placeholder buckets are global-by-manager-name in this app (build_portfolio_
    view looks them up by name across all clients), matching the legacy
    /placeholder_buckets contract. We union all clients' saved buckets and
    NORMALIZE each to sum=1.0 so values match what /placeholder_buckets writes,
    regardless of which path last saved them. No-op when the DB is unavailable.
    """
    if not db_enabled():
        return
    try:
        merged = {}
        weights, _ = client_db.load_weights_from_db()
        for cname in weights:
            for row in (client_db.repository.load_client_portfolio(cname) or []):
                nb = _normalize_buckets(row.get('style_buckets'))
                if nb:
                    merged[row['manager_name']] = nb
        state['placeholder_buckets'] = merged
    except Exception as e:  # noqa: BLE001
        print(f"[db] failed to sync placeholder buckets: {e}")


def _ensure_norm_skill_for_tab(tab):
    """Populate normalized skill for one tab only when the cache is missing.

    This keeps startup fast while preserving skill data for tabs that were
    absent from older cache files.
    """
    if not tab or state.get('running') or not state.get('clone_results'):
        return
    ns_by_tab = state.get('norm_skill_by_tab') or {}
    if ns_by_tab.get(tab):
        return
    try:
        print(f"Lazy normalized skill recompute for {tab}...")
        recompute_norm_skill(tabs=[tab])
        save_cache()
    except Exception as e:
        print(f"Lazy norm-skill compute skipped for {tab}: {e}")

def _norm_skill_for(tab, name):
    """Return normalized-skill fields for a manager.

    Uses an exact-key lookup first, then a whitespace-stripped / lower-cased
    fallback so minor name variants (e.g. trailing space from Excel) never
    cause a silent miss.
    """
    _ensure_norm_skill_for_tab(tab)
    ns_tab = (state.get('norm_skill_by_tab') or {}).get(tab, {})
    # 1 — exact match
    rec = ns_tab.get(name)
    # 2 — whitespace-normalised fallback
    if rec is None:
        key_norm = str(name or '').strip().lower()
        for k, v in ns_tab.items():
            if str(k or '').strip().lower() == key_norm:
                rec = v
                break
    if not rec:
        return {'ns_z': None, 'ns_skill': None, 'ns_n_obs': None,
                'ns_n_peers': None, 'ns_adj_method': None,
                'ns_last_month': None}
    return {
        'ns_z':          rec.get('z'),
        'ns_skill':      rec.get('skill'),
        'ns_n_obs':      rec.get('n_obs'),
        'ns_n_peers':    rec.get('n_peers'),
        'ns_adj_method': rec.get('adj_method'),
        'ns_last_month': rec.get('last_month'),
    }


def _resolve_manager_record(tab, mgr_name):
    """Return (resolved_name, clone_result_dict) for a manager lookup."""
    cr_tab = (state.get('clone_results') or {}).get(tab, {})
    if not isinstance(cr_tab, dict):
        return None, None
    d = cr_tab.get(mgr_name)
    exact_key = mgr_name
    if d is None:
        key_norm = str(mgr_name or '').strip().lower()
        for k, v in cr_tab.items():
            if str(k or '').strip().lower() == key_norm:
                d = v
                exact_key = k
                break
    return exact_key, d


def _safe_metric(value):
    """Return a finite float or None for downstream scoring."""
    try:
        f = float(value)
        if np.isnan(f) or np.isinf(f):
            return None
        return f
    except (TypeError, ValueError):
        return None


def _filled_manager_returns(manager_result, max_months=36):
    """Return recent manager returns, back-filling holes with the static clone."""
    mgr = manager_result.get('manager_returns') or []
    clone = manager_result.get('static_clone_full') or []
    out = []
    for idx in range(max(len(mgr), len(clone))):
        mv = _safe_metric(mgr[idx]) if idx < len(mgr) else None
        if mv is None and idx < len(clone):
            mv = _safe_metric(clone[idx])
        if mv is None:
            continue
        out.append(mv)
        if len(out) >= max_months:
            break
    return out


def _annualized_volatility(returns):
    if len(returns) < 12:
        return None
    arr = np.array(returns, dtype=float)
    return float(np.std(arr, ddof=0) * np.sqrt(12.0))


def _annualized_downside_dev(returns):
    if len(returns) < 12:
        return None
    arr = np.minimum(np.array(returns, dtype=float), 0.0)
    return float(np.sqrt(np.mean(arr ** 2)) * np.sqrt(12.0))


def _build_portfolio_contribution_rows(managers):
    """Compute contribution rows for an arbitrary manager list."""
    import math, pandas as pd
    from data_loader import get_benchmark, load_factor_returns

    if not state['clone_results']:
        return {'error': 'No results'}

    if state.get('factor_df') is None and 'factor_returns' in state['files']:
        state['factor_df'] = load_factor_returns(state['files']['factor_returns'])

    fdf = state.get('factor_df')

    def period_stats(mgr, clone, bench, n):
        pairs = [(mgr[i], clone[i], bench[i])
                 for i in range(min(n, len(mgr), len(clone), len(bench)))
                 if mgr[i] is not None and clone[i] is not None and bench[i] is not None
                 and not math.isnan(float(mgr[i])) and not math.isnan(float(clone[i]))
                 and not math.isnan(float(bench[i]))]
        if len(pairs) < max(1, n // 2):
            return None, None, None, None
        m_tot = float(np.prod([1 + p[0] for p in pairs]) - 1)
        c_tot = float(np.prod([1 + p[1] for p in pairs]) - 1)
        b_tot = float(np.prod([1 + p[2] for p in pairs]) - 1)
        nm = len(pairs)
        if nm >= 12:
            ann = lambda r: (1 + r) ** (12 / nm) - 1
            m_tot, c_tot, b_tot = ann(m_tot), ann(c_tot), ann(b_tot)
        return m_tot, m_tot - b_tot, c_tot - b_tot, m_tot - c_tot

    managers_out = []
    for m in managers:
        tab = m.get('tab')
        mgr_name = m.get('matched_name') or m.get('name')
        if not tab or not mgr_name:
            continue

        d = (state['clone_results'].get(tab, {}) or {}).get(mgr_name, {})
        if not d:
            continue

        mgr_rets = d.get('manager_returns', [])
        clone_rets = d.get('static_clone_full', [])
        dates = d.get('dates', [])

        bench_name = get_benchmark(tab, mgr_name)
        bench_rets = []
        if fdf is not None and bench_name in fdf.columns:
            for ds in dates:
                try:
                    dt = pd.Timestamp(ds)
                    v = fdf.loc[dt, bench_name] if dt in fdf.index else None
                    bench_rets.append(None if (v is None or math.isnan(float(v))) else float(v))
                except Exception:
                    bench_rets.append(None)
        else:
            bench_rets = [None] * len(dates)

        row = {
            'name': mgr_name,
            'weight': float(m.get('current_weight', 0) or 0),
            'benchmark_name': bench_name,
            'vg_full': m.get('vg_full', d.get('vg_full', 0)),
        }
        for label, n in [('qtd', 3), ('t1', 12), ('t3', 36)]:
            mr, be, ps, sk = period_stats(mgr_rets, clone_rets, bench_rets, n)
            row[f'{label}_mgr'] = round(mr * 100, 4) if mr is not None else None
            row[f'{label}_bench'] = round(be * 100, 4) if be is not None else None
            row[f'{label}_style'] = round(ps * 100, 4) if ps is not None else None
            row[f'{label}_skill'] = round(sk * 100, 4) if sk is not None else None

        managers_out.append(row)

    return {'managers': managers_out, 'unmatched': []}


def _manager_recommender_features(tab, name, manager_result):
    """Build one comparable feature bundle for manager-to-manager ranking."""
    style_buckets = {
        bucket: float((manager_result.get('style_buckets') or {}).get(bucket, 0) or 0)
        for bucket in _STYLE_BUCKET_KEYS
    }
    recent = _filled_manager_returns(manager_result, max_months=36)
    record = {
        'name': name,
        'tab': tab,
        'style_buckets': style_buckets,
        'vg_full': _safe_metric(manager_result.get('vg_full')) or 0.0,
        'pct_small': _safe_metric(manager_result.get('pct_small')) or 0.0,
        'pct_em': _safe_metric(manager_result.get('pct_em')) or 0.0,
        'r2_full': _safe_metric(manager_result.get('r2_full')),
        'vol_36m': _annualized_volatility(recent),
        'downside_dev_36m': _annualized_downside_dev(recent),
    }
    record.update(_norm_skill_for(tab, name))
    return record


def _bounded_similarity(gap, scale, default=0.5):
    if gap is None:
        return default
    if scale <= 0:
        return default
    return max(0.0, 1.0 - min(abs(gap) / scale, 1.0))


def _relative_similarity(left, right, floor=0.02, default=0.5):
    if left is None or right is None:
        return default
    denom = max(abs(left), abs(right), floor)
    return max(0.0, 1.0 - min(abs(left - right) / denom, 1.0))


def _manager_recommendation_payload(reference, candidate):
    """Rank a candidate versus the reference manager using current clone metrics."""
    style_gap = sum(
        abs(reference['style_buckets'].get(bucket, 0.0) - candidate['style_buckets'].get(bucket, 0.0))
        for bucket in _STYLE_BUCKET_KEYS
    ) / 2.0
    style_similarity = max(0.0, 1.0 - min(style_gap, 1.0))
    vg_gap = candidate['vg_full'] - reference['vg_full']
    small_gap = candidate['pct_small'] - reference['pct_small']
    em_gap = candidate['pct_em'] - reference['pct_em']
    vol_gap = None if reference['vol_36m'] is None or candidate['vol_36m'] is None else candidate['vol_36m'] - reference['vol_36m']
    downside_gap = None if reference['downside_dev_36m'] is None or candidate['downside_dev_36m'] is None else candidate['downside_dev_36m'] - reference['downside_dev_36m']
    ns_delta = None if reference['ns_z'] is None or candidate['ns_z'] is None else candidate['ns_z'] - reference['ns_z']
    r2_delta = None if reference['r2_full'] is None or candidate['r2_full'] is None else candidate['r2_full'] - reference['r2_full']

    vg_similarity = _bounded_similarity(vg_gap, 0.75)
    small_similarity = _bounded_similarity(small_gap, 0.35)
    em_similarity = _bounded_similarity(em_gap, 0.35)
    risk_similarity = (
        _relative_similarity(reference['vol_36m'], candidate['vol_36m']) +
        _relative_similarity(reference['downside_dev_36m'], candidate['downside_dev_36m'])
    ) / 2.0
    footprint_similarity = (
        0.6 * style_similarity +
        0.2 * vg_similarity +
        0.1 * small_similarity +
        0.1 * em_similarity
    )
    skill_component = 0.5 if ns_delta is None else max(0.0, min(1.0, 0.5 + (ns_delta / 2.0)))
    overall_score = 100.0 * (
        0.6 * footprint_similarity +
        0.25 * risk_similarity +
        0.15 * skill_component
    )

    reasons = []
    if footprint_similarity >= 0.88:
        reasons.append('Very similar style footprint')
    elif footprint_similarity >= 0.76:
        reasons.append('Close style footprint')
    if ns_delta is not None and ns_delta >= 0.2:
        reasons.append('Higher normalized skill')
    if vol_gap is not None and vol_gap <= -0.01:
        reasons.append('Lower trailing volatility')
    if downside_gap is not None and downside_gap <= -0.01:
        reasons.append('Lower downside deviation')
    if not reasons:
        reasons.append('Closest overall match in this peer tab')

    return {
        'name': candidate['name'],
        'tab': candidate['tab'],
        'score': round(overall_score, 1),
        'footprint_similarity': round(footprint_similarity * 100, 1),
        'risk_similarity': round(risk_similarity * 100, 1),
        'vg_full': candidate['vg_full'],
        'vg_gap': round(vg_gap, 4),
        'pct_small': candidate['pct_small'],
        'pct_small_gap': round(small_gap, 4),
        'pct_em': candidate['pct_em'],
        'pct_em_gap': round(em_gap, 4),
        'r2_full': candidate['r2_full'],
        'r2_delta': None if r2_delta is None else round(r2_delta, 4),
        'ns_z': candidate['ns_z'],
        'ns_z_delta': None if ns_delta is None else round(ns_delta, 4),
        'vol_36m': candidate['vol_36m'],
        'vol_delta': None if vol_gap is None else round(vol_gap, 4),
        'downside_dev_36m': candidate['downside_dev_36m'],
        'downside_delta': None if downside_gap is None else round(downside_gap, 4),
        'reasons': reasons,
    }


@app.route('/manager_recommendations/<tab>/<path:mgr_name>')
def manager_recommendations(tab, mgr_name):
    """Return same-tab manager alternatives ranked by footprint, risk, and skill."""
    if not state.get('clone_results'):
        return jsonify({'error': 'No results'})

    exact_key, base_record = _resolve_manager_record(tab, mgr_name)
    if not base_record:
        return jsonify({'error': f'{tab}/{mgr_name} not found'})

    reference = _manager_recommender_features(tab, exact_key, base_record)
    peer_tab = (state.get('clone_results') or {}).get(tab, {}) or {}
    candidates = []
    for cand_name, cand_record in peer_tab.items():
        if cand_name == exact_key:
            continue
        features = _manager_recommender_features(tab, cand_name, cand_record)
        candidates.append(_manager_recommendation_payload(reference, features))

    closest_matches = sorted(
        candidates,
        key=lambda row: (-row['score'], -(row['footprint_similarity']), row['name'].lower()),
    )[:5]
    skill_upgrades = sorted(
        [
            row for row in candidates
            if row['ns_z_delta'] is not None and row['ns_z_delta'] >= 0.15 and row['footprint_similarity'] >= 72.0
        ],
        key=lambda row: (
            -(0.65 * row['footprint_similarity'] + 35.0 * (row['ns_z_delta'] or 0.0)),
            row['name'].lower(),
        ),
    )[:3]
    lower_risk_matches = sorted(
        [
            row for row in candidates
            if row['footprint_similarity'] >= 72.0 and (
                (row['vol_delta'] is not None and row['vol_delta'] <= -0.01) or
                (row['downside_delta'] is not None and row['downside_delta'] <= -0.01)
            )
        ],
        key=lambda row: (
            -(0.55 * row['footprint_similarity'] + 0.45 * row['risk_similarity']),
            row['name'].lower(),
        ),
    )[:3]

    return jsonify({
        'reference': reference,
        'closest_matches': closest_matches,
        'skill_upgrades': skill_upgrades,
        'lower_risk_matches': lower_risk_matches,
        'scope': 'same_tab_only',
    })


@app.route('/portfolio/<client_name>')
def portfolio(client_name):
    if not state['clone_results']:
        return jsonify({'error': 'No results'})
    if not state['weights'] or client_name not in state['weights']:
        return jsonify({'error': 'Client not found'})
    data = build_portfolio_view(client_name, state['weights'][client_name],
                                state['clone_results'], state['manager_dfs'],
                                universe_clone_results=state.get('universe_clone_results'),
                                placeholder_buckets=state.get('placeholder_buckets') or {})
    # Enrich each manager with normalized-skill fields (placeholders skip
    # this — they have no clone, no skill Z, and the lookup would just
    # return None across the board)
    for m in data.get('managers', []):
        if m.get('is_placeholder'):
            # Placeholders still get qualitative fields (they may be a known
            # firm with FactSet data but no clone), just not norm-skill.
            m.update(_qual_fields(m.get('matched_name'), m.get('weight_file_name')))
            continue
        m.update(_norm_skill_for(m['tab'], m['matched_name']))
        m.update(_qual_fields(m['matched_name'], m.get('weight_file_name')))
    # Restore any saved proposed weights from the DB so an edited draft
    # survives a refresh. current_weight comes from state['weights'] (the
    # book the pipeline uses); proposed_weight is a what-if overlay stored
    # only in the DB, so we apply it here per matched manager name.
    if db_enabled():
        try:
            saved = client_db.repository.load_client_portfolio(client_name) or []
            # Key by the STABLE weight-file name (same identity used on save),
            # NOT matched_name. proposed_weight is nullable: NULL means "no
            # draft" so we skip it and keep build_portfolio_view's sensible
            # proposed==current default.
            proposed_by_name = {
                r['manager_name']: r.get('proposed_weight')
                for r in saved
                if r.get('proposed_weight') is not None
            }
            for m in data.get('managers', []):
                key = m.get('weight_file_name') or m.get('matched_name')
                pw = proposed_by_name.get(key)
                if pw is not None:
                    m['proposed_weight'] = pw
        except Exception as e:  # noqa: BLE001
            print(f"[db] could not apply saved proposed weights: {e}")
    # Surface the client's benchmark (as given in the weights file) so the
    # UI can label the risk/exposures tables with the right benchmark name.
    data['client_benchmark'] = (state.get('client_benchmarks') or {}).get(client_name)
    return jsonify(data)


@app.route('/placeholder_buckets', methods=['POST'])
def update_placeholder_buckets():
    """Update style buckets for one placeholder manager (or clear them).
    POST JSON: { name: str, buckets: {bucket: weight, ...} | null }
    Setting buckets to null/{} resets to the default Core=1.0.
    Bucket weights are normalised to sum to 1.0; the UI sends fractions
    or percentages — both work.
    """
    data    = request.get_json(silent=True) or {}
    name    = data.get('name')
    buckets = data.get('buckets')
    if not name:
        return jsonify({'status': 'error', 'message': 'Missing name'})
    if 'placeholder_buckets' not in state or state['placeholder_buckets'] is None:
        state['placeholder_buckets'] = {}
    if not buckets:
        state['placeholder_buckets'].pop(name, None)
    else:
        # Normalise to sum=1.0 via the shared helper so this path and the DB
        # sync path produce identical values for the same manager.
        state['placeholder_buckets'][name] = _normalize_buckets(buckets) or {}
    save_cache()
    return jsonify({'status': 'ok',
                    'buckets': state['placeholder_buckets'].get(name)})


def _enumerate_placeholder_candidates():
    """Return the set of manager names that should be eligible for the
    Placeholder peer group: any name appearing in the FactSet exposures
    file, the Risk Summary file, the Security-Level Risk file, or any
    client's weights file, that does NOT fuzzy-match into clone_results
    or universe_clone_results. These are managers the user has FactSet
    data for but no clone (typically <3 years of returns), so they can
    be added to a portfolio with manually-set style buckets."""
    from data_loader import fuzzy_match
    candidates = set()
    ed = state.get('exposures_data') or {}
    for n in (ed.get('managers') or {}).keys():
        candidates.add(n)
    rd = state.get('risk_data') or {}
    for n in rd.get('manager_names', []) or []:
        candidates.add(n)
    srd = state.get('security_risk_data') or {}
    for n in (srd.get('managers') or {}).keys():
        candidates.add(n)
    weights = state.get('weights') or {}
    for mgr_weights in weights.values():
        for n in mgr_weights.keys():
            candidates.add(n)
    mgr_dfs = state.get('manager_dfs') or {}
    placeholders = set()
    for name in candidates:
        t, m = fuzzy_match(name, mgr_dfs)
        cr_tab = (state.get('clone_results') or {}).get(t, {}) if t else {}
        ucr_tab = (state.get('universe_clone_results') or {}).get(t, {}) if t else {}
        if not (m and (m in cr_tab or m in ucr_tab)):
            placeholders.add(name)
    return placeholders


@app.route('/all_managers')
def all_managers():
    out = []
    if state.get('clone_results'):
        for tab, td in state['clone_results'].items():
            for name, d in td.items():
                row = {'name': name, 'tab': tab, 'peer_group': tab,
                       'vg_full': d.get('vg_full', 0), 'vg_3factor': d.get('vg_3factor', 0),
                       'r2_full': d.get('r2_full'),
                       'is_placeholder': False}
                row.update(_norm_skill_for(tab, name))
                out.append(row)
    # Placeholder candidates: managers with FactSet data but no clone.
    # Bucket overrides come from state['placeholder_buckets'] (default
    # Core=1.0). Surfacing these in /all_managers makes them selectable
    # from the Add Manager dialog so the user can pull e.g. 'IMC Global'
    # into a portfolio even though it has no clone yet.
    ph_buckets = state.get('placeholder_buckets') or {}
    for name in sorted(_enumerate_placeholder_candidates(), key=str.lower):
        sb = dict(ph_buckets.get(name) or {'Core': 1.0})
        v = sum(w for b, w in sb.items() if b == 'Value')
        y = sum(w for b, w in sb.items() if b == 'Yield')
        g = sum(w for b, w in sb.items() if b == 'Growth')
        vg = round((v + y) - g, 6)
        out.append({
            'name':           name,
            'tab':            'Placeholder',
            'peer_group':     'Placeholder',
            'vg_full':        vg,
            'vg_3factor':     vg,
            'r2_full':        None,
            'style_buckets':  sb,
            'is_placeholder': True,
        })
    return jsonify({'managers': out})

@app.route('/manager_detail/<tab>/<path:mgr_name>')
def manager_detail(tab, mgr_name):
    if not state['clone_results']:
        return jsonify({'error': 'No results'})
    exact_key, d = _resolve_manager_record(tab, mgr_name)
    if not d:
        return jsonify({'error': f'{tab}/{mgr_name} not found'})
    out = dict(d)
    out.update(_norm_skill_for(tab, exact_key))
    return jsonify(out)

@app.route('/peer_group_view/<tab>')
def peer_group_view(tab):
    if not state['clone_results']:
        return jsonify({'error': 'No results'})

    # Special peer group: synthesise placeholders from every weights-file
    # manager that doesn't fuzzy-match into clone_results / universe_clone_results.
    # Bucket overrides come from state['placeholder_buckets'] (default Core=1.0).
    # (Ported from clone_tool — the sibling /peer_skill_summary kept this; this
    # endpoint had accidentally dropped it, leaving the Placeholder table empty.)
    if tab == 'Placeholder':
        ph_buckets = state.get('placeholder_buckets') or {}
        out = []
        for name in sorted(_enumerate_placeholder_candidates(), key=str.lower):
            sb = dict(ph_buckets.get(name) or {'Core': 1.0})
            v = sum(w for b, w in sb.items() if b == 'Value')
            y = sum(w for b, w in sb.items() if b == 'Yield')
            g = sum(w for b, w in sb.items() if b == 'Growth')
            vg = round((v + y) - g, 6)
            out.append({
                'name':            name,
                'r2_full':         None,
                'r2_3factor':      None,
                'vg_full':         vg,
                'vg_3factor':      vg,
                'style_buckets':   sb,
                'pct_small':       sb.get('Small Cap', 0) or 0,
                'pct_em':          0,
                'betas_full':      {},
                'betas_3factor':   {},
                'dates':           [],
                'manager_returns': [],
                'static_clone_full': [],
                'is_placeholder':  True,
            })
        return jsonify({'tab': tab, 'managers': out})

    td = state['clone_results'].get(tab, {})
    out = []
    for name, d in td.items():
        row = {
            'name': name, 'r2_full': d.get('r2_full'), 'r2_3factor': d.get('r2_3factor'),
            'vg_full': d.get('vg_full', 0), 'vg_3factor': d.get('vg_3factor', 0),
            'style_buckets': d.get('style_buckets', {}),
            'pct_small': d.get('pct_small', 0), 'pct_em': d.get('pct_em', 0),
            'betas_full': d.get('betas_full', {}), 'betas_3factor': d.get('betas_3factor', {}),
            'dates': d.get('dates', [])[:24],
            'manager_returns': d.get('manager_returns', [])[:24],
            'static_clone_full': d.get('static_clone_full', [])[:24],
        }
        row.update(_norm_skill_for(tab, name))
        out.append(row)
    return jsonify({'tab': tab, 'managers': out})

@app.route('/compute_portfolio_stats', methods=['POST'])
def compute_portfolio_stats():
    managers = request.json.get('managers', [])
    def agg(wkey):
        keys = ['vg_full','vg_3factor','Core','Growth','Value','Yield',
                'Quality','Dynamic','Defensive','Low Vol','pct_small','pct_em']
        stats = {k: 0.0 for k in keys}
        total = sum(m.get(wkey, 0) for m in managers)
        if not total: return stats
        for m in managers:
            w = m.get(wkey, 0) / total
            stats['vg_full']    += w * m.get('vg_full', 0)
            stats['vg_3factor'] += w * m.get('vg_3factor', 0)
            stats['pct_small']  += w * m.get('pct_small', 0)
            stats['pct_em']     += w * m.get('pct_em', 0)
            sb = m.get('style_buckets', {})
            for b in ['Core','Growth','Value','Yield','Quality','Dynamic','Defensive','Low Vol']:
                stats[b] += w * sb.get(b, 0)
        return stats
    cur  = agg('current_weight')
    prop = agg('proposed_weight')

    # Portfolio Edge = weighted-average Normalized Skill Z, computed over
    # managers with a valid z score. Missing-z managers are excluded from the
    # weighted average (their weight is dropped), and we report the covered
    # weight so the UI can flag coverage gaps.
    def wavg_z(wkey):
        wz_sum  = 0.0
        w_cov   = 0.0
        w_total = 0.0
        for m in managers:
            w = float(m.get(wkey, 0) or 0)
            w_total += w
            z = m.get('ns_z')
            if z is None: continue
            w_cov  += w
            wz_sum += z * w
        z = (wz_sum / w_cov) if w_cov > 1e-9 else None
        return {'z': z, 'covered_weight': w_cov, 'total_weight': w_total}

    return jsonify({
        'current': cur, 'proposed': prop,
        'delta':   {k: prop[k]-cur[k] for k in cur},
        'edge_current':  wavg_z('current_weight'),
        'edge_proposed': wavg_z('proposed_weight'),
    })

# ── New endpoints for Manager Detail tab ──────────────────────────────────
@app.route('/manager_skill_summary/<tab>/<path:mgr_name>')
def manager_skill_summary(tab, mgr_name):
    """Return skill periods + cumulative skill + period returns for detail tab."""
    import math, pandas as pd
    from data_loader import (compute_skill_periods, compute_cumulative_skill,
                             compute_manager_period_returns, get_benchmark,
                             load_factor_returns)
    if not state['clone_results']:
        return jsonify({'error': 'No results'})
    d = state['clone_results'].get(tab, {}).get(mgr_name)
    if not d:
        return jsonify({'error': f'{tab}/{mgr_name} not found'})

    mgr_rets_all   = d.get('manager_returns', [])
    clone_rets_all = d.get('static_clone_full', [])
    dates_all      = d.get('dates', [])

    # Trim to manager inception: find last index where manager has valid data
    # (most recent first, so find the last non-None value)
    last_valid = len(mgr_rets_all) - 1
    while last_valid >= 0 and (mgr_rets_all[last_valid] is None or
          (isinstance(mgr_rets_all[last_valid], float) and math.isnan(mgr_rets_all[last_valid]))):
        last_valid -= 1
    # Also trim where clone starts (first 24 months have backfilled clone)
    first_clone = 0
    while first_clone < len(clone_rets_all) and (clone_rets_all[first_clone] is None or
          (isinstance(clone_rets_all[first_clone], float) and math.isnan(clone_rets_all[first_clone]))):
        first_clone += 1

    # Use the range where both manager and clone are valid
    inception_idx = min(last_valid, len(mgr_rets_all) - 1)
    mgr_rets   = mgr_rets_all[:inception_idx + 1]
    clone_rets = clone_rets_all[:inception_idx + 1]
    dates      = dates_all[:inception_idx + 1]

    # Ensure factor_df is loaded
    if state.get('factor_df') is None and 'factor_returns' in state['files']:
        state['factor_df'] = load_factor_returns(state['files']['factor_returns'])

    # Get benchmark returns
    bench_name = get_benchmark(tab, mgr_name)
    bench_rets = []
    if state.get('factor_df') is not None and bench_name in state['factor_df'].columns:
        fdf = state['factor_df']
        for ds in dates_all:
            try:
                dt = pd.Timestamp(ds)
                if dt in fdf.index:
                    v = fdf.loc[dt, bench_name]
                    bench_rets.append(None if (v is None or math.isnan(float(v))) else float(v))
                else:
                    bench_rets.append(None)
            except:
                bench_rets.append(None)
        bench_rets = bench_rets[:inception_idx + 1]
    else:
        bench_rets = [None] * len(dates)

    skill_periods  = compute_skill_periods(mgr_rets, clone_rets)
    cumulative     = compute_cumulative_skill(mgr_rets, clone_rets)
    period_returns = compute_manager_period_returns(mgr_rets, clone_rets, bench_rets)

    return jsonify({
        'name': mgr_name, 'tab': tab,
        'benchmark_name': bench_name,
        'skill_periods': skill_periods,
        'cumulative_skill': cumulative,
        'dates': dates,
        'period_returns': period_returns,
        'betas_full': d.get('betas_full', {}),
        'style_buckets': d.get('style_buckets', {}),
        'r2_full': d.get('r2_full'),
    })


@app.route('/peer_skill_summary/<tab>')
def peer_skill_summary(tab):
    """Return skill periods for all managers in a tab — for the sortable table."""
    from data_loader import compute_skill_periods
    if not state['clone_results']:
        return jsonify({'error': 'No results'})
    # Synthesise the Placeholder peer group: every manager appearing in
    # the FactSet exposures / risk / weights files but lacking a clone,
    # with bucket overrides from state['placeholder_buckets'] applied.
    if tab == 'Placeholder':
        ph_buckets = state.get('placeholder_buckets') or {}
        out = []
        for name in sorted(_enumerate_placeholder_candidates(), key=str.lower):
            sb = dict(ph_buckets.get(name) or {'Core': 1.0})
            v = sum(w for b, w in sb.items() if b == 'Value')
            y = sum(w for b, w in sb.items() if b == 'Yield')
            g = sum(w for b, w in sb.items() if b == 'Growth')
            vg = round((v + y) - g, 6)
            out.append({
                'name':            name,
                'r2_full':         None,
                'vg_full':         vg,
                'style_buckets':   sb,
                'pct_small':       sb.get('Small Cap', 0) or 0,
                'pct_em':          0,
                'betas_full':      {},
                'betas_3factor':   {},
                't1_skill': None, 't3_skill': None,
                't5_skill': None, 'si_skill': None,
                'is_placeholder':  True,
            })
        return jsonify({'tab': tab, 'managers': out})

    td = state['clone_results'].get(tab, {})
    out = []
    for name, d in td.items():
        sp = compute_skill_periods(d.get('manager_returns', []),
                                   d.get('static_clone_full', []))
        row = {
            'name': name,
            'r2_full': d.get('r2_full'),
            'vg_full': d.get('vg_full', 0),
            'style_buckets': d.get('style_buckets', {}),
            'pct_small': d.get('pct_small', 0),
            'pct_em': d.get('pct_em', 0),
            'betas_full': d.get('betas_full', {}),
            'betas_3factor': d.get('betas_3factor', {}),
            't1_skill':  sp.get('t1'),
            't3_skill':  sp.get('t3'),
            't5_skill':  sp.get('t5'),
            'si_skill':  sp.get('si'),
        }
        row.update(_norm_skill_for(tab, name))
        row.update(_qual_fields(name))
        out.append(row)
    return jsonify({'tab': tab, 'managers': out})


# ── Portfolio contribution stats endpoint ─────────────────────────────────
@app.route('/portfolio_contribution/<client_name>')
def portfolio_contribution(client_name):
    """
    Per-manager stats for Portfolio Contribution tables.
    Returns manager_return, benchmark_excess, passive_style, skill
    for QTD, T1, T3 for each manager in the client portfolio.
    """
    if not state['clone_results']:
        return jsonify({'error': 'No results'})
    if not state['weights'] or client_name not in state['weights']:
        return jsonify({'error': 'Client not found'})
    from data_loader import build_portfolio_view
    portfolio = build_portfolio_view(client_name, state['weights'][client_name],
                                     state['clone_results'], state['manager_dfs'],
                                     universe_clone_results=state.get('universe_clone_results'),
                                     placeholder_buckets=state.get('placeholder_buckets') or {})
    result = _build_portfolio_contribution_rows(portfolio['managers'])
    result['unmatched'] = portfolio.get('unmatched', [])
    return jsonify(result)


@app.route('/portfolio_contribution_preview', methods=['POST'])
def portfolio_contribution_preview():
    """Preview contribution tables for a client-edited portfolio payload."""
    if not state['clone_results']:
        return jsonify({'error': 'No results'})

    payload = request.get_json(silent=True) or {}
    managers = payload.get('managers', [])
    if not managers:
        return jsonify({'managers': [], 'unmatched': []})

    return jsonify(_build_portfolio_contribution_rows(managers))


# ── FactSet Risk Summary ──────────────────────────────────────────────────
STYLE_FACTORS = [
    'Beta', 'Book to Price', 'Dividend Yield', 'Earnings Yield',
    'Growth', 'Leverage', 'Liquidity', 'Momentum',
    'Profitability', 'Size', 'Volatility'
]

# Section markers that appear in column A of the Risk Summary export. These
# delineate the top-level groups of factor rows.
_RISK_SECTIONS = ('Style', 'Industry', 'Country', 'Currency')

# Benchmark-name heuristics (same families used by the exposures parser).
_RISK_BENCH_PREFIXES = ('MSCI ', 'FTSE ', 'RUSSELL ', 'S&P ', 'BLOOMBERG ',
                         'STOXX ', 'NIKKEI ', 'TOPIX ')
_RISK_BENCH_KEYWORDS = ('BENCHMARK', 'INDEX')


def _is_risk_bench_name(nm):
    """Match the convention FactSet uses in portfolio lists: any column
    whose header starts with a known index-family prefix, or contains
    'Benchmark'/'Index', is treated as a benchmark row. Everything else
    is a manager portfolio."""
    if not nm:
        return False
    up = str(nm).upper()
    return up.startswith(_RISK_BENCH_PREFIXES) or any(k in up for k in _RISK_BENCH_KEYWORDS)


def parse_risk_summary(filepath):
    """Parse a FactSet Risk Summary export into a structured dict.

    The export is expected to list managers AND candidate benchmarks as
    columns on the same sheet (in the 'Active vs USD' configuration, which
    is equivalent to absolute exposure for our purposes — see risk doc).
    Row structure:
      Row ~7:  column header names (managers + benchmarks mixed)
      Row 10:  'Active Risk Factor Decomp'
      Row 11:  'Active Exposure'
      Rows 12+: Market / Global Market / Style (aggregate) / 11 style factors
               / Industry (aggregate) + individual industries
               / Country (aggregate) + individual countries
               / Currency (aggregate) + individual currencies

    Returns:
      {
        'manager_names':   [str, ...],              # in file order
        'benchmark_names': [str, ...],              # ditto
        'style_factors':   {factor: {col_name: float or None}},
        'industries':      {industry: {col_name: float or None}},
        'countries':       {country:  {col_name: float or None}},
        'currencies':      {currency: {col_name: float or None}},
      }
    Downstream code combines managers' absolute exposures by weight and
    subtracts the client benchmark's absolute exposures to produce true
    portfolio-vs-benchmark active exposures (the 'bottom-up' computation).
    """
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the header row: first row where col B is a non-empty string
    # other than 'Data'. The FactSet export repeats the header — use the
    # first occurrence.
    col_names = []
    for row in rows:
        if len(row) < 2 or not row[1]:
            continue
        if str(row[1]).strip().lower() == 'data':
            continue
        non_none = [x for x in row[1:] if x is not None and str(x).strip() != '']
        if non_none and isinstance(non_none[0], str):
            col_names = [str(x).strip() if x is not None else None
                         for x in row[1:]]
            # Trim trailing Nones (columns past the last named one)
            while col_names and not col_names[-1]:
                col_names.pop()
            break

    if not col_names:
        return {'manager_names': [], 'benchmark_names': [],
                'style_factors': {}, 'industries': {}, 'countries': {},
                'currencies': {}}

    # Classify each column
    manager_names   = [c for c in col_names if c and not _is_risk_bench_name(c)]
    benchmark_names = [c for c in col_names if c and _is_risk_bench_name(c)]

    def _coerce(v):
        if v is None or v == '':
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def _row_to_map(row):
        """Given a data row, return {col_name: float or None} keyed by the
        column headers (skipping unlabelled columns)."""
        out = {}
        # row[0] is the label, row[1:] align with col_names
        vals = list(row[1:1 + len(col_names)])
        for name, v in zip(col_names, vals):
            if not name:
                continue
            out[name] = _coerce(v)
        return out

    # Walk the file, tracking which section we're in. The aggregate rows
    # labelled exactly 'Style'/'Industry'/'Country'/'Currency' flip the
    # section; every subsequent non-empty label until the next marker is a
    # member row of that section.
    style_factors = {}
    industries    = {}
    countries     = {}
    currencies    = {}

    section = None
    buckets = {
        'Style':    style_factors,
        'Industry': industries,
        'Country':  countries,
        'Currency': currencies,
    }

    for row in rows:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if label in _RISK_SECTIONS:
            # Start of a new section; skip the aggregate row itself.
            section = label
            continue
        if section is None:
            continue
        # Ignore section-level aggregate labels that repeat header text.
        if label in ('Market', 'Global Market', 'Active Exposure',
                     'Active Risk Factor Decomp'):
            continue
        # If the row has no numeric data under any column, treat it as
        # stray padding and skip.
        data_map = _row_to_map(row)
        if not any(v is not None for v in data_map.values()):
            continue
        buckets[section][label] = data_map

    return {
        'manager_names':   manager_names,
        'benchmark_names': benchmark_names,
        'style_factors':   style_factors,
        'industries':      industries,
        'countries':       countries,
        'currencies':      currencies,
    }


@app.route('/upload_risk', methods=['POST'])
def upload_risk():
    if 'risk_summary' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file'})
    f = request.files['risk_summary']
    if not f.filename:
        return jsonify({'status': 'error', 'message': 'No filename'})
    path = save_uploaded_file(f, secure_filename(f.filename), app.config['UPLOAD_FOLDER'])
    try:
        local = resolve_path(path, app.config['UPLOAD_FOLDER'], suffix='.xlsx')
        parsed = parse_risk_summary(local)
        state['risk_data'] = parsed
        state['files']['risk_summary'] = path
        save_cache()
        return jsonify({
            'status':    'ok',
            'managers':  parsed['manager_names'],
            'benchmarks': parsed['benchmark_names'],
            'factors':   list(parsed['style_factors'].keys()),
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/upload_security_risk', methods=['POST'])
def upload_security_risk():
    """Upload a FactSet Security-Level Risk DNA workbook. Parses manager
    holdings (stock weights + raw factor exposures + country) and benchmark
    absolute factor exposures from the Risk Summary tab. Stores in
    state['security_risk_data'] for use by /compute_security_risk_exposures."""
    if 'security_risk' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file provided'})
    f = request.files['security_risk']
    if not f.filename:
        return jsonify({'status': 'error', 'message': 'Empty filename'})
    path = save_uploaded_file(f, secure_filename(f.filename), app.config['UPLOAD_FOLDER'])
    try:
        from security_risk_engine import parse_security_risk_file
        # Resolve to a local path first — in S3 mode `path` is an S3 key, not a
        # local file, so the parser must read the downloaded copy.
        local = resolve_path(path, app.config['UPLOAD_FOLDER'], suffix='.xlsx')
        parsed = parse_security_risk_file(local)
        if not parsed['managers']:
            return jsonify({'status': 'error',
                            'message': 'No manager sections found. Check the '
                                       'file has a Security-Level Risk DNA sheet.'})
        state['security_risk_data'] = parsed
        state['files']['security_risk'] = path
        save_cache()
        return jsonify({
            'status':      'ok',
            'managers':    list(parsed['managers'].keys()),
            'benchmarks':  parsed['available_benchmarks'],
            'factors':     parsed['factors'],
            'as_of_date':  parsed.get('as_of_date'),
        })
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'message': str(e),
                        'detail': traceback.format_exc()})


@app.route('/sleeve_options', methods=['POST'])
def sleeve_options():
    """Return the available sleeve breakdown options for a client given their
    benchmark and the benchmarks present in the uploaded risk files."""
    data = request.get_json(silent=True) or {}
    client_name = data.get('client_name')
    bench_str   = (state.get('client_benchmarks') or {}).get(client_name or '')
    available   = (state.get('security_risk_data') or {}).get('available_benchmarks', [])
    if not available:
        # Mirror the fallback in /compute_security_risk_exposures: the
        # Security-Level Risk DNA file may have no embedded 'Risk Summary'
        # sheet, in which case the benchmark side comes from the separate
        # Benchmark Risk Summaries upload.
        available = list((state.get('risk_data') or {}).get('benchmark_names', []))
    from security_risk_engine import get_sleeve_options
    options = get_sleeve_options(bench_str or '', available)
    return jsonify({'options': options, 'benchmark': bench_str})


@app.route('/compute_security_risk_exposures', methods=['POST'])
def compute_security_risk_exposures():
    """Stock-level (bottom-up) active factor exposure computation.

    Accepts JSON:
      { managers: [...], client_name: str, sleeve: str|null, bench: str|null }

    sleeve is one of: null (full portfolio), 'US', 'Non-US', 'EM'
    bench is the exact benchmark column name from the Risk Summary tab.

    Returns the same response shape as /compute_risk_exposures so the
    frontend's renderRiskExposures() works unchanged, with an extra
    'sleeve_info' key carrying coverage % and country flags.
    """
    if not state.get('security_risk_data'):
        return jsonify({'error': 'No security risk data loaded. Upload a '
                                 'Security-Level Risk DNA file on the Setup tab.'})
    payload     = request.get_json(silent=True) or {}
    managers    = payload.get('managers', [])
    client_name = payload.get('client_name')
    sleeve      = payload.get('sleeve')   # None | 'US' | 'Non-US' | 'EM'
    bench_name  = payload.get('bench')    # exact column name from file
    if not managers:
        return jsonify({'error': 'No managers provided.'})

    # Use security_risk_data as-is, but if the security file has no
    # embedded 'Risk Summary' sheet (so available_benchmarks is empty),
    # splice in benchmarks from the separately-uploaded Benchmark Risk
    # Summaries file. Without this fallback the active-vs-benchmark math
    # silently degrades to 'absolute portfolio exposures' even when the
    # user has clearly uploaded benchmark data — they just put it in the
    # second file rather than as a sheet inside the security file.
    sec_data = state['security_risk_data']
    if not sec_data.get('available_benchmarks'):
        rd = state.get('risk_data') or {}
        rd_bench_names   = rd.get('benchmark_names', [])
        rd_style_factors = rd.get('style_factors', {})  # {factor: {col: val}}
        if rd_bench_names:
            from security_risk_engine import STYLE_FACTORS as _SF
            merged = {bn: {} for bn in rd_bench_names}
            for factor, col_map in rd_style_factors.items():
                if factor not in _SF:
                    continue
                for bn in rd_bench_names:
                    v = (col_map or {}).get(bn)
                    if v is not None:
                        merged[bn][factor] = v
            sec_data = dict(sec_data)
            sec_data['benchmarks']           = merged
            sec_data['available_benchmarks'] = list(rd_bench_names)

    # Always look up the breakdown's sleeve set so we know whether an EM
    # sleeve exists for this client's benchmark. classify_country needs
    # this to correctly route EM / unclassified countries:
    #   3-way (US/Non-US Dev/EM) → EM and unclassified go to EM.
    #   2-way (US/Non-US)        → EM and unclassified lump into Non-US.
    # Determined by the BREAKDOWN, not by which sleeve the user has
    # picked — so picking 'Non-US Dev' inside a 3-way ACWI ex-US split
    # correctly excludes EM stocks from Non-US Dev.
    from security_risk_engine import get_sleeve_options
    bench_str = (state.get('client_benchmarks') or {}).get(client_name or '')
    all_opts = get_sleeve_options(bench_str or '', sec_data.get('available_benchmarks', []))
    has_em_sleeve = any(o.get('sleeve') == 'EM' for o in all_opts)

    # Fall back to client benchmark from weights file if bench not supplied
    if not bench_name:
        bench_name = all_opts[0]['bench'] if all_opts else None

    from security_risk_engine import compute_exposures
    try:
        result = compute_exposures(
            portfolio_managers=managers,
            security_data=sec_data,
            benchmark_name=bench_name,
            sleeve=sleeve,
            has_em_sleeve=has_em_sleeve,
        )
        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detail': traceback.format_exc()})


@app.route('/compute_risk_exposures', methods=['POST'])
def compute_risk_exposures():
    """
    Compute portfolio-vs-client-benchmark active style exposures from the
    bottom up, using the new FactSet Risk Summary format where both managers
    AND candidate benchmarks sit side-by-side as absolute exposures
    (active vs USD, which is mathematically equivalent to absolute for
    our purposes).

    For each style factor:
      portfolio_current_abs  = Σ(w_current_i  × manager_i_abs)
      portfolio_proposed_abs = Σ(w_proposed_i × manager_i_abs)
      benchmark_abs          = risk_file[factor][client_benchmark]
      current_active  = portfolio_current_abs  − benchmark_abs
      proposed_active = portfolio_proposed_abs − benchmark_abs
      delta           = proposed_active − current_active    (same as
                        proposed_abs − current_abs — benchmark cancels)

    The client benchmark is resolved from state['client_benchmarks'] (the
    weights file). If that benchmark doesn't match a column in the risk
    file, we fall back to computing without a benchmark (i.e. raw portfolio
    absolute) and flag the condition so the UI can show a warning.

    Weights are renormalised across the managers that DID match the risk
    file, so a portfolio with one un-matched manager still produces a
    meaningful active number for the rest — matching the old endpoint's
    behaviour.
    """
    if not state['risk_data']:
        return jsonify({'error': 'No risk data loaded'})

    data        = request.json or {}
    managers    = data.get('managers', [])
    client_name = data.get('client_name')
    rd          = state['risk_data']

    # Backwards compat: if the loaded cache is from the OLD parser (which
    # returned {'managers': [...], 'style_factors': {factor: [list]}}),
    # refuse to run rather than produce garbage. User needs to re-upload.
    if 'manager_names' not in rd:
        return jsonify({'error': 'Risk summary cache is in the old format. '
                                 'Re-upload the Risk Summary file to use the '
                                 'new bottom-up computation.'})

    mgr_cols   = rd.get('manager_names', [])
    bench_cols = rd.get('benchmark_names', [])
    factors    = rd.get('style_factors', {})   # {factor: {col_name: value}}

    # ── Manager-name matching ────────────────────────────────────────────
    # Matching is a three-tier cascade designed to handle two situations
    # the previous implementation got wrong:
    #   (1) The risk file may contain BOTH 'Mgr EAFE' and 'Mgr EAFE SC' for
    #       the same firm's LC and SC sleeves — our normaliser strips
    #       'eafe sc' and 'eafe' both to nothing, so naive dict-comprehension
    #       lookups silently dropped one variant and returned the wrong
    #       sleeve's data. Client 1's size exposure flipped from -0.32 to
    #       +0.17 because of this.
    #   (2) A buy-list manager's name may differ in its regional tag from
    #       the risk-file column ('Ballina ISC' in the buy-list vs 'Ballina
    #       EAFE SC' in the risk file).
    #
    # The fix: keep a LIST of candidates per normalised key (not a dict that
    # overwrites collisions) and break ties by the manager's peer tab. ISC
    # and USSC tabs prefer risk columns whose name carries a small-cap
    # marker; EAFE/US/ACWI/EM tabs prefer columns without one.
    import re
    def norm(s):
        s = str(s).lower().strip()
        s = re.sub(r'\([^)]*\)', '', s)
        s = re.sub(r'[\./\-_,]+', ' ', s)
        suffixes = [
            'international small cap equity',
            'international small cap',
            'international small company',
            'non us small cap',
            'non us equity',
            'international equity',
            'smid cap', 'small/mid cap', 'small mid cap', 'small cap',
            'small company', 'micro cap', 'emerging markets', 'emerging mkts',
            'eafe sc', 'acwi sc', 'xus sc', 'em sc',
            'eafe', 'acwi', 'xus', 'em', 'us', 'global', 'isc',
            'non us', 'international',
            'equity', 'equities',
            'composite', 'portfolio', 'strategy', 'fund',
            'mid cap', 'large cap',
            'growth', 'value', 'blend', 'core',
            'sc', 'lc',
        ]
        for suffix in suffixes:
            s = re.sub(r'\b' + re.escape(suffix) + r'\b', '', s)
        return re.sub(r'\s+', ' ', s).strip()

    # Normalised key → LIST of candidate column names (not just the last one)
    mgr_key_to_cols = {}
    for c in mgr_cols:
        mgr_key_to_cols.setdefault(norm(c), []).append(c)
    mgr_raw_lower = [(n, str(n).lower().strip()) for n in mgr_cols]

    _SC_TOKENS = (' sc', ' small cap', ' smallcap', ' small-cap', ' isc')

    def _col_is_smallcap(col):
        """Is this risk-file column a small-cap sleeve? Token-match on the
        column name so 'Ballina EAFE SC' → True, 'Ballina EAFE' → False."""
        c = ' ' + str(col).lower() + ' '
        return any(tok + ' ' in c or tok == c.rstrip() for tok in _SC_TOKENS)

    def _pick_by_tab(candidates, tab):
        """When a normalised key has multiple candidate columns, use the
        manager's peer tab to pick the right one. Falls back to the first
        candidate if the tab doesn't disambiguate."""
        if len(candidates) == 1:
            return candidates[0]
        prefer_sc = tab in ('ISC', 'USSC')
        sc_cands    = [c for c in candidates if _col_is_smallcap(c)]
        nonsc_cands = [c for c in candidates if not _col_is_smallcap(c)]
        if prefer_sc and sc_cands:
            return sc_cands[0]
        if (not prefer_sc) and nonsc_cands:
            return nonsc_cands[0]
        return candidates[0]

    def match_manager_col(mgr):
        """Return the exact manager-column name in the risk file that best
        matches this buy-list manager, or None. `mgr` is the dict from the
        frontend, with 'matched_name' (or 'name') and 'tab' fields."""
        name = mgr.get('matched_name') or mgr.get('name', '')
        tab  = mgr.get('tab')
        if not name:
            return None
        raw_lower = str(name).lower().strip()

        # (1) Exact raw (case-insensitive) match — preserves LC/SC distinction
        # in the common case where buy-list and risk-file use identical names.
        for orig, rl in mgr_raw_lower:
            if rl == raw_lower:
                return orig

        # (2) Normalised-key match with collision-aware selection
        key = norm(name)
        cands = mgr_key_to_cols.get(key)
        if cands:
            return _pick_by_tab(cands, tab)

        # (3) Short-stem prefix / word-boundary match
        if key and len(key) <= 5:
            hits = []
            for orig, rl in mgr_raw_lower:
                if rl.startswith(key + ' ') or rl == key \
                   or (' ' + key + ' ') in (' ' + rl + ' '):
                    hits.append(orig)
            if hits:
                return _pick_by_tab(hits, tab)

        # (4) Fuzzy fallback — WRatio handles partial/length-asymmetric matches
        from rapidfuzz import process, fuzz
        candidates = [k for k in mgr_key_to_cols.keys() if k]
        if candidates and key:
            m = process.extractOne(key, candidates,
                                   scorer=fuzz.WRatio, score_cutoff=75)
            if m:
                return _pick_by_tab(mgr_key_to_cols[m[0]], tab)
        return None

    # ── Benchmark-column matching ────────────────────────────────────────
    # Client's benchmark comes from the weights file (state['client_benchmarks']).
    # Risk-file benchmark columns are things like 'MSCI EAFE', 'MSCI EAFE SC',
    # 'MSCI EAFE + Canada' etc. Names in the weights file may differ slightly
    # ('MSCI EAFE+CANADA' vs 'MSCI EAFE + Canada') — normalise whitespace and
    # punctuation before matching.
    def _bench_norm(s):
        t = str(s or '').lower()
        for ch in ['+', '-', '/', ',', '.']:
            t = t.replace(ch, ' ')
        return re.sub(r'\s+', ' ', t).strip()

    bench_lookup = {_bench_norm(b): b for b in bench_cols}

    client_bench_str = (data.get('benchmark_name')
                        or (state.get('client_benchmarks') or {}).get(client_name or ''))
    matched_bench_col = None
    if client_bench_str:
        key = _bench_norm(client_bench_str)
        matched_bench_col = bench_lookup.get(key)
        if matched_bench_col is None:
            # Fuzzy fallback — handles 'MSCI EAFE SC' vs 'MSCI EAFE Small Cap'
            from rapidfuzz import process, fuzz
            cand = list(bench_lookup.keys())
            if cand:
                m = process.extractOne(key, cand, scorer=fuzz.WRatio,
                                        score_cutoff=80)
                if m:
                    matched_bench_col = bench_lookup[m[0]]

    # Per-factor benchmark absolute values (None if benchmark not found)
    def bench_val(factor):
        if not matched_bench_col:
            return None
        return factors.get(factor, {}).get(matched_bench_col)

    # ── Per-manager factor lookups ───────────────────────────────────────
    def mgr_val(mgr, factor):
        col = match_manager_col(mgr)
        if not col:
            return None
        return factors.get(factor, {}).get(col)

    # ── Bottom-up sumproduct → active ────────────────────────────────────
    def sumproduct(weight_key):
        """Weighted-average absolute exposures across matched managers.
        Weights are renormalised across matched managers only — an
        unmatched manager doesn't poison the aggregate."""
        out = {}
        for factor in STYLE_FACTORS:
            matched_pairs = []   # list of (w, value) for matched managers
            for m in managers:
                w = m.get(weight_key, 0) or 0
                if w <= 0:
                    continue
                v = mgr_val(m, factor)
                if v is None:
                    continue
                matched_pairs.append((w, v))
            total_w = sum(w for w, _ in matched_pairs)
            if total_w <= 0:
                out[factor] = None
                continue
            out[factor] = sum((w / total_w) * v for w, v in matched_pairs)
        return out

    cur_abs  = sumproduct('current_weight')
    prop_abs = sumproduct('proposed_weight')

    # Active = absolute − benchmark_absolute. When benchmark is unavailable
    # we fall back to absolute (matches legacy behaviour and keeps the UI
    # responsive — warning is surfaced separately).
    current  = {}
    proposed = {}
    for f in STYLE_FACTORS:
        b = bench_val(f)
        if cur_abs[f] is None:
            current[f] = None
        elif b is None:
            current[f] = round(cur_abs[f], 6)
        else:
            current[f] = round(cur_abs[f] - b, 6)
        if prop_abs[f] is None:
            proposed[f] = None
        elif b is None:
            proposed[f] = round(prop_abs[f], 6)
        else:
            proposed[f] = round(prop_abs[f] - b, 6)
    delta = {f: round(proposed[f] - current[f], 6)
             if proposed[f] is not None and current[f] is not None else None
             for f in STYLE_FACTORS}

    # Flag managers with non-zero weight whose exposures couldn't be matched
    unmatched = []
    for m in managers:
        if (m.get('current_weight', 0) or 0) > 0 or (m.get('proposed_weight', 0) or 0) > 0:
            display_name = m.get('matched_name') or m.get('name', '?')
            if mgr_val(m, 'Beta') is None:
                unmatched.append(display_name)

    # Benchmark-resolution notes for the UI
    benchmark_info = {
        'requested':         client_bench_str,
        'matched_column':    matched_bench_col,
        'available_columns': bench_cols,
        'fallback_absolute': bool(client_bench_str and not matched_bench_col),
    }

    return jsonify({
        'factors':     STYLE_FACTORS,
        'current':     current,
        'proposed':    proposed,
        'delta':       delta,
        'unmatched':   unmatched,
        'rd_managers': mgr_cols,
        'benchmark':   benchmark_info,
    })


# ── Risk analysis endpoint (Scenario / Marginal Contribution / Regime) ────
@app.route('/risk_analysis', methods=['POST'])
def risk_analysis():
    """
    Accepts JSON: { managers: [{matched_name, tab, current_weight, proposed_weight}, ...] }
    Returns the full risk analysis payload.
    Sending managers from the frontend (rather than reading from state) lets
    the user edit proposed weights and see updated numbers without reloading.
    """
    from risk_engine import compute_risk_analysis
    from data_loader import load_factor_returns

    if not state['clone_results']:
        return jsonify({'error': 'Run the clone engine first.'})

    # Ensure factor_df is loaded
    if state.get('factor_df') is None:
        if 'factor_returns' not in state['files']:
            return jsonify({'error': 'Factor returns file missing.'})
        state['factor_df'] = load_factor_returns(state['files']['factor_returns'])

    payload = request.get_json(silent=True) or {}
    managers = payload.get('managers', [])
    client_name = payload.get('client_name')
    peer_benchmark = payload.get('peer_benchmark')
    if not managers:
        return jsonify({'error': 'No managers provided.'})

    # If the frontend didn't pick an explicit peer benchmark, try the
    # client's benchmark from the weights file. This supersedes the
    # hardcoded CLIENT_BENCHMARK_OVERRIDE inside compute_risk_analysis.
    if not peer_benchmark and client_name:
        from data_loader import resolve_peer_group
        bench_str = (state.get('client_benchmarks') or {}).get(client_name)
        resolved  = resolve_peer_group(bench_str)
        if resolved:
            peer_benchmark = resolved

    # Defensive coercion — the frontend should already send clean floats
    clean = []
    for m in managers:
        clean.append({
            'matched_name':    m.get('matched_name') or m.get('name', ''),
            'tab':             m.get('tab'),
            'current_weight':  float(m.get('current_weight', 0) or 0),
            'proposed_weight': float(m.get('proposed_weight', 0) or 0),
        })

    try:
        result = compute_risk_analysis(
            clean, state['clone_results'], state['factor_df'],
            client_name=client_name, peer_benchmark=peer_benchmark,
        )
    except Exception as e:
        import traceback
        return jsonify({'error': f'{e}', 'trace': traceback.format_exc()})
    return jsonify(result)


@app.route('/ideal_complement', methods=['POST'])
def ideal_complement():
    """
    Find the manager whose track record best offsets the *proposed* portfolio's
    underperformance.

    Logic (locked in via user spec):
      1. Build proposed portfolio monthly returns from clone_results + weights.
      2. Compute proposed excess vs the client benchmark (same yardstick used
         for the rest of the report). Find months where excess < 0.
      3. Eligible candidates = managers whose `tab` matches a peer group with
         non-zero proposed weight AND who are not already held in the portfolio.
      4. For each candidate, on the proposed-portfolio-underperformance months
         that the candidate also has data for: compute
           - hit_rate  = # months candidate beat the benchmark / # overlap months
           - avg_excess = mean(candidate_return − benchmark_return) over those months
      5. Require ≥ 12 overlapping months — anything less is too noisy.
      6. Rank by hit_rate desc, tiebreak by avg_excess desc. Return the top one
         along with its full manager metadata so the UI can render it in the
         portfolio table.
    """
    import numpy as np
    import pandas as pd
    from risk_engine import (
        build_return_matrix, build_benchmark_series, portfolio_return_series,
        PEER_BENCHMARKS, CLIENT_BENCHMARK_OVERRIDE,
    )
    from data_loader import load_factor_returns, resolve_peer_group

    if not state['clone_results']:
        return jsonify({'error': 'Run the clone engine first.'})
    if state.get('factor_df') is None:
        if 'factor_returns' not in state['files']:
            return jsonify({'error': 'Factor returns file missing.'})
        state['factor_df'] = load_factor_returns(state['files']['factor_returns'])

    payload = request.get_json(silent=True) or {}
    managers = payload.get('managers', [])
    client_name = payload.get('client_name')
    if not managers:
        return jsonify({'error': 'No managers provided.'})

    # ── 1. Held managers + eligible peer groups (proposed_weight > 0) ─────
    held_names = set()
    eligible_tabs = set()
    proposed_managers = []
    for m in managers:
        nm  = m.get('matched_name') or m.get('name', '')
        tab = m.get('tab')
        pw  = float(m.get('proposed_weight', 0) or 0)
        if nm and pw > 0:
            held_names.add(nm)
            if tab:
                eligible_tabs.add(tab)
                proposed_managers.append({'matched_name': nm, 'tab': tab, 'proposed_weight': pw})
    if not proposed_managers:
        return jsonify({'error': 'No managers with proposed weight > 0.'})
    if not eligible_tabs:
        return jsonify({'error': 'Cannot determine peer groups from proposed managers.'})

    # ── 2. Resolve client benchmark peer group ────────────────────────────
    peer_benchmark = None
    if client_name:
        bench_str = (state.get('client_benchmarks') or {}).get(client_name)
        peer_benchmark = resolve_peer_group(bench_str)
    # Fallback: per-client hardcoded override, else the dominant peer group
    if not peer_benchmark and client_name in CLIENT_BENCHMARK_OVERRIDE:
        peer_benchmark = CLIENT_BENCHMARK_OVERRIDE[client_name]
    if not peer_benchmark:
        # Pick the tab with the largest proposed weight as the benchmark anchor
        tab_weights = {}
        for pm in proposed_managers:
            tab_weights[pm['tab']] = tab_weights.get(pm['tab'], 0) + pm['proposed_weight']
        peer_benchmark = max(tab_weights, key=tab_weights.get)

    # ── 3. Proposed portfolio returns + benchmark series ──────────────────
    ret_mtx = build_return_matrix(proposed_managers, state['clone_results'])
    if ret_mtx.empty:
        return jsonify({'error': 'No return data for proposed managers.'})

    bench_df = build_benchmark_series(peer_benchmark, state['factor_df'], ret_mtx.index)
    if bench_df is None or 'core' not in bench_df.columns:
        return jsonify({'error': f'Benchmark series unavailable for peer group {peer_benchmark}.'})
    bench_core = bench_df['core'].dropna()

    weights_dict = {pm['matched_name']: pm['proposed_weight'] for pm in proposed_managers}
    port_rets = portfolio_return_series(ret_mtx, weights_dict)
    excess    = (port_rets - bench_core).dropna()
    under_idx = excess[excess < 0].index
    if len(under_idx) < 12:
        return jsonify({
            'error': f'Proposed portfolio only has {len(under_idx)} underperformance month(s). '
                     f'Need at least 12 to compute a meaningful complement.',
            'n_underperform_months': int(len(under_idx)),
        })

    # ── 4. Score each candidate on the underperformance subset ────────────
    MIN_OVERLAP = 12
    candidates = []
    for tab in eligible_tabs:
        for cname, cdata in (state['clone_results'].get(tab, {}) or {}).items():
            if cname in held_names:
                continue
            dates  = cdata.get('dates') or []
            mrets  = cdata.get('manager_returns') or []
            if not dates or not mrets:
                continue
            n = min(len(dates), len(mrets))
            try:
                idx = pd.DatetimeIndex(pd.to_datetime(dates[:n]))
            except Exception:
                continue
            cs = pd.Series([float(v) if v is not None and not (isinstance(v, float) and np.isnan(v)) else np.nan
                            for v in mrets[:n]], index=idx, dtype=float).dropna()
            cs = cs[~cs.index.duplicated(keep='first')]

            common = cs.index.intersection(under_idx)
            if len(common) < MIN_OVERLAP:
                continue
            cand_on = cs.loc[common]
            bench_on = bench_core.reindex(common).dropna()
            common  = cand_on.index.intersection(bench_on.index)
            if len(common) < MIN_OVERLAP:
                continue
            cand_on  = cand_on.loc[common]
            bench_on = bench_on.loc[common]
            diff = cand_on - bench_on
            hit_rate   = float((diff > 0).sum()) / float(len(diff))
            avg_excess = float(diff.mean())

            ns = _norm_skill_for(tab, cname) or {}
            candidates.append({
                'name': cname,
                'matched_name': cname,
                'tab': tab,
                'peer_group': cdata.get('peer_group', tab),
                'hit_rate': hit_rate,
                'avg_excess': avg_excess,
                'n_months': int(len(diff)),
                'vg_full':     cdata.get('vg_full', 0),
                'vg_3factor':  cdata.get('vg_3factor', 0),
                'r2_full':     cdata.get('r2_full', 0),
                'r2_3factor':  cdata.get('r2_3factor', 0),
                'style_buckets': cdata.get('style_buckets', {}) or {},
                'pct_small':   cdata.get('pct_small', 0),
                'pct_em':      cdata.get('pct_em', 0),
                'ns_z':        ns.get('z'),
                'ns_skill':    ns.get('skill'),
                'ns_n_obs':    ns.get('n_obs'),
                'ns_n_peers':  ns.get('n_peers'),
                'current_weight':  0.0,
                'proposed_weight': 0.0,
            })

    if not candidates:
        return jsonify({
            'error': f'No candidates with at least {MIN_OVERLAP} overlapping underperformance months.',
            'n_underperform_months': int(len(under_idx)),
            'eligible_tabs': sorted(eligible_tabs),
        })

    # ── 5. Rank: hit_rate desc, then avg_excess desc ──────────────────────
    candidates.sort(key=lambda c: (-c['hit_rate'], -c['avg_excess']))
    best = candidates[0]
    return jsonify({
        'best': best,
        'n_underperform_months': int(len(under_idx)),
        'n_candidates_considered': len(candidates),
        'eligible_tabs': sorted(eligible_tabs),
        'peer_benchmark': peer_benchmark,
        'benchmark_name': (PEER_BENCHMARKS.get(peer_benchmark) or {}).get('core'),
    })


@app.route('/portfolio_report/<client_name>')
def portfolio_report(client_name):
    """Compose the data the printable Report tab needs for one client.
    Returns the same shape as the legacy REPORT_MOCK, computed from real
    clone_results + weights + factor_df. Includes both the backtested
    portfolio performance series and an Aapryl-factor + FactSet-risk
    complement pick (in addition to the manager complement)."""
    import numpy as np
    import pandas as pd
    from risk_engine import (
        build_return_matrix, build_benchmark_series, portfolio_return_series,
        PEER_BENCHMARKS, CLIENT_BENCHMARK_OVERRIDE,
    )
    from data_loader import (
        build_portfolio_view, load_factor_returns, resolve_peer_group,
    )
    from report_engine import (
        compute_trailing_periods, compute_calendar_years,
        compute_quarterly_excess, compute_factor_complements,
        find_inception_date,
    )

    if not state.get('clone_results'):
        return jsonify({'error': 'Run the clone engine first.'})
    if not state.get('weights') or client_name not in state['weights']:
        return jsonify({'error': f'Client {client_name!r} not found.'})

    # Ensure factor_df is loaded (same lazy pattern as /ideal_complement)
    if state.get('factor_df') is None:
        if 'factor_returns' not in state.get('files', {}):
            return jsonify({'error': 'Factor returns file missing.'})
        state['factor_df'] = load_factor_returns(state['files']['factor_returns'])
    factor_df = state['factor_df']

    # ── 1. Resolve managers + weights via the same path /portfolio uses ──
    portfolio = build_portfolio_view(
        client_name, state['weights'][client_name],
        state['clone_results'], state['manager_dfs'],
        universe_clone_results=state.get('universe_clone_results'),
        placeholder_buckets=state.get('placeholder_buckets') or {},
    )
    managers = [m for m in portfolio.get('managers', [])
                if not m.get('is_placeholder')
                and float(m.get('current_weight', 0) or 0) > 0]
    if not managers:
        return jsonify({'error': 'Client has no managers with weight data.'})

    # ── 2. Resolve benchmark peer group ──────────────────────────────────
    bench_str = (state.get('client_benchmarks') or {}).get(client_name) or ''
    peer_benchmark = resolve_peer_group(bench_str)
    if not peer_benchmark and client_name in CLIENT_BENCHMARK_OVERRIDE:
        peer_benchmark = CLIENT_BENCHMARK_OVERRIDE[client_name]
    if not peer_benchmark:
        tab_weights = {}
        for m in managers:
            t = m.get('tab')
            if t:
                tab_weights[t] = tab_weights.get(t, 0) + float(m.get('current_weight', 0) or 0)
        peer_benchmark = max(tab_weights, key=tab_weights.get) if tab_weights else None
    if not peer_benchmark:
        return jsonify({'error': 'Could not resolve a benchmark peer group for this client.'})

    bench_meta = PEER_BENCHMARKS.get(peer_benchmark) or {}
    benchmark_name = bench_meta.get('core') or bench_str

    # ── 3. Build portfolio + benchmark return series ─────────────────────
    ret_mtx = build_return_matrix(managers, state['clone_results'])
    if ret_mtx.empty:
        return jsonify({'error': 'No return data for portfolio managers.'})

    bench_df = build_benchmark_series(peer_benchmark, factor_df, ret_mtx.index)
    if bench_df is None or 'core' not in bench_df.columns:
        return jsonify({'error': f'Benchmark series unavailable for peer group {peer_benchmark}.'})

    weights_dict = {m['matched_name']: float(m.get('current_weight', 0) or 0) for m in managers}
    port_rets = portfolio_return_series(ret_mtx, weights_dict).dropna()
    bench_core = bench_df['core'].dropna()
    # "Clone" series = portfolio return constructed from each manager's static
    # clone (not the manager's actual returns). Available when static_clone_full
    # is present in clone_results — same pattern as build_return_matrix's
    # clone-fallback path, but isolated to the clone-only return.
    clone_rets = _build_clone_return_series(managers, state['clone_results'], weights_dict)

    perf_periods = compute_trailing_periods(port_rets, bench_core, clone_rets)
    perf_calendar = compute_calendar_years(port_rets, bench_core, clone_rets, n_years=5)
    perf_quarterly = compute_quarterly_excess(port_rets, bench_core, clone_rets, n_quarters=21)
    inception = find_inception_date(port_rets) or '—'

    # ── 4. Manager-side ideal complement (reuse the same logic) ──────────
    excess = (port_rets - bench_core).dropna()
    under_idx = excess[excess < 0].index
    manager_complement = _best_manager_complement(
        state['clone_results'], managers, under_idx, bench_core, min_overlap=12,
    )

    # ── 5. Aapryl-factor and FactSet-risk complements ────────────────────
    aapryl_factors = ['Value', 'Growth', 'Yield', 'Quality', 'Defensive',
                      'Low Vol', 'Dynamic', 'Momentum']
    aapryl_complement = compute_factor_complements(
        port_rets, bench_core, factor_df, aapryl_factors,
        min_overlap=12, category_label='Aapryl Style Factor',
    )
    factset_factors = ['Beta', 'Book to Price', 'Dividend Yield', 'Earnings Yield',
                       'Growth', 'Leverage', 'Liquidity', 'Momentum',
                       'Profitability', 'Size', 'Volatility']
    factset_complement = compute_factor_complements(
        port_rets, bench_core, factor_df, factset_factors,
        min_overlap=12, category_label='FactSet Global Risk Factor',
    )

    n_underperf = int(len(under_idx))

    # ── 6. Holdings + portfolio V-G (pulled from build_portfolio_view) ───
    holdings_payload = []
    total_weight = sum(float(m.get('current_weight', 0) or 0) for m in managers)
    for m in managers:
        ns = _norm_skill_for(m.get('tab'), m.get('matched_name')) or {}
        holdings_payload.append({
            'name':       m.get('matched_name'),
            'tab':        m.get('tab'),
            'weight':     float(m.get('current_weight', 0) or 0),
            'vg_3factor': float(m.get('vg_3factor', 0) or 0),
            'vg_full':    float(m.get('vg_full', 0) or 0),
            'ns_z':       ns.get('z'),
        })

    # Portfolio-level V-G (weighted average over current weights)
    def _wavg(field):
        return (sum(float(m.get(field, 0) or 0) * float(m.get('current_weight', 0) or 0)
                    for m in managers)
                / total_weight) if total_weight > 0 else 0.0
    portfolio_vg = {
        'vg_3factor': _wavg('vg_3factor'),
        'vg_full':    _wavg('vg_full'),
    }

    return jsonify({
        'client':           client_name,
        'benchmark':        bench_str or benchmark_name,
        'benchmark_name':   benchmark_name,
        'peer_benchmark':   peer_benchmark,
        'as_of':            str(port_rets.index.max().date()) if not port_rets.empty else None,
        'inception_date':   inception,
        'n_underperf_months': n_underperf,
        'managers':         holdings_payload,
        'portfolio_vg':     portfolio_vg,
        'perf_backtested': {
            'name':             'Current Portfolio (Backtested)',
            'inception_date':   inception,
            'periods':          perf_periods,
            'calendar':         perf_calendar,
            'quarterly_excess': perf_quarterly,
        },
        'complements_backtested': {
            'benchmark_name':       benchmark_name,
            'n_underperf_months':   n_underperf,
            'manager':              manager_complement,
            'aapryl_factor':        aapryl_complement,
            'factset_risk':         factset_complement,
        },
    })


def _build_clone_return_series(managers, clone_results, weights_dict):
    """Synthesize a portfolio return series from each manager's static clone
    (vs their actual returns). Used in the report to surface 'Portfolio Clone'
    next to the actual portfolio backtest."""
    import numpy as np
    import pandas as pd
    series_list = []
    for m in managers:
        tab = m.get('tab')
        name = m.get('matched_name')
        d = (clone_results.get(tab) or {}).get(name, {})
        dates = d.get('dates') or []
        clone = d.get('static_clone_full') or []
        n = min(len(dates), len(clone))
        if n == 0:
            continue
        idx = pd.DatetimeIndex(pd.to_datetime(dates[:n]))
        vals = [np.nan if v is None or (isinstance(v, float) and np.isnan(v)) else float(v)
                for v in clone[:n]]
        s = pd.Series(vals, index=idx, name=name, dtype=float)
        s = s[~s.index.duplicated(keep='first')]
        series_list.append(s)
    if not series_list:
        return pd.Series(dtype=float)
    df = pd.concat(series_list, axis=1).sort_index()
    weights = np.array([weights_dict.get(c, 0) for c in df.columns], dtype=float)
    s = (df.values * weights).sum(axis=1)
    # Renormalize per-row by the sum of weights of the non-NaN columns so
    # months where some clones are missing don't drag the portfolio down.
    mask = (~np.isnan(df.values)).astype(float)
    denom = (mask * weights).sum(axis=1)
    denom = np.where(denom > 0, denom, np.nan)
    s = np.nansum(df.values * weights, axis=1) / denom
    return pd.Series(s, index=df.index, dtype=float).dropna()


def _best_manager_complement(clone_results, held_managers, under_idx, bench_core, min_overlap=12):
    """Find the manager with the highest hit rate on portfolio-underperformance
    months. Restricted to managers from peer-group tabs the portfolio already
    holds, and excludes managers already in the portfolio."""
    import numpy as np
    import pandas as pd
    if len(under_idx) < min_overlap:
        return None
    held_names = {m.get('matched_name') for m in held_managers}
    eligible_tabs = {m.get('tab') for m in held_managers if m.get('tab')}

    best = None
    for tab in eligible_tabs:
        for cname, cdata in (clone_results.get(tab, {}) or {}).items():
            if cname in held_names:
                continue
            dates = cdata.get('dates') or []
            mrets = cdata.get('manager_returns') or []
            n = min(len(dates), len(mrets))
            if n == 0:
                continue
            try:
                idx = pd.DatetimeIndex(pd.to_datetime(dates[:n]))
            except Exception:
                continue
            cs = pd.Series(
                [float(v) if v is not None and not (isinstance(v, float) and np.isnan(v)) else np.nan
                 for v in mrets[:n]],
                index=idx, dtype=float,
            ).dropna()
            cs = cs[~cs.index.duplicated(keep='first')]
            common = cs.index.intersection(under_idx)
            if len(common) < min_overlap:
                continue
            cand_on = cs.loc[common]
            bench_on = bench_core.reindex(common).dropna()
            common = cand_on.index.intersection(bench_on.index)
            if len(common) < min_overlap:
                continue
            diff = cand_on.loc[common] - bench_on.loc[common]
            hit = float((diff > 0).sum()) / float(len(diff))
            avg = float(diff.mean())
            ns = _norm_skill_for(tab, cname) or {}
            rec = {
                'name':       cname,
                'tab':        tab,
                'vg_3factor': cdata.get('vg_3factor', 0),
                'vg_full':    cdata.get('vg_full', 0),
                'ns_z':       ns.get('z'),
                'hit_rate':   hit,
                'avg_excess': avg,
                'n_months':   int(len(diff)),
            }
            if best is None or (rec['hit_rate'], rec['avg_excess']) > (best['hit_rate'], best['avg_excess']):
                best = rec
    return best


# ── Market Cycle chart endpoint ───────────────────────────────────────────
# Map each peer tab to its core benchmark for downside-capture comparisons.
MARKET_CYCLE_BENCH = {
    'EAFE': 'MSCI EAFE NR USD',
    'ACWI': 'MSCI ACWI NR USD',
    'ISC':  'MSCI ACWI Ex USA Small NR USD',
    'EM':   'MSCI EM NR USD',
    'US':   'Russell 1000 TR USD',
    'USSC': 'Russell 2000 TR USD',
}
# Temporary per-client override, matches risk_engine.CLIENT_BENCHMARK_OVERRIDE
CLIENT_CYCLE_BENCH = {
    'Client 1': 'MSCI EAFE Small Cap NR USD',
    'Client 2': 'MSCI EAFE NR USD',
}


@app.route('/reset_universe', methods=['POST'])
def reset_universe():
    """Wipe all universe clone results and normalised-skill data from state
    and cache. Used when you want to re-run universe clones from scratch
    (e.g. after uploading a new universe file with different managers or a
    new factor file). Buy-list clone results are NOT affected."""
    if state.get('running'):
        return jsonify({'status': 'error',
                        'message': 'A run is in progress — wait for it to finish first.'})
    state['universe_clone_results'] = None
    state['universe_dfs']           = None
    state['norm_skill_by_tab']      = {}
    # Remove staged universe file paths so the user must re-upload
    if isinstance(state['files'].get('universe_returns'), dict):
        state['files'].pop('universe_returns', None)
    save_cache()
    cb("Universe clone results and normalized-skill data cleared.")
    return jsonify({'status': 'ok',
                    'message': 'Universe clones and normalized skill cleared. '
                               'Re-upload your universe file and click Run Universe Clones.'})



@app.route('/debug_portfolio_ns/<client_name>/<path:wt_name>')
def debug_portfolio_ns(client_name, wt_name):
    """Diagnostic: trace exactly how a weight-file manager name resolves
    through fuzzy_match → clone_results → norm_skill_by_tab.
    Open in browser: http://localhost:5050/debug_portfolio_ns/Client%201/EAM%20EAFE%20SC
    """
    from data_loader import fuzzy_match
    out = {'wt_name': wt_name, 'client': client_name}

    # 1. fuzzy_match result
    tab, mgr_name = fuzzy_match(wt_name, state['manager_dfs'])
    out['fuzzy_tab']       = tab
    out['fuzzy_name']      = mgr_name
    out['fuzzy_name_repr'] = repr(mgr_name)

    # 2. clone_results lookup
    cr_tab  = (state['clone_results'] or {}).get(tab or '', {})
    key_norm = (mgr_name or '').strip().lower()
    ci_cr   = next((k for k in cr_tab  if k.strip().lower() == key_norm), None)
    out['in_clone_results_exact'] = mgr_name in cr_tab
    out['in_clone_results_ci']    = ci_cr
    out['cr_eam_keys'] = [repr(k) for k in cr_tab  if 'eam' in str(k).lower()]

    # 3. universe lookup
    ucr_tab = (state.get('universe_clone_results') or {}).get(tab or '', {})
    ci_ucr  = next((k for k in ucr_tab if k.strip().lower() == key_norm), None)
    out['in_ucr_exact'] = mgr_name in ucr_tab
    out['in_ucr_ci']    = ci_ucr

    # 4. resolved exact_key and data source
    d, exact_key = None, mgr_name
    if tab:
        if mgr_name in cr_tab:
            d, exact_key = cr_tab[mgr_name], mgr_name
        elif ci_cr:
            d, exact_key = cr_tab[ci_cr], ci_cr
        elif mgr_name in ucr_tab:
            d, exact_key = ucr_tab[mgr_name], mgr_name
        elif ci_ucr:
            d, exact_key = ucr_tab[ci_ucr], ci_ucr
    out['exact_key']      = exact_key
    out['exact_key_repr'] = repr(exact_key)
    out['data_source']    = ('clone_results' if (mgr_name in cr_tab or ci_cr)
                             else 'universe' if d else 'not_found')
    out['r2_full'] = d.get('r2_full') if d else None

    # 5. norm_skill lookup
    ns_tab  = (state.get('norm_skill_by_tab') or {}).get(tab or '', {})
    ek_norm = (exact_key or '').strip().lower()
    ci_ns   = next((k for k in ns_tab if k.strip().lower() == ek_norm), None)
    out['ns_exact_match']      = exact_key in ns_tab
    out['ns_ci_match']         = ci_ns
    out['ns_ci_match_repr']    = repr(ci_ns)
    rec = ns_tab.get(exact_key) or (ns_tab.get(ci_ns) if ci_ns else None)
    out['ns_z']    = rec.get('z')    if rec else None
    out['ns_n_obs'] = rec.get('n_obs') if rec else None
    out['ns_eam_keys'] = [repr(k) for k in ns_tab if 'eam' in str(k).lower()]

    return jsonify(out)


def universe_status():
    """Tell the UI which peer tabs have universe data loaded."""
    ucr = state.get('universe_clone_results') or {}
    tabs_with_data = {tab: len(mgrs) for tab, mgrs in ucr.items() if mgrs}
    return jsonify({
        'tabs':            tabs_with_data,
        'managers_by_tab': tabs_with_data,   # alias consumed by runUniverseClones confirm
    })


@app.route('/universe_files')
def universe_files():
    """Tell the UI which peer-tab universe files have been staged on the
    server (whether or not they have cached clone results yet). Used to
    enable the 'Run Universe Clones' button on the Setup tab.
    """
    uni_map = state.get('files', {}).get('universe_returns', {})
    if not isinstance(uni_map, dict):
        uni_map = {}
    # Present the consolidated file as 'consolidated' (not '__all__') for UX
    tabs = sorted(['consolidated' if k == '__all__' else k for k in uni_map.keys()])
    return jsonify({'tabs': tabs})


def _start_universe_run():
    """Guard, then launch the universe clone in a background thread.

    Shared by the /run_universe route and the startup auto-run hook. Returns a
    (status, message|tabs) tuple describing what happened; it never raises.
    """
    uni_map = state.get('files', {}).get('universe_returns', {})
    if not isinstance(uni_map, dict) or not uni_map:
        return ('error', 'No universe files staged. Upload at least one universe file first.')
    if 'factor_returns' not in state['files']:
        return ('error', 'Factor returns not loaded.')
    if state['running']:
        return ('error', 'Another run is in progress')
    reset_progress(phase='Preparing universe clone run', total=0)
    state.update({'running': True, 'error': None})

    def worker():
        try:
            from data_loader import (load_universe_returns,
                                     load_universe_returns_consolidated,
                                     load_factor_returns, run_universe_cloning)
            if state.get('factor_df') is None:
                set_progress_phase('Loading factor returns')
                state['factor_df'] = load_factor_returns(state['files']['factor_returns'])
            if state['universe_clone_results'] is None:
                state['universe_clone_results'] = {}
            if state['universe_dfs'] is None:
                state['universe_dfs'] = {}

            # Pre-load everything first so we know the grand total up front.
            # This makes the progress bar meaningful from the first manager;
            # otherwise we'd only know per-tab totals as they streamed in.
            set_progress_phase('Loading universe files')
            loaded_tabs = []   # list of (peer_tab, tab_df) in run order
            consolidated_path = uni_map.get('__all__')
            if consolidated_path and os.path.exists(consolidated_path):
                cb("Loading consolidated universe file...")
                cons_map = load_universe_returns_consolidated(consolidated_path)
                if not cons_map:
                    cb("WARNING: consolidated file had no recognised sheets")
                else:
                    cb(f"Parsed {len(cons_map)} peer tab(s): {sorted(cons_map.keys())}")
                    for peer_tab in sorted(cons_map.keys()):
                        loaded_tabs.append((peer_tab, cons_map[peer_tab]))

            for peer_tab, p in sorted(uni_map.items()):
                if peer_tab == '__all__':
                    continue
                cb(f"Loading universe file for {peer_tab}...")
                if not os.path.exists(p):
                    cb(f"WARNING: file missing for {peer_tab}: {p}")
                    continue
                one_map = load_universe_returns(p, peer_tab)
                if not one_map or peer_tab not in one_map:
                    cb(f"WARNING: could not parse universe file for {peer_tab}")
                    continue
                loaded_tabs.append((peer_tab, one_map[peer_tab]))

            if not loaded_tabs:
                raise RuntimeError("No universe data could be loaded from staged files.")

            grand_total = sum(df.shape[1] for _, df in loaded_tabs)
            set_progress_phase('Cloning universe managers',
                               total=grand_total, reset_current=True)
            state['_progress_phase_base'] = 0
            cb(f"Starting clones: {grand_total} managers across {len(loaded_tabs)} tab(s).")

            for peer_tab, tab_df in loaded_tabs:
                set_progress_phase(f'Cloning {peer_tab} ({tab_df.shape[1]} managers)',
                                   total=grand_total)
                already = len((state['universe_clone_results'] or {}).get(peer_tab) or {})
                if already:
                    cb(f"Resuming {peer_tab}: {already} already done, "
                       f"{tab_df.shape[1] - already} remaining...")
                else:
                    cb(f"Running clones for {tab_df.shape[1]} {peer_tab} universe managers...")

                def make_checkpoint(ptab):
                    """Return a closure that merges a partial tab result into
                    state and flushes to disk every 50 managers."""
                    def _checkpoint(tab, partial):
                        if state['universe_clone_results'] is None:
                            state['universe_clone_results'] = {}
                        state['universe_clone_results'][tab] = partial
                        save_cache()
                        cb(f"  checkpoint: {len(partial)} {tab} managers saved.")
                    return _checkpoint

                new_results = run_universe_cloning(
                    {peer_tab: tab_df},
                    state['factor_df'],
                    progress_callback=cb_counting,
                    checkpoint_callback=make_checkpoint(peer_tab),
                    checkpoint_every=50,
                    existing_results=state.get('universe_clone_results'),
                )
                state['universe_clone_results'].update(new_results)
                state['universe_dfs'][peer_tab] = tab_df
                absorb_regional_into_core({peer_tab: state['universe_clone_results'].get(peer_tab, {})})
                set_progress_phase(f'Computing normalized skill ({peer_tab})',
                                   total=grand_total)
                recompute_norm_skill(tabs=[peer_tab])
                # Precompute the market-cycle universe state for this tab
                # while we're already paying the load — eliminates the
                # multi-second 'first /market_cycle call after restart' lag.
                try:
                    from market_cycle import precompute_universe_state
                    from risk_engine import PEER_BENCHMARKS as _PB
                    bench_map = {t: _PB.get(t, {}).get('core') for t in _PB}
                    if state.get('factor_df') is None:
                        state['factor_df'] = load_factor_returns(state['files']['factor_returns'])
                    if 'mc_universe_cache' not in state or state['mc_universe_cache'] is None:
                        state['mc_universe_cache'] = {}
                    us = precompute_universe_state(
                        state['universe_clone_results'], peer_tab,
                        state['factor_df'], bench_map,
                    )
                    if us is not None:
                        state['mc_universe_cache'][peer_tab] = us
                        cb(f"  market-cycle universe state cached for {peer_tab}.")
                except Exception as e:
                    cb(f"  market-cycle precompute failed for {peer_tab}: {e}")
                save_cache()
                advance_progress_base()

            set_progress_phase('Done', total=grand_total)
            state['progress_current'] = grand_total
            cb("DONE")
        except Exception as e:
            import traceback
            state['error'] = str(e) + "\n" + traceback.format_exc()
            cb(f"ERROR: {e}")
        finally:
            state['running'] = False

    threading.Thread(target=worker, daemon=True).start()
    return ('started', list(uni_map.keys()))


@app.route('/run_universe', methods=['POST'])
def run_universe():
    """Re-run the clone engine against every staged universe file. Each peer
    tab is loaded and cloned in turn; results merge into universe_clone_results
    per-tab (so a partial failure on one tab doesn't wipe others).
    """
    status, payload = _start_universe_run()
    if status == 'error':
        return jsonify({'status': 'error', 'message': payload})
    return jsonify({'status': 'started', 'tabs': payload})


@app.route('/market_cycle', methods=['POST'])
def market_cycle():
    """Classify current and proposed portfolio managers onto market-cycle buckets.
    Request JSON: { client_name?, managers: [{matched_name, tab, current_weight, proposed_weight}, ...] }
    Returns {current: [...], proposed: [...], peer_group, missing_universe: bool}.
    A manager appears in the 'current' array only if current_weight > 0, and
    likewise for 'proposed'. Classification is the same per manager regardless
    of which portfolio — this is just a convenience split for charting.
    """
    from market_cycle import classify_peer_group
    from data_loader import load_factor_returns
    from risk_engine import PEER_BENCHMARKS

    if not state['clone_results']:
        return jsonify({'error': 'Run the clone engine first.'})
    ucr = state.get('universe_clone_results') or {}
    if not ucr:
        return jsonify({'error': 'No universe data loaded — upload a universe file first.',
                        'missing_universe': True})

    if state.get('factor_df') is None:
        if 'factor_returns' not in state['files']:
            return jsonify({'error': 'Factor returns file missing.'})
        state['factor_df'] = load_factor_returns(state['files']['factor_returns'])

    payload = request.get_json(silent=True) or {}
    raw_managers = payload.get('managers', [])
    if not raw_managers:
        return jsonify({'error': 'No managers provided.'})

    # Group buy-list managers by peer tab. Market cycle classification is
    # peer-group-specific, so each tab's managers get classified against
    # their own universe.
    by_tab = {}
    for m in raw_managers:
        tab = m.get('tab')
        if not tab:
            continue
        # Pull the buy-list clone data to get dates/returns/betas/buckets
        d = state['clone_results'].get(tab, {}).get(m.get('matched_name') or m.get('name'), {})
        if not d:
            continue
        by_tab.setdefault(tab, []).append({
            'name':              m.get('matched_name') or m.get('name'),
            'tab':               tab,
            'current_weight':    float(m.get('current_weight', 0) or 0),
            'proposed_weight':   float(m.get('proposed_weight', 0) or 0),
            # Use posted style_buckets when provided (for client-side overrides),
            # otherwise fall back to the computed buckets in clone_results.
            'style_buckets':     m.get('style_buckets') or d.get('style_buckets', {}),
            'manager_returns':   d.get('manager_returns', []),
            'dates':             d.get('dates', []),
            'static_clone_full': d.get('static_clone_full', []),
            'betas_full':        d.get('betas_full', {}),
        })

    bench_map = {tab: PEER_BENCHMARKS.get(tab, {}).get('core') for tab in PEER_BENCHMARKS}
    # EAFE_SC isn't a peer tab in the buy list — it's a benchmark override —
    # so no mapping needed here for market cycle purposes.

    # The universe-side classification state is precomputed at the end of
    # each /run_universe tab and persisted in state['mc_universe_cache']
    # (saved/loaded with the cache pickle). Look it up here; only fall
    # back to on-demand compute if the entry is missing — that path is
    # for caches predating the precompute hook or for users who haven't
    # re-run universe since this code shipped.
    from market_cycle import precompute_universe_state
    if 'mc_universe_cache' not in state or state['mc_universe_cache'] is None:
        state['mc_universe_cache'] = {}
    mc_cache = state['mc_universe_cache']

    def _get_universe_state(tab):
        us = mc_cache.get(tab)
        if us is not None:
            return us
        # Lazy fallback: build it now and stash for next time. Save not
        # required immediately — the next save_cache() call (after a clone
        # run, weights reload, etc.) will persist it.
        us = precompute_universe_state(ucr, tab, state['factor_df'], bench_map)
        if us is not None:
            mc_cache[tab] = us
        return us

    all_placements = {}  # name -> placement dict (with tab, weights)
    tabs_without_universe = []
    for tab, mgrs in by_tab.items():
        if tab not in ucr or not ucr[tab]:
            tabs_without_universe.append(tab)
            continue
        try:
            placements = classify_peer_group(
                mgrs, ucr, state['factor_df'], bench_map,
                universe_state=_get_universe_state(tab),
            )
        except Exception as e:
            import traceback
            return jsonify({'error': f'Classification failed for {tab}: {e}',
                            'trace': traceback.format_exc()})
        for m in mgrs:
            p = placements.get(m['name'])
            if p:
                p = dict(p)  # copy
                p['name']            = m['name']
                p['tab']             = tab
                p['current_weight']  = m['current_weight']
                p['proposed_weight'] = m['proposed_weight']
                all_placements[m['name']] = p

    current_list  = [p for p in all_placements.values() if p['current_weight']  > 0]
    proposed_list = [p for p in all_placements.values() if p['proposed_weight'] > 0]

    return jsonify({
        'current':              current_list,
        'proposed':             proposed_list,
        'tabs_without_universe': tabs_without_universe,
        'n_total':              len(all_placements),
    })


# ── Exposures endpoints ──────────────────────────────────────────────────
@app.route('/upload_exposures', methods=['POST'])
def upload_exposures():
    """Accept a FactSet-style contribution XLSX, parse it, cache it."""
    from exposures_engine import parse_exposures_file
    f = request.files.get('exposures')
    if not f:
        return jsonify({'status': 'error', 'message': 'No file provided.'})
    fname = secure_filename(f.filename)
    path  = save_uploaded_file(f, fname, app.config['UPLOAD_FOLDER'])
    state['files']['exposures'] = path
    try:
        local = resolve_path(path, app.config['UPLOAD_FOLDER'], suffix='.xlsx')
        data = parse_exposures_file(local)
        state['exposures_data'] = data
        save_cache()
        return jsonify({
            'status':    'ok',
            'benchmark': data['benchmark_name'],
            'managers':  data['manager_names'],
            'n_benchmark': len(data['benchmark']),
        })
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'message': str(e),
                        'traceback': traceback.format_exc()})


# Per-client FactSet-exposures benchmark override. The exposures file is a
# snapshot of stocks and factor values across multiple candidate benchmarks
# and manager portfolios; this dict tells the tool which benchmark each
# client should be compared to. Falls back to the default (first-listed)
# benchmark in the file if the named one isn't present.
CLIENT_EXPOSURE_BENCHMARK = {
    'Client 1': 'MSCI EAFE Small Cap',
    'Client 2': 'MSCI EAFE',
    'Client 3': 'MSCI World',
}


@app.route('/portfolio_exposures', methods=['POST'])
def portfolio_exposures():
    """
    Compute portfolio exposure to a given grouping.
    POST JSON: {
        managers: [{matched_name, current_weight, proposed_weight}],
        grouping: str,
        client_name: str (optional — picks the client-specific benchmark),
        benchmark_name: str (optional — explicit override, trumps client_name)
    }
    """
    from exposures_engine import compute_portfolio_exposures, get_grouping_menu
    if state['exposures_data'] is None:
        return jsonify({'error': 'No exposure data loaded. Upload a FactSet file on the Setup tab.'})
    payload      = request.get_json(silent=True) or {}
    grouping     = payload.get('grouping')
    sub_grouping = payload.get('sub_grouping')
    managers     = payload.get('managers', [])
    client_name  = payload.get('client_name')
    bmk_name     = payload.get('benchmark_name')

    if not grouping:
        # Return the grouping menu + available benchmarks so the frontend can
        # build its selector and benchmark-picker.
        return jsonify({
            'menu':                 get_grouping_menu(),
            'available_benchmarks': list((state['exposures_data'].get('benchmarks') or {}).keys()),
            'default_benchmark':    state['exposures_data'].get('benchmark_name'),
        })
    if not managers:
        return jsonify({'error': 'No managers provided.'})

    # Resolve benchmark: explicit override > weights-file benchmark for the
    # client > hardcoded CLIENT_EXPOSURE_BENCHMARK > file default. The
    # weights-file name is used verbatim against exposures-file columns; if
    # it doesn't match, compute_portfolio_exposures falls back to the file
    # default and flags benchmark_fallback: true.
    if not bmk_name and client_name:
        bmk_name = (state.get('client_benchmarks') or {}).get(client_name)
    if not bmk_name and client_name:
        bmk_name = CLIENT_EXPOSURE_BENCHMARK.get(client_name)

    result = compute_portfolio_exposures(
        managers, state['exposures_data'], grouping,
        benchmark_name=bmk_name,
        sub_grouping=sub_grouping,
    )
    return jsonify(result)


# ─────────────────────────────────────────────────────────────────────────
#  Holdings-overlap matrix (Portfolio tab)
# ─────────────────────────────────────────────────────────────────────────
def _resolve_overlap_benchmark(client_name, bmk_name):
    """Same benchmark-resolution cascade the exposures tab uses, so the
    overlap matrix matches managers to the same exposure-file sleeves."""
    if not bmk_name and client_name:
        bmk_name = (state.get('client_benchmarks') or {}).get(client_name)
    if not bmk_name and client_name:
        bmk_name = CLIENT_EXPOSURE_BENCHMARK.get(client_name)
    return bmk_name


@app.route('/holdings_overlap', methods=['POST'])
def holdings_overlap():
    """
    Pairwise holdings-overlap matrix for the current or proposed portfolio.

    POST JSON: {
        managers: [{matched_name, weight_file_name?, current_weight, proposed_weight}],
        client_name:  str (optional — resolves the benchmark matching hint),
        weight_state: 'current' | 'proposed'  (default 'current'),
        benchmark_name: str (optional explicit override)
    }
    Returns the matrix (see overlap_engine.compute_holdings_overlap).
    """
    from overlap_engine import compute_holdings_overlap
    if state['exposures_data'] is None:
        return jsonify({'error': 'No exposure data loaded. Upload a FactSet '
                                 'exposures file on the Setup tab.'})
    payload      = request.get_json(silent=True) or {}
    managers     = payload.get('managers', [])
    client_name  = payload.get('client_name')
    weight_state = payload.get('weight_state', 'current')
    match_basis  = payload.get('match_basis', 'sedol')
    bmk_name     = _resolve_overlap_benchmark(client_name,
                                              payload.get('benchmark_name'))
    if not managers:
        return jsonify({'error': 'No managers provided.'})
    result = compute_holdings_overlap(
        managers, state['exposures_data'],
        benchmark_name=bmk_name, weight_state=weight_state,
        match_basis=match_basis,
    )
    return jsonify(result)


@app.route('/holdings_overlap_detail', methods=['POST'])
def holdings_overlap_detail():
    """
    Shared-security detail for one manager pair (drill-down panel).

    POST JSON: {
        managers: [...same as /holdings_overlap...],
        name_i: str, name_j: str,
        client_name: str (optional), weight_state: 'current'|'proposed',
        top_n: int (optional — cap the returned rows)
    }
    """
    from overlap_engine import compute_pair_detail
    if state['exposures_data'] is None:
        return jsonify({'error': 'No exposure data loaded.'})
    payload      = request.get_json(silent=True) or {}
    managers     = payload.get('managers', [])
    name_i       = payload.get('name_i')
    name_j       = payload.get('name_j')
    client_name  = payload.get('client_name')
    weight_state = payload.get('weight_state', 'current')
    match_basis  = payload.get('match_basis', 'sedol')
    top_n        = payload.get('top_n')
    bmk_name     = _resolve_overlap_benchmark(client_name,
                                              payload.get('benchmark_name'))
    if not (managers and name_i and name_j):
        return jsonify({'error': 'managers, name_i and name_j are required.'})
    result = compute_pair_detail(
        managers, state['exposures_data'], name_i, name_j,
        benchmark_name=bmk_name, weight_state=weight_state, top_n=top_n,
        match_basis=match_basis,
    )
    return jsonify(result)


@app.route('/optimize_portfolio', methods=['POST'])
def optimize_portfolio_endpoint():
    """Run a MILP optimization to find the buy-list portfolio that
    maximizes weighted norm skill subject to user constraints.

    Request JSON:
      {
        "client_name": "Client 1",            # used to resolve peer group if not given
        "peer_group": "EAFE",                 # optional; resolved from client if omitted
        "forced_managers": [
          {"name": "...", "tab": "EAFE", "weight": 0.15},
          ...
        ],
        # optional overrides — all default to the user's spec
        "min_weight": 0.05, "max_weight": 0.20,
        "min_managers": 4, "max_managers": 8,
        "vg_band": 0.07, "vg_center": 0.0
      }

    Candidates are restricted to the buy list of `peer_group` only.
    Forced managers may come from any peer-group tab on the buy list.
    """
    from portfolio_optimizer import optimize_portfolio
    from data_loader import resolve_peer_group
    from risk_engine import CLIENT_BENCHMARK_OVERRIDE

    if not state['clone_results']:
        return jsonify({'status': 'error', 'error': 'Run the clone engine first.'})

    payload = request.get_json(silent=True) or {}
    client_name = payload.get('client_name')
    peer_group  = payload.get('peer_group')
    forced      = payload.get('forced_managers') or []

    # ── Resolve peer group from client benchmark if not provided ─────────
    if not peer_group and client_name:
        bench = (state.get('client_benchmarks') or {}).get(client_name)
        peer_group = resolve_peer_group(bench) if bench else None
        if not peer_group and client_name in CLIENT_BENCHMARK_OVERRIDE:
            peer_group = CLIENT_BENCHMARK_OVERRIDE[client_name]

    # ── Map benchmark-side pseudo peer groups to actual buy-list tabs ────
    # resolve_peer_group may return 'EAFE_SC' or 'ACWI_xUS', which are used
    # for benchmark/risk routing but are NOT buy-list tabs. Translate:
    BENCH_TO_BUYLIST = {
        'EAFE_SC':  'ISC',   # MSCI EAFE SC / World ex-US SC → ISC buy list
        'ACWI_xUS': 'EAFE',  # MSCI ACWI ex-US → EAFE buy list (developed-ex-US managers)
    }
    if peer_group in BENCH_TO_BUYLIST:
        peer_group = BENCH_TO_BUYLIST[peer_group]

    if not peer_group:
        return jsonify({
            'status': 'error',
            'error': 'Cannot resolve peer group for the optimizer. '
                     'Specify peer_group explicitly or ensure the client has a benchmark.'
        })

    if peer_group not in (state.get('clone_results') or {}):
        return jsonify({
            'status': 'error',
            'error': f'No buy-list managers found for peer group "{peer_group}".'
        })

    # ── Optional constraint overrides (with sane fallbacks) ──────────────
    def _f(k, d):
        try:
            v = payload.get(k)
            return float(v) if v is not None else d
        except (TypeError, ValueError):
            return d
    def _i(k, d):
        try:
            v = payload.get(k)
            return int(v) if v is not None else d
        except (TypeError, ValueError):
            return d

    result = optimize_portfolio(
        peer_group=peer_group,
        forced_managers=forced,
        clone_results=state['clone_results'],
        norm_skill_by_tab=state.get('norm_skill_by_tab') or {},
        min_weight=_f('min_weight', 0.05),
        max_weight=_f('max_weight', 0.20),
        min_managers=_i('min_managers', 4),
        max_managers=_i('max_managers', 8),
        vg_band=_f('vg_band', 0.07),
        vg_center=_f('vg_center', 0.0),
    )

    # Surface the peer group used (in case it was resolved server-side)
    result['peer_group'] = peer_group
    return jsonify(result)


@app.route('/optimize_candidates/<peer_group>')
def optimize_candidates(peer_group):
    """List the buy-list managers eligible as forced inputs for the optimizer.
    Returns ALL buy-list managers across all peer-group tabs (so the user can
    force, e.g., an ISC manager into an EAFE optimization).

    Filtered to only include managers with a norm-skill score so the user
    sees what's actually usable. (Forced managers without a score are still
    permitted server-side, but we don't surface them in the picker.)
    """
    if not state['clone_results']:
        return jsonify({'candidates': []})
    ns = state.get('norm_skill_by_tab') or {}
    out = []
    for tab, mgrs in (state.get('clone_results') or {}).items():
        ns_tab = ns.get(tab, {}) or {}
        for name, d in mgrs.items():
            rec = {
                'name':      name,
                'tab':       tab,
                'vg_3factor': d.get('vg_3factor'),
                'vg_full':    d.get('vg_full'),
                'ns_z':       (ns_tab.get(name) or {}).get('z'),
                'r2_full':    d.get('r2_full'),
            }
            out.append(rec)
    # Sort: same peer group first (when peer_group matches), then by name
    out.sort(key=lambda r: (r['tab'] != peer_group, r['tab'], r['name'].lower()))
    return jsonify({'candidates': out, 'peer_group': peer_group})


# ─────────────────────────────────────────────────────────────────────────
#  Qualitative firm/strategy data — diverse/woman-owned ownership tracking
# ─────────────────────────────────────────────────────────────────────────
@app.route('/upload_qualitative', methods=['POST'])
def upload_qualitative():
    """Accept the layered firm/strategy qualitative XLSX, parse it, cache it."""
    from qualitative_loader import parse_qualitative_file
    f = request.files.get('qualitative')
    if not f:
        return jsonify({'status': 'error', 'message': 'No file provided.'})
    fname = secure_filename(f.filename)
    path  = save_uploaded_file(f, fname, app.config['UPLOAD_FOLDER'])
    state['files']['qualitative'] = path
    try:
        local = resolve_path(path, app.config['UPLOAD_FOLDER'], suffix='.xlsx')
        data = parse_qualitative_file(local)
        state['qualitative_data'] = data
        _QUAL_MATCH_CACHE.clear()
        save_cache()
        return jsonify({
            'status':       'ok',
            'n_firms':      data['n_firms'],
            'n_strategies': data['n_strategies'],
            'warnings':     data.get('warnings', []),
        })
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'message': str(e),
                        'traceback': traceback.format_exc()})


# ── Qualitative join: manager name → firm/strategy record ──────────────────
# Cache the fuzzy match per (qualitative-version, name) so the peer tables and
# portfolio view don't re-run rapidfuzz for every row on every render.
_QUAL_MATCH_CACHE = {}


def _qual_lookup(name, weight_file_name=None):
    """Return the qualitative record for a manager, matching its name (or the
    weights-file label) against the strategy keys in the qualitative file.
    Cascade: exact → whitespace/case-normalised → fuzzy (WRatio ≥ 86).
    Returns None if no confident match or no qualitative data loaded."""
    qd = state.get('qualitative_data')
    if not qd or not qd.get('strategies'):
        return None
    strategies = qd['strategies']

    for cand in (name, weight_file_name):
        if not cand:
            continue
        c = str(cand).strip()
        # 1 — exact
        if c in strategies:
            return strategies[c]
        # 2 — whitespace/case-normalised (handles 'Ativo ' vs 'Ativo',
        #     'Castleark' vs 'CastleArk')
        cl = c.lower().strip()
        for k, v in strategies.items():
            if str(k).lower().strip() == cl:
                return v

    # 3 — fuzzy fallback, guarded by a high cutoff so we don't mis-attach
    from rapidfuzz import process, fuzz
    keys = list(strategies.keys())
    ck = f"{id(qd)}::{name}::{weight_file_name}"
    if ck in _QUAL_MATCH_CACHE:
        hit = _QUAL_MATCH_CACHE[ck]
        return strategies.get(hit) if hit else None
    best = None
    for cand in (name, weight_file_name):
        if not cand:
            continue
        m = process.extractOne(str(cand), keys, scorer=fuzz.WRatio, score_cutoff=86)
        if m:
            best = m[0]
            break
    _QUAL_MATCH_CACHE[ck] = best
    return strategies.get(best) if best else None


def _qual_fields(name, weight_file_name=None):
    """Flat dict of qualitative fields for a manager, always the same shape
    (None-filled when unmatched) so the frontend can render uniformly."""
    rec = _qual_lookup(name, weight_file_name)
    if not rec:
        return {'q_firm': None, 'q_firm_aum': None, 'q_strategy_aum': None,
                'q_ownership': None, 'q_diverse_pct': None}
    return {
        'q_firm':         rec.get('firm'),
        'q_firm_aum':     rec.get('firm_aum'),
        'q_strategy_aum': rec.get('strategy_aum'),
        'q_ownership':    rec.get('ownership'),
        'q_diverse_pct':  rec.get('diverse_pct'),
    }


@app.route('/diverse_ownership', methods=['POST'])
def diverse_ownership():
    """
    Portfolio-level diverse/woman-owned exposure. Rolls managers up to firms
    (a client can hold several strategies from one firm), then answers:
    how much of the portfolio sits with firms whose diverse/woman ownership
    is >= threshold (default 50%), by weight and by firm count.
    """
    qd = state.get('qualitative_data')
    if not qd or not qd.get('strategies'):
        return jsonify({'has_data': False})
    payload   = request.get_json(silent=True) or {}
    managers  = payload.get('managers', [])
    threshold = float(payload.get('threshold', 50) or 50)

    def rollup(wkey):
        firms = {}
        unknown_w = 0.0
        total_w = 0.0
        for m in managers:
            w = float(m.get(wkey, 0) or 0)
            if w <= 0:
                continue
            total_w += w
            rec = _qual_lookup(m.get('matched_name'), m.get('weight_file_name'))
            if not rec or not rec.get('firm'):
                unknown_w += w
                continue
            f = rec['firm']
            if f not in firms:
                firms[f] = {'weight': 0.0, 'diverse_pct': rec.get('diverse_pct')}
            firms[f]['weight'] += w
        n_firms = len(firms)
        div_firms = [f for f, d in firms.items()
                     if d['diverse_pct'] is not None and d['diverse_pct'] >= threshold]
        div_w = sum(firms[f]['weight'] for f in div_firms)
        wpct = (100.0 * div_w / total_w) if total_w > 0 else 0.0
        return {
            'weight_pct':   round(wpct, 2),
            'n_diverse':    len(div_firms),
            'n_firms':      n_firms,
            'ratio_pct':    round(100.0 * len(div_firms) / n_firms, 1) if n_firms else 0.0,
            'unknown_weight_pct': round(100.0 * unknown_w / total_w, 1) if total_w > 0 else 0.0,
        }

    return jsonify({
        'has_data':  True,
        'threshold': threshold,
        'current':   rollup('current_weight'),
        'proposed':  rollup('proposed_weight'),
    })


# ─────────────────────────────────────────────────────────────────────────
#  PowerPoint "Print Memo Report" export (Portfolio tab)
# ─────────────────────────────────────────────────────────────────────────
@app.route('/export_portfolio_pptx', methods=['POST'])
def export_portfolio_pptx():
    """Build the 'Print Memo Report' PowerPoint deck for the current portfolio.

    Request JSON (assembled by the frontend from already-rendered panels):
      {
        "client_name":      str,
        "benchmark_name":   str | null,
        "images": {portfolio_managers, vg_positioning, factset_risk, mcr, market_cycle}
      }
    Each image is a base64 data-URL PNG (or null to skip that slide).

    Returns the binary .pptx with Content-Disposition set for download.
    If a section's data is missing the slide is skipped; the response header
    'X-Skipped-Slides' communicates that back to the UI.
    """
    from pptx_export import build_portfolio_pptx
    from flask import Response

    payload = request.get_json(silent=True) or {}
    images = payload.get('images') or {}
    if not any(images.get(k) for k in
               ('portfolio_managers', 'vg_positioning', 'factset_risk', 'mcr', 'market_cycle')):
        return jsonify({'status': 'error',
                        'error': 'No panel screenshots provided — '
                                 'select a client and wait for the page to fully render.'}), 400

    try:
        data, summary = build_portfolio_pptx(payload)
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'error': str(e),
                        'trace': traceback.format_exc()}), 500

    # Filename: {ClientName}_Memo_Report_{YYYY-MM-DD}.pptx
    from datetime import date
    client = (payload.get('client_name') or 'Portfolio').replace(' ', '_')
    safe   = ''.join(ch for ch in client if ch.isalnum() or ch in '_-')
    fname  = f"{safe}_Memo_Report_{date.today().isoformat()}.pptx"

    resp = Response(
        data,
        mimetype='application/vnd.openxmlformats-officedocument.presentationml.presentation',
    )
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    resp.headers['Content-Length']      = str(len(data))
    # Comma-separated list of section names that were omitted, if any
    if summary.get('skipped'):
        resp.headers['X-Skipped-Slides'] = '; '.join(summary['skipped'])
    return resp


def _auto_run_universe_on_startup():
    """If a universe file is staged on disk but its clones haven't been run yet,
    kick off the universe clone automatically in the background at startup — so
    the Market Cycle view populates without the user clicking 'Run Universe
    Clones'. The first boot pays the multi-minute compute once; results are
    cached to results.pkl, so every later boot sees cached results here and
    skips the run. Guards keep it from firing when it isn't needed or safe:
      • a universe file must actually be staged (and exist on disk),
      • factor returns must be loaded (the clone needs them),
      • no universe clone results may already be cached (else we'd redo work),
      • no run may already be in progress.
    Failures are swallowed — a bad auto-run must never stop the server booting.
    """
    try:
        uni_map = state.get('files', {}).get('universe_returns', {})
        if not isinstance(uni_map, dict) or not uni_map:
            return
        if not any(isinstance(p, str) and os.path.exists(p) for p in uni_map.values()):
            return
        if 'factor_returns' not in state.get('files', {}):
            return
        if state.get('universe_clone_results'):
            return  # already cloned — cached results are enough
        if state.get('running'):
            return
        print("Auto-running universe clones on startup (staged file, no cached results)...")
        status, payload = _start_universe_run()
        print(f"  universe auto-run: {status} ({payload})")
    except Exception as e:
        print(f"  universe auto-run skipped due to error: {e}")


# Fire the auto-run at import time so it happens under both `python run.py`
# (which does `from app import app`) and `python app.py`.
_auto_run_universe_on_startup()


if __name__ == '__main__':
    # Direct entry point. The canonical launcher is `python run.py` (port 3001
    # + dependency check + browser auto-open). This block only fires when
    # someone does `python app.py` directly — keep the port consistent so
    # behaviour matches.
    app.run(host='0.0.0.0', port=3001, debug=False)
