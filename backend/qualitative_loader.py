"""
Qualitative-data loader — parses the layered firm/strategy workbook the user
uploads alongside the buy-list returns.

Expected layout (single sheet), a header row followed by repeating blocks:

    Firm / Strategy         | Firm AUM ($mm) | Ownership            | Strategy AUM ($mm)
    Arga                    | 12500          | 100% employee-owned  |
      Arga                  |                |                      | 4200
      Arga ISC              |                |                      | 1800
      Arga xUS              |                |                      | 2100
    (blank)
    Ativo                   | 3100           | 60% minority-owned   |
      Ativo                 |                |                      | 1200
      ...

Firm rows are identified by having a Firm-AUM and/or Ownership value; strategy
rows have those blank and carry only a Strategy name (matching the returns-file
column header exactly) and that strategy's own AUM. Blank rows are ignored.

Returns
-------
{
    'firms': {
        firm_name: {'firm_aum': float|None, 'ownership': str|None,
                    'strategies': [strategy_name, ...]}
    },
    'strategies': {
        strategy_name: {'firm': firm_name, 'strategy_aum': float|None,
                        'firm_aum': float|None, 'ownership': str|None}
    },
    'columns': [extra column labels beyond the four core ones, in order],
    'n_firms': int, 'n_strategies': int,
    'warnings': [str, ...],
}

The 'strategies' dict is the primary lookup the app uses: given a portfolio
manager's name it fuzzy-matches to a strategy key and reads firm + AUM +
ownership in one shot.
"""

import re
import openpyxl


# Header synonyms → canonical field. Matching is case-insensitive and ignores
# punctuation / unit parentheticals so "Firm AUM ($mm)" and "firm aum" both hit.
_FIELD_SYNONYMS = {
    'firm_name':    ('firm / strategy', 'firm/strategy', 'firm', 'firm name',
                     'manager', 'strategy', 'name'),
    'firm_aum':     ('firm aum', 'firm aum ($mm)', 'firm aum $mm', 'total firm aum',
                     'firm assets'),
    'ownership':    ('ownership', 'ownership structure', 'firm ownership'),
    'diverse_pct':  ('diverse/female ownership', 'diverse female ownership',
                     'diverse/female ownership %', 'diverse female ownership pct',
                     'diverse ownership', 'female ownership', 'diverse/female',
                     'diverse female', 'women/minority ownership',
                     'minority/women ownership', 'diverse pct'),
    'strategy_aum': ('strategy aum', 'strategy aum ($mm)', 'strategy aum $mm',
                     'strat aum', 'product aum', 'strategy assets'),
}


def _norm_hdr(s):
    s = str(s or '').lower().strip()
    s = re.sub(r'\([^)]*\)', '', s)          # drop unit parens
    s = re.sub(r'[^a-z0-9 /%]+', ' ', s)     # keep slash + percent
    s = s.replace('%', '')                   # then drop the % sign itself
    return re.sub(r'\s+', ' ', s).strip()


def _norm_name(s):
    """Normalise a firm/strategy name for prefix matching: lower-cased,
    punctuation flattened to spaces, whitespace collapsed."""
    s = str(s or '').lower().strip()
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def match_firm(strategy_name, qd):
    """Return (firm_name, firm_record) for a strategy via PREFIX match, or
    (None, None).

    A strategy belongs to a firm when its normalised name equals, or begins with,
    the firm's normalised name at a whole-token boundary — so firm "Arga" claims
    "Arga ISC" and "Arga xUS" but never "Argatroban". When several firms could
    match, the LONGEST firm name wins, so a specific "Arga Global" entry beats
    the generic "Arga".

    This is the fallback for firm-level workbooks that carry no per-strategy
    rows (the newer one-row-per-firm format). The layered format still resolves
    through the exact strategy keys first — see _qual_lookup.
    """
    if not qd or not qd.get('firms'):
        return None, None
    sn = _norm_name(strategy_name)
    if not sn:
        return None, None
    best = None
    best_len = -1
    for firm, rec in qd['firms'].items():
        fn = _norm_name(firm)
        if not fn:
            continue
        if sn == fn or sn.startswith(fn + ' '):
            if len(fn) > best_len:
                best, best_len = (firm, rec), len(fn)
    return best if best else (None, None)


def match_description(manager_name, qd):
    """Return the Description-tab write-up for a manager, or None.

    Same prefix rule as match_firm: the manager's normalised name equals, or
    begins with, a Description-tab firm name at a whole-token boundary —
    'Empiric' matches 'Empiric'; 'IMC ACWI ex US' matches 'IMC'. Longest
    firm name wins when several match.
    """
    descs = (qd or {}).get('descriptions') or {}
    mn = _norm_name(manager_name)
    if not descs or not mn:
        return None
    best, best_len = None, -1
    for firm, text in descs.items():
        fn = _norm_name(firm)
        if not fn:
            continue
        if mn == fn or mn.startswith(fn + ' '):
            if len(fn) > best_len:
                best, best_len = text, len(fn)
    return best


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('$', '')
    if s == '' or s == '-':
        return None
    # tolerate a trailing unit like "12.5B" / "900M"
    m = re.match(r'^(-?\d+(?:\.\d+)?)\s*([bmk])?$', s, re.I)
    if m:
        val = float(m.group(1))
        mult = {'b': 1000.0, 'm': 1.0, 'k': 0.001}.get((m.group(2) or 'm').lower(), 1.0)
        return val * mult
    try:
        return float(s)
    except ValueError:
        return None


def parse_qualitative_file(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # Firm write-ups live on a separate "Description" tab: manager/firm name
    # in column A, prose description in column B. Used by the Word memo's
    # New Managers section.
    desc_rows = []
    for sn in wb.sheetnames:
        if _norm_hdr(sn) in ('description', 'descriptions'):
            desc_rows = list(wb[sn].iter_rows(values_only=True))
            break
    wb.close()

    descriptions = {}
    for row in desc_rows:
        if not row or row[0] is None or len(row) < 2 or row[1] is None:
            continue
        name, text = str(row[0]).strip(), str(row[1]).strip()
        if not name or not text:
            continue
        if _norm_hdr(text) == 'description':   # header row
            continue
        descriptions[name] = text

    warnings = []
    if not rows:
        return {'firms': {}, 'strategies': {}, 'columns': [],
                'n_firms': 0, 'n_strategies': 0,
                'descriptions': descriptions,
                'warnings': ['Sheet is empty.']}

    # ── Locate the header row (first row that maps a name col + at least one
    #    firm/strategy field). Scan the first ~10 rows to tolerate a title. ──
    header_idx = None
    col_map = {}   # column index → canonical field
    extra_cols = {}  # column index → original label (non-core columns)
    for r_idx in range(min(10, len(rows))):
        row = rows[r_idx]
        cand = {}
        cand_extra = {}
        for c_idx, val in enumerate(row):
            h = _norm_hdr(val)
            if not h:
                continue
            matched = None
            for field, syns in _FIELD_SYNONYMS.items():
                if h in syns:
                    matched = field
                    break
            if matched:
                # first column wins for a given field
                if matched not in cand.values():
                    cand[c_idx] = matched
            else:
                cand_extra[c_idx] = str(val).strip()
        # A valid header needs a name column and at least one of the firm/strategy fields
        if 'firm_name' in cand.values() and (
                'firm_aum' in cand.values() or 'ownership' in cand.values()
                or 'diverse_pct' in cand.values() or 'strategy_aum' in cand.values()):
            header_idx = r_idx
            col_map = cand
            extra_cols = cand_extra
            break

    if header_idx is None:
        return {'firms': {}, 'strategies': {}, 'columns': [],
                'n_firms': 0, 'n_strategies': 0,
                'descriptions': descriptions,
                'warnings': ['Could not find a header row. Expected columns like '
                             '"Firm / Strategy", "Firm AUM", "Ownership", '
                             '"Diverse/Female Ownership %", "Strategy AUM" '
                             'within the first 10 rows.']}

    # invert col_map: field → column index
    idx = {field: c for c, field in col_map.items()}
    name_c  = idx.get('firm_name')
    faum_c  = idx.get('firm_aum')
    own_c   = idx.get('ownership')
    div_c   = idx.get('diverse_pct')
    saum_c  = idx.get('strategy_aum')

    firms = {}
    strategies = {}
    current_firm = None
    orphan_strats = []

    for row in rows[header_idx + 1:]:
        name = row[name_c] if name_c is not None and name_c < len(row) else None
        name = str(name).strip() if name is not None else ''
        faum = _to_float(row[faum_c]) if faum_c is not None and faum_c < len(row) else None
        own  = row[own_c] if own_c is not None and own_c < len(row) else None
        own  = str(own).strip() if own not in (None, '') else None
        div  = _to_float(row[div_c]) if div_c is not None and div_c < len(row) else None
        saum = _to_float(row[saum_c]) if saum_c is not None and saum_c < len(row) else None

        if not name and faum is None and own is None and div is None and saum is None:
            continue  # blank separator row

        is_firm_row = (faum is not None) or (own is not None) or (div is not None)

        if is_firm_row:
            # Start a new firm block. (If the same firm name recurs, merge.)
            if name in firms:
                warnings.append(f'Firm "{name}" appears more than once — merged.')
                if faum is not None:
                    firms[name]['firm_aum'] = faum
                if own is not None:
                    firms[name]['ownership'] = own
                if div is not None:
                    firms[name]['diverse_pct'] = div
            else:
                firms[name] = {'firm_aum': faum, 'ownership': own,
                               'diverse_pct': div, 'strategies': []}
            current_firm = name
            # A firm row may ALSO carry a strategy AUM if the firm's flagship
            # strategy shares the firm name and the user put its AUM inline.
            if saum is not None:
                _add_strategy(strategies, firms, name, current_firm, saum,
                              extra_row=None)
        else:
            # Strategy row.
            if current_firm is None:
                orphan_strats.append(name)
                strategies[name] = {'firm': None, 'strategy_aum': saum,
                                    'firm_aum': None, 'ownership': None,
                                    'diverse_pct': None}
                continue
            _add_strategy(strategies, firms, name, current_firm, saum, extra_row=None)

    if orphan_strats:
        warnings.append('Strategy rows before any firm row (no firm assigned): '
                        + ', '.join(orphan_strats[:8])
                        + ('…' if len(orphan_strats) > 8 else ''))

    return {
        'firms':        firms,
        'strategies':   strategies,
        'columns':      list(extra_cols.values()),
        'n_firms':      len(firms),
        'n_strategies': len(strategies),
        'descriptions': descriptions,
        'warnings':     warnings,
    }


def _add_strategy(strategies, firms, strat_name, firm_name, strat_aum, extra_row):
    if not strat_name:
        return
    finfo = firms.get(firm_name, {})
    strategies[strat_name] = {
        'firm':         firm_name,
        'strategy_aum': strat_aum,
        'firm_aum':     finfo.get('firm_aum'),
        'ownership':    finfo.get('ownership'),
        'diverse_pct':  finfo.get('diverse_pct'),
    }
    if firm_name in firms and strat_name not in firms[firm_name]['strategies']:
        firms[firm_name]['strategies'].append(strat_name)
