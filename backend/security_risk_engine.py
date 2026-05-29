"""
security_risk_engine.py — Stock-level risk factor engine.

Validated against FactSet aggregates: Σ(wᵢ·factorᵢ) bottom-up from
Security-Level Risk DNA, subtracted from benchmark absolute exposures from
Risk Summary tab. Supports US / Non-US / EM regional sleeve breakdown.
"""
from __future__ import annotations
import re, math
from typing import Optional
import openpyxl

STYLE_FACTORS = [
    'Beta', 'Book to Price', 'Dividend Yield', 'Earnings Yield',
    'Growth', 'Leverage', 'Liquidity', 'Momentum',
    'Profitability', 'Size', 'Volatility',
]

# ── Country classification ────────────────────────────────────────────────
EAFE_COUNTRIES = frozenset([
    'Austria','Belgium','Denmark','Finland','France','Germany','Ireland',
    'Israel','Italy','Netherlands','Norway','Portugal','Spain','Sweden',
    'Switzerland','United Kingdom',
    'Australia','New Zealand','Japan','Hong Kong','Singapore',
])
_CANADA = frozenset(['Canada'])
_US     = frozenset(['United States'])
_EM     = frozenset([
    'Brazil','Chile','Colombia','Mexico','Peru','Czech Republic','Egypt',
    'Greece','Hungary','Kuwait','Poland','Qatar','Saudi Arabia',
    'South Africa','Turkey','United Arab Emirates','China','India',
    'Indonesia','South Korea','Malaysia','Philippines','Taiwan','Thailand',
])
EAFE_CANADA  = EAFE_COUNTRIES | _CANADA
WORLD        = EAFE_CANADA | _US
ALL_STANDARD = WORLD | _EM

SLEEVE_US    = 'US'
SLEEVE_NONUS = 'Non-US'
SLEEVE_EM    = 'EM'


def classify_country(country: str, has_em_sleeve: bool = False):
    """Return (region_label, flag_or_None) for routing a security into
    one of the sleeve buckets.

    `has_em_sleeve` describes the BREAKDOWN, not the current selection: it
    is True whenever the parent benchmark has an EM sleeve option (3-way
    US/Non-US Dev/EM split, e.g. ACWI / ACWI ex-US / ACWI ex-US SC) and
    False for 2-way US/Non-US splits (e.g. MSCI World). Passing the
    current sleeve here would silently route EM stocks into Non-US Dev
    when the user picks the Non-US Dev sleeve — that's the bug we're
    fixing. The classifier returns the *correct* destination sleeve
    regardless of which sleeve the user is currently filtering to.

    Routing rules:
      US country               -> US sleeve.
      EAFE Developed + Canada  -> Non-US Dev sleeve.
      Listed EM country        -> EM if 3-way; else lumped into Non-US.
      Unclassified non-US      -> EM if 3-way; else lumped into Non-US.
    """
    if not country or country.strip() in ('--', ''):
        return None, None
    c = country.strip()
    if c in _US:            return SLEEVE_US,    None
    if c in EAFE_COUNTRIES: return SLEEVE_NONUS, None
    if c in _CANADA:        return SLEEVE_NONUS, None
    if c in _EM:
        if has_em_sleeve:   return SLEEVE_EM, None
        return SLEEVE_NONUS, f'{c} is EM — included in Non-US (no EM sleeve)'
    flag = f'{c} not in standard classification — treated as {"EM" if has_em_sleeve else "Non-US"}'
    if has_em_sleeve:       return SLEEVE_EM, flag
    return SLEEVE_NONUS, flag


def _norm_bench(s):
    """Normalize a benchmark name for cross-file matching.
    Strips punctuation (`+ - / , .`) and collapses whitespace, then
    canonicalises 'Small Cap' → 'sc' so 'MSCI EAFE Small Cap' and 'MSCI
    EAFE SC' resolve to the same key. Used both for _SLEEVES lookups and
    for matching client benchmark strings against the available column
    list in the uploaded risk file."""
    t = re.sub(r'[\s\+\-/,\.]+', ' ', (s or '').lower()).strip()
    t = re.sub(r'\bsmall\s+cap\b', 'sc', t)
    return re.sub(r'\s+', ' ', t).strip()


# ── Sleeve options per benchmark ─────────────────────────────────────────
# (display_label, sleeve_constant, preferred_bench_key)
# IMPORTANT: keys MUST be in `_norm_bench` form so that any spelling
# variation in the weights file ('MSCI EAFE+CANADA', 'MSCI EAFE + Canada',
# 'MSCI EAFE+Canada', etc.) resolves to the same entry. Don't add a key
# with raw hyphens, plus signs, or 'Small Cap' — those won't match.
_SLEEVES = {
    'msci world': [
        ('Full Portfolio',                None,        'MSCI World'),
        ('US vs Russell 1000',            SLEEVE_US,   'Russell 1000'),
        ('Non-US vs MSCI EAFE + Canada',  SLEEVE_NONUS,'MSCI EAFE + Canada'),
    ],
    'msci acwi': [
        ('Full Portfolio',                   None,        'MSCI ACWI'),
        ('US vs Russell 1000',               SLEEVE_US,   'Russell 1000'),
        ('Non-US Dev vs MSCI EAFE + Canada', SLEEVE_NONUS,'MSCI EAFE + Canada'),
        ('EM vs MSCI EM',                    SLEEVE_EM,   'MSCI EM'),
    ],
    'msci acwi ex us': [
        ('Full Portfolio',                   None,        'MSCI ACWI ex-US'),
        ('Non-US Dev vs MSCI EAFE + Canada', SLEEVE_NONUS,'MSCI EAFE + Canada'),
        ('EM vs MSCI EM',                    SLEEVE_EM,   'MSCI EM'),
    ],
    # ACWI ex-US Small Cap variant. Same regional split logic as ACWI ex-US
    # but with SC-specific benchmark targets. The EM SC option only renders
    # if the benchmark column is actually present in the uploaded risk
    # file — get_sleeve_options filters by `avail_norm` — so users without
    # an 'MSCI EM SC' column will see Full + Non-US SC only.
    'msci acwi ex us sc': [
        ('Full Portfolio',                       None,         'MSCI ACWI ex-US SC'),
        ('Non-US Dev SC vs MSCI World ex-US SC', SLEEVE_NONUS, 'MSCI World ex US SC'),
        ('EM SC vs MSCI EM SC',                  SLEEVE_EM,    'MSCI EM SC'),
    ],
    'msci eafe canada': [('Full Portfolio', None, 'MSCI EAFE + Canada')],
    'msci eafe sc':     [('Full Portfolio', None, 'MSCI EAFE Small Cap')],
    'msci eafe':        [('Full Portfolio', None, 'MSCI EAFE')],
    'msci em':          [('Full Portfolio', None, 'MSCI EM')],
    'russell 1000':     [('Full Portfolio', None, 'Russell 1000')],
    'russell 2000':     [('Full Portfolio', None, 'Russell 2000')],
}


def get_sleeve_options(client_bench_str: str, available_benchmarks: list) -> list:
    """Return [{label, sleeve, bench}] for the client's benchmark, filtered
    to only options whose bench column is actually in the uploaded file."""
    key = _norm_bench(client_bench_str)
    options = _SLEEVES.get(key)
    if not options:
        options = [('Full Portfolio', None, client_bench_str or '')]
    avail_norm = {_norm_bench(b): b for b in available_benchmarks}
    result = []
    for label, sleeve, pref in options:
        actual = avail_norm.get(_norm_bench(pref))
        if actual:
            result.append({'label': label, 'sleeve': sleeve, 'bench': actual})
    if not result:
        result.append({'label': 'Full Portfolio', 'sleeve': None, 'bench': None})
    return result


# ── Parser ────────────────────────────────────────────────────────────────
def _coerce(v):
    if v is None or v == '': return None
    try:
        f = float(v)
        return f if math.isfinite(f) else None
    except (TypeError, ValueError):
        return None


def parse_security_risk_file(filepath: str) -> dict:
    """Parse a FactSet Security-Level Risk DNA workbook.

    Returns:
      managers              : {mgr_name: [{security, weight, factors{}, country}]}
      benchmarks            : {bench_name: {factor: float}}
      available_benchmarks  : [str]
      factors               : [str]   style factors found in the file
      as_of_date            : str | None
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    result = {'managers': {}, 'benchmarks': {}, 'available_benchmarks': [],
              'factors': [], 'as_of_date': None}

    if 'Security-Level Risk DNA' not in wb.sheetnames:
        wb.close(); return result

    ws   = wb['Security-Level Risk DNA']
    rows = list(ws.iter_rows(values_only=True))

    for row in rows[:6]:
        if row[0] and isinstance(row[0], str) and '-' in str(row[0]):
            result['as_of_date'] = str(row[0]).strip(); break

    header_row = rows[7]  # factor names
    factor_cols = {i: str(v).strip() for i, v in enumerate(header_row)
                   if v and str(v).strip() in STYLE_FACTORS}
    result['factors'] = [f for f in STYLE_FACTORS if f in factor_cols.values()]

    current = None
    for row in rows[9:]:
        if not row or row[0] is None: continue
        name = str(row[0]).strip()
        wt   = _coerce(row[1]) if len(row) > 1 else None
        if wt is not None and abs(wt - 100) < 0.5:
            current = name
            result['managers'].setdefault(current, [])
            continue
        if current is None or wt is None: continue
        country = (str(row[2]).strip()
                   if len(row) > 2 and row[2] and str(row[2]).strip() != '--'
                   else None)
        factors = {fname: _coerce(row[ci]) for ci, fname in factor_cols.items()
                   if ci < len(row) and _coerce(row[ci]) is not None}
        result['managers'][current].append(
            {'security': name, 'weight': wt, 'country': country, 'factors': factors})

    if 'Risk Summary' in wb.sheetnames:
        ws2   = wb['Risk Summary']
        rows2 = list(ws2.iter_rows(values_only=True))
        bhdrs = {i: str(v).strip() for i, v in enumerate(rows2[6]) if v and i >= 1}
        result['available_benchmarks'] = list(bhdrs.values())
        for b in bhdrs.values():
            result['benchmarks'][b] = {}
        in_style = False
        for row in rows2[10:]:
            if not row or row[0] is None: continue
            label = str(row[0]).strip()
            if label == 'Style':    in_style = True;  continue
            if label == 'Industry': in_style = False; continue
            if in_style and label in STYLE_FACTORS:
                for ci, bname in bhdrs.items():
                    v = _coerce(row[ci]) if ci < len(row) else None
                    if v is not None:
                        result['benchmarks'][bname][label] = v

    wb.close()
    return result


# ── Name matching ─────────────────────────────────────────────────────────
def _norm_name(s):
    """Normalise a manager name for matching across files. Strips
    regional descriptors ('EAFE', 'Canada', 'xUS', etc.) and generic
    style/wrapper words but PRESERVES the size sleeve marker (sc/lc/mc).
    Reason: previously stripping 'sc' and 'small cap' caused 'CastleArk
    EAFE + Canada SC' and 'CastleArk EAFE + Canada' to collapse to the
    same key — the LC sleeve then silently overwrote the SC sleeve in
    the lookup dict, swapping in the wrong portfolio's exposures and
    flipping sign on the Size factor for SC clients."""
    s = str(s).lower().strip()
    s = re.sub(r'\([^)]*\)', '', s)
    # Treat +/-/./_/,/space combos as separators (handles 'EAFE+Canada'
    # vs 'EAFE + Canada' uniformly).
    s = re.sub(r'[\./\-_,\+]+', ' ', s)
    # Canonicalise size markers BEFORE stripping anything else
    s = re.sub(r'\bsmall\s+cap\b', 'sc', s)
    s = re.sub(r'\blarge\s+cap\b', 'lc', s)
    s = re.sub(r'\bmid\s+cap\b',   'mc', s)
    # Compound region+size shorthand. 'ISC' = International Small Cap,
    # 'USSC' = US Small Cap — both encode the SC sleeve. Keeping them as
    # raw tokens caused 'Ballina ISC' to lose its size signal during
    # region stripping and silently match the LC sleeve in the risk file.
    s = re.sub(r'\bisc\b',  'sc', s)
    s = re.sub(r'\bussc\b', 'sc', s)
    # Strip regional / geographic descriptors only — never size words
    for region in ['international', 'eafe', 'acwi', 'em', 'us', 'usa',
                   'world', 'global', 'xus', 'ex', 'non', 'canada']:
        s = re.sub(r'\b' + re.escape(region) + r'\b', '', s)
    # Strip generic wrapper / style words
    for generic in ['composite', 'portfolio', 'strategy', 'fund',
                    'equity', 'equities', 'growth', 'value', 'blend', 'core']:
        s = re.sub(r'\b' + re.escape(generic) + r'\b', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def _match_mgr(weight_name: str, security_names: list) -> Optional[str]:
    """Map a weights-file manager name to the corresponding column in the
    security risk file. When several file rows share a normalised key
    (e.g. 'Hillsdale xUS SC' and 'Hillsdale EAFE + Canada SC' both norm
    to 'hillsdale sc'), break ties on full-string similarity to the
    weights name — picks the EAFE+Canada variant for 'Hillsdale EAFE SC'
    and the xUS variant for 'Hillsdale xUS SC'."""
    wn = _norm_name(weight_name)
    # Group candidates by normalised key — keep ALL collisions so we can
    # tiebreak rather than silently overwriting like the previous dict
    # comprehension did.
    by_norm: dict = {}
    for n in security_names:
        by_norm.setdefault(_norm_name(n), []).append(n)

    def _best_of(originals):
        if len(originals) == 1:
            return originals[0]
        # Tiebreak by original-word-token overlap. WRatio fuzzy didn't
        # always pick the right candidate when the normalised key collapses
        # multiple sleeves: 'Hillsdale EAFE Small Cap' (weights) vs
        # 'Hillsdale xUS SC' AND 'Hillsdale EAFE + Canada SC' (file) is the
        # canonical case. Counting how many original word-tokens each
        # candidate shares with the weights name picks the EAFE+Canada
        # variant cleanly because it shares 'eafe' with the weights name.
        wt_tokens = set(re.split(r'[\s\+\-\.,/]+', weight_name.lower()))
        wt_tokens.discard('')
        ranked = sorted(
            originals,
            key=lambda o: -len(set(re.split(r'[\s\+\-\.,/]+', o.lower())) & wt_tokens),
        )
        return ranked[0]

    if wn in by_norm:
        return _best_of(by_norm[wn])
    try:
        from rapidfuzz import process, fuzz
        m = process.extractOne(wn, list(by_norm.keys()),
                               scorer=fuzz.WRatio, score_cutoff=70)
        if m:
            return _best_of(by_norm[m[0]])
    except ImportError:
        pass
    return None


# ── Core computation ──────────────────────────────────────────────────────
def _mgr_abs(stocks: list, factors: list) -> dict:
    tw = sum(s['weight'] for s in stocks)
    if tw <= 0: return {}
    totals = {f: 0.0 for f in factors}
    for s in stocks:
        for f in factors:
            v = s['factors'].get(f)
            if v is not None: totals[f] += s['weight'] * v
    return {f: totals[f] / tw for f in factors}


def compute_exposures(
    portfolio_managers: list,
    security_data: dict,
    benchmark_name: Optional[str],
    sleeve: Optional[str] = None,
    has_em_sleeve: Optional[bool] = None,
) -> dict:
    """Compute current + proposed active style exposures.

    `has_em_sleeve` reflects the breakdown context (does the parent
    benchmark have a 3-way US/Non-US Dev/EM split?), not the user's
    current selection. When None, defaults to `sleeve == 'EM'` for
    backwards compat with old call sites — but new callers should pass
    it explicitly to get correct routing for Non-US Dev sleeves in 3-way
    breakdowns.

    Returns same schema as /compute_risk_exposures (manager-level path) so
    the existing renderRiskExposures() function works without changes.
    """
    factors    = security_data.get('factors') or STYLE_FACTORS
    mgr_data   = security_data.get('managers', {})
    bench_abs  = (security_data.get('benchmarks', {}).get(benchmark_name) or {}) \
                 if benchmark_name else {}
    if has_em_sleeve is None:
        has_em_sleeve = (sleeve == SLEEVE_EM)
    has_em     = has_em_sleeve
    sec_names  = list(mgr_data.keys())

    def compute_side(weight_key: str):
        totals      = {f: 0.0 for f in factors}
        sleeve_w    = 0.0
        # Portfolio-wide weight in stocks whose country tag is missing or
        # '--'. Surfaced separately in the response so the UI can explain
        # why summed sleeve coverage falls short of 100%. Computed even in
        # Full Portfolio mode (sleeve=None) because the user may want to
        # know the total no-country drag regardless of split.
        no_country_w = 0.0
        total_cw    = sum(max(m.get(weight_key, 0) or 0, 0) for m in portfolio_managers)
        if total_cw <= 0:
            return {f: None for f in factors}, [], 0.0, set(), 0.0

        unmatched = []
        flags     = set()

        for m in portfolio_managers:
            cw = m.get(weight_key, 0) or 0
            if cw <= 0: continue
            # Prefer weight_file_name for matching — it's the original
            # weights-file label ('Mac Alpha EAFE + Canada SC') and carries
            # sleeve-specific info that the buy-list matched_name often
            # loses (e.g. 'Mac Alpha' alone). matched_name remains the
            # canonical key for clone_results/norm_skill lookups elsewhere.
            match_input = (m.get('weight_file_name')
                           or m.get('matched_name')
                           or m.get('name', ''))
            mgr_key = _match_mgr(match_input, sec_names)
            if mgr_key is None:
                unmatched.append(m.get('matched_name') or m.get('name', '?'))
                continue
            all_stocks = mgr_data.get(mgr_key, [])
            if not all_stocks:
                continue

            mgr_total = sum(s['weight'] for s in all_stocks) or 1
            # No-country accounting (independent of sleeve filter)
            no_country_stocks_w = sum(
                s['weight'] for s in all_stocks
                if not s.get('country')
                or str(s.get('country')).strip() in ('--', '')
            )
            no_country_w += (cw / total_cw) * (no_country_stocks_w / mgr_total)

            if sleeve is not None:
                stocks = []
                for s in all_stocks:
                    region, flag = classify_country(s.get('country'), has_em)
                    if flag: flags.add(flag)
                    if region == sleeve: stocks.append(s)
            else:
                stocks = all_stocks

            if not stocks: continue

            if sleeve is not None:
                contrib   = (cw / total_cw) * (sum(s['weight'] for s in stocks) / mgr_total)
            else:
                contrib   = cw / total_cw

            sleeve_w += contrib
            ma = _mgr_abs(stocks, factors)
            for f in factors:
                if f in ma: totals[f] += contrib * ma[f]

        if sleeve_w <= 0:
            return {f: None for f in factors}, unmatched, 0.0, flags, no_country_w
        return {f: totals[f] / sleeve_w for f in factors}, unmatched, sleeve_w, flags, no_country_w

    cur_abs,  unc_cur,  cur_sw,  cur_flags,  cur_nc  = compute_side('current_weight')
    prop_abs, unc_prop, prop_sw, prop_flags, prop_nc = compute_side('proposed_weight')

    def to_active(abs_vals):
        return {f: round(abs_vals[f] - bench_abs[f], 4)
                if abs_vals.get(f) is not None and bench_abs.get(f) is not None
                else abs_vals.get(f)
                for f in factors}

    current  = to_active(cur_abs)
    proposed = to_active(prop_abs)
    delta    = {f: round(proposed[f] - current[f], 4)
                if proposed.get(f) is not None and current.get(f) is not None else None
                for f in factors}

    return {
        'factors':    factors,
        'current':    current,
        'proposed':   proposed,
        'delta':      delta,
        'unmatched':  sorted(set(unc_cur) | set(unc_prop)),
        'benchmark': {
            'matched_column':    benchmark_name,
            'requested':         benchmark_name,
            'fallback_absolute': not bool(bench_abs),
            'available_columns': security_data.get('available_benchmarks', []),
        },
        'sleeve_info': {
            'sleeve':         sleeve,
            'sleeve_wt':      round(cur_sw, 4),
            'sleeve_wt_proposed': round(prop_sw, 4),
            # Portfolio-wide weight in stocks whose country tag is missing
            # in the FactSet export (typically cash, ADRs, or derivatives).
            # The sum of sleeve coverages won't reach 100% by exactly this
            # amount; surfaced so the UI can render an explicit accounting
            # line for the user.
            'no_country_wt':           round(cur_nc, 4),
            'no_country_wt_proposed':  round(prop_nc, 4),
            'flags':     sorted(cur_flags | prop_flags),
        },
    }
