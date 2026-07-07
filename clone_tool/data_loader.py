"""
Data Loader - reads uploaded Excel files, runs cloning, returns structured results
"""
import hashlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import process, fuzz
import re
import os
import concurrent.futures
from clone_engine import (clone_fun, FACTOR_CATEGORIES, STYLE_BUCKET_MAP,
                          SMALL_CAP_FACTORS, EM_FACTORS, VALUE_FACTORS, GROWTH_FACTORS)

PEER_TABS = ['EAFE', 'ACWI', 'ISC', 'EM', 'US', 'USSC']


def _input_hash(*parts):
    """Stable fingerprint of the inputs that feed a single clone_fun call.
    Used by the determinism diagnostic to tell file drift apart from
    algorithmic non-determinism: if two runs produce different betas but
    the input hash is identical, the change is internal to the LASSO/QP
    pipeline (typically threaded-BLAS FP reordering flipping a near-zero
    LASSO coefficient in or out of the active set)."""
    h = hashlib.sha1()
    for p in parts:
        if p is None:
            h.update(b'\x00NONE\x00')
        elif isinstance(p, np.ndarray):
            arr = np.ascontiguousarray(p, dtype=np.float64)
            h.update(b'\x00ARR\x00' + str(arr.shape).encode() + b'\x00')
            h.update(arr.tobytes())
        elif isinstance(p, (list, tuple)):
            h.update(b'\x00LST\x00' + '|'.join(map(str, p)).encode() + b'\x00')
        else:
            h.update(b'\x00STR\x00' + str(p).encode() + b'\x00')
    return h.hexdigest()[:16]


def diagnose_clone_determinism(prev_results, new_results, beta_change_threshold=0.01):
    """Compare two clone_results dicts and produce a list of progress-log
    lines summarising what changed. Splits material beta changes into
    'inputs SAME' (algorithmic non-determinism — e.g. threaded BLAS) vs
    'inputs CHANGED' (file drift). Skipped silently when there is no prior
    run to compare against, or when prior results predate the input-hash
    addition (no _input_hash field).
    """
    if not prev_results:
        return ['Determinism diagnostic: no prior run cached - baseline established.']

    materially_changed = []
    algo_drift = []
    input_drift = []
    no_hash_count = 0
    only_in_prev = []
    only_in_new = []
    total_compared = 0

    all_tabs = sorted(set(prev_results) | set(new_results))
    for tab in all_tabs:
        prev_tab = prev_results.get(tab) or {}
        new_tab  = new_results.get(tab) or {}
        if not isinstance(prev_tab, dict) or not isinstance(new_tab, dict):
            continue
        for mgr in sorted(set(prev_tab) | set(new_tab)):
            if mgr not in new_tab:
                only_in_prev.append(f'{tab}/{mgr}')
                continue
            if mgr not in prev_tab:
                only_in_new.append(f'{tab}/{mgr}')
                continue
            total_compared += 1
            old, new = prev_tab[mgr], new_tab[mgr]
            max_db = 0.0
            for key in ('betas_full', 'betas_3factor'):
                old_b = old.get(key) or {}
                new_b = new.get(key) or {}
                for f in set(old_b) | set(new_b):
                    db = abs((new_b.get(f) or 0) - (old_b.get(f) or 0))
                    if db > max_db:
                        max_db = db
            if max_db > beta_change_threshold:
                old_h = old.get('_input_hash')
                new_h = new.get('_input_hash')
                if old_h is None or new_h is None:
                    tag = 'no-prior-hash'
                    no_hash_count += 1
                elif old_h == new_h:
                    tag = 'inputs SAME -> algo non-determinism'
                    algo_drift.append((tab, mgr, max_db))
                else:
                    tag = 'inputs CHANGED -> file drift'
                    input_drift.append((tab, mgr, max_db))
                materially_changed.append((tab, mgr, max_db, tag))

    lines = ['=== Clone determinism diagnostic ===',
             f'Compared {total_compared} managers vs prior run.']
    if materially_changed:
        lines.append(f'{len(materially_changed)} managers with max |dB| > {beta_change_threshold}:')
        for tab, mgr, db, tag in sorted(materially_changed, key=lambda x: -x[2])[:25]:
            lines.append(f'  {tab}/{mgr}: max |dB|={db:.4f}  [{tag}]')
        if len(materially_changed) > 25:
            lines.append(f'  ... and {len(materially_changed)-25} more')
    else:
        lines.append('No material beta changes - runs are stable.')
    if algo_drift:
        lines.append(f'{len(algo_drift)} with identical inputs (algorithmic non-determinism).')
    if input_drift:
        lines.append(f'{len(input_drift)} with changed inputs (file drift).')
    if no_hash_count:
        lines.append(f'{no_hash_count} prior results lacked input hashes - re-run once more for full attribution.')
    if only_in_prev:
        lines.append(f'{len(only_in_prev)} managers dropped vs prior run.')
    if only_in_new:
        lines.append(f'{len(only_in_new)} managers added vs prior run.')
    lines.append('=== end diagnostic ===')
    return lines


def _resolve_factors_ci(expected, actual_columns):
    """Map expected factor names → actual column names from the factor file,
    case- and whitespace-insensitive. Returns the list of actual column
    names that matched, in the order of `expected`. This guards against
    capitalization drift between the FACTOR_CATEGORIES list and FactSet
    exports (e.g. 'EAFe SC Div Yld' vs 'EAFE SC Div Yld') that would
    otherwise cause silent factor dropouts."""
    def _key(s):
        return ' '.join(str(s).split()).lower()
    actual_by_key = {_key(c): c for c in actual_columns}
    out = []
    for f in expected:
        col = actual_by_key.get(_key(f))
        if col is not None:
            out.append(col)
    return out


def _headline_r2(r2_array):
    """Extract the most-recent rolling-window R² from a clone_fun result.

    clone_fun returns R² in MOST-RECENT-FIRST order (the array is reversed
    on the way out — see clone_engine.py line 120). The R reference workflow
    reports `output_sta$R2[1,1]` which is row 1 = first row = most recent,
    so we do the same: return the first non-NaN value.

    A previous version of the code computed `np.nanmean(r2s[-12:])` AFTER
    filtering NaNs, which silently grabbed the OLDEST 12 R²s — those come
    from the earliest 24-month rolling windows and are systematically
    lower than the headline. This caused managers like CastleArk EAFE SC
    to display R²=0.75 instead of the correct 0.87 even when the clone
    itself was right."""
    if r2_array is None or len(r2_array) == 0:
        return None
    for v in r2_array:
        if not np.isnan(v):
            return float(v)
    return None


def load_factor_returns(filepath):
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    factor_names = [str(f).strip() if f else None for f in rows[0][1:]]
    data_rows = []
    for row in rows[4:]:
        if row[0] is None:
            continue
        date_str = str(row[0])
        if ' - ' in date_str:
            date_str = date_str.split(' - ')[1].strip()
        try:
            date = pd.to_datetime(date_str) + pd.offsets.MonthEnd(0)
        except:
            continue
        def sf(v):
            if v is None: return np.nan
            try: return float(v) / 100.0
            except: return np.nan
        values = [sf(v) for v in row[1:len(factor_names)+1]]
        data_rows.append([date] + values)
    df = pd.DataFrame(data_rows, columns=['Date'] + factor_names)
    df = df.set_index('Date').sort_index(ascending=False)
    return df


def load_manager_returns(filepath):
    """Load a buy-list manager-returns workbook.

    Each sheet is a peer group; column A is the date, remaining columns are
    manager return series. We accept several sheet-naming conventions:
      - The internal peer-tab codes themselves: 'EAFE', 'ACWI', 'ISC', 'EM',
        'US', 'USSC'
      - Display variants matching the universe-file convention: 'US LC' →
        'US', 'US SC' → 'USSC', 'Global' → 'ACWI'
    Sheets whose names don't map (e.g. 'Map') are ignored.
    """
    # Sheet-name aliases — display strings the user might write, mapped to
    # our internal peer-tab codes. Kept in lock-step with the universe-file
    # mapping (_UNIVERSE_SHEET_TO_TAB) so buy-list and universe sheet names
    # can be identical.
    sheet_aliases = {
        'US LC': 'US',
        'US SC': 'USSC',
        'Global': 'ACWI',
    }

    wb = load_workbook(filepath, read_only=True)
    result = {}
    # Build an explicit list of (sheet_name_in_file, peer_tab) pairs we
    # actually want to read. Skip duplicates if the file contains both
    # 'US' and 'US LC' — first match wins, with PEER_TABS prioritised.
    sheets_to_read = []
    seen_tabs = set()
    for sheet_name in wb.sheetnames:
        # Direct match on a peer-tab code
        if sheet_name in PEER_TABS:
            tab = sheet_name
        # Alias match
        elif sheet_name in sheet_aliases:
            tab = sheet_aliases[sheet_name]
        else:
            continue
        if tab in seen_tabs:
            continue
        seen_tabs.add(tab)
        sheets_to_read.append((sheet_name, tab))

    for sheet_name, peer_tab in sheets_to_read:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        managers = [str(m).strip() if m else None for m in rows[0][1:]]
        managers = [m for m in managers if m]
        data_rows = []
        for row in rows[1:]:
            if row[0] is None:
                continue
            try:
                date = pd.to_datetime(row[0])
            except:
                continue
            def sfm(v):
                if v is None: return np.nan
                try: return float(v)
                except: return np.nan
            values = [sfm(v) for v in row[1:len(managers)+1]]
            data_rows.append([date] + values)
        df = pd.DataFrame(data_rows, columns=['Date'] + managers)
        df = df.set_index('Date').sort_index(ascending=False)
        result[peer_tab] = df
    wb.close()
    return result


def load_universe_returns(filepath, peer_tab):
    """Load an eVestment 'Monthly Return Series' peer universe file.

    Format:
      Row 1-4: metadata (eVestment ID, "Results displayed in USD", blank, "Created On: ...")
      Row 5:   strategy/product name (used as the manager identifier)
      Row 6:   firm name
      Row 7:   full product name with vehicle suffix e.g. "(SA-GF)"
      Row 8+:  date in col A ascending, returns as decimals, '---' for missing
      Col 1 (index 1): the eVestment peer group aggregate e.g.
                       'International Equity Small Cap' — excluded from the universe.

    Returns {peer_tab: DataFrame} with the DataFrame indexed by date (descending)
    and manager names as columns, to match the shape of load_manager_returns.
    Returns empty dict if the file doesn't have the expected structure.
    """
    wb = load_workbook(filepath, read_only=True)
    target_sheet = None
    for sn in wb.sheetnames:
        if sn.lower().strip() in ('monthly return series', 'returns', 'monthly returns'):
            target_sheet = sn
            break
    if target_sheet is None:
        # fallback: first non-disclaimer sheet
        for sn in wb.sheetnames:
            if 'disclaim' not in sn.lower():
                target_sheet = sn
                break
    if target_sheet is None:
        wb.close()
        return {}

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the product-name header row: in the eVestment format, col A = 'Date'
    # in row 5 and the row contains strategy/product names starting at col 2.
    header_row_idx = None
    for i, r in enumerate(rows[:12]):
        col_a = r[0]
        if isinstance(col_a, str) and col_a.strip().lower() == 'date':
            header_row_idx = i
            break
    if header_row_idx is None:
        return {}

    header = rows[header_row_idx]
    # Use the strategy name as the manager identifier. Col index 1 is the
    # peer-group aggregate (e.g., "International Equity Small Cap") — exclude.
    managers = []          # list of (col_idx, name)
    for col_idx in range(2, len(header)):
        name = header[col_idx]
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue
        managers.append((col_idx, name))

    # Data rows start 3 rows after the header (row 5 header, row 6 firm, row 7
    # full name with vehicle, row 8 first data row).
    data_start = header_row_idx + 3

    def sfm(v):
        if v is None or (isinstance(v, str) and v.strip() in ('---', '', 'N/A', 'NA')):
            return np.nan
        try:
            return float(v)
        except (TypeError, ValueError):
            return np.nan

    # De-duplicate manager names by keeping the column with the most non-NaN values
    # (covers the SA-GF vs CF-GF vehicle duplication).
    col_data = {}           # name -> {col_idx: [values...], ...}
    dates = []
    for r in rows[data_start:]:
        if r[0] is None:
            continue
        try:
            dt = pd.to_datetime(r[0])
        except (TypeError, ValueError):
            continue
        dates.append(dt)
        for col_idx, name in managers:
            v = sfm(r[col_idx] if col_idx < len(r) else None)
            col_data.setdefault(name, {}).setdefault(col_idx, []).append(v)

    if not dates:
        return {}

    # Collapse duplicates: pick the column with fewest NaNs per name
    final = {}
    for name, variants in col_data.items():
        best_col = max(variants, key=lambda c: sum(1 for v in variants[c] if not np.isnan(v)))
        final[name] = variants[best_col]

    df = pd.DataFrame(final, index=pd.DatetimeIndex(dates))
    # Flip to descending (most recent first) to match load_manager_returns convention
    df = df.sort_index(ascending=False)
    # Drop any all-NaN columns (shouldn't happen but safety)
    df = df.dropna(axis=1, how='all')
    return {peer_tab: df}


# Mapping from the new consolidated file's sheet names to our peer-group tabs
_UNIVERSE_SHEET_TO_TAB = {
    'US LC':  'US',
    'US SC':  'USSC',
    'ISC':    'ISC',
    'EAFE':   'EAFE',
    'Global': 'ACWI',
    # Forgiving variants
    'US':     'US',
    'USSC':   'USSC',
    'ACWI':   'ACWI',
    'EM':     'EM',
}


def _repair_descending_dates(dates):
    """Fix a year-corruption artifact observed in FactSet/eVestment exports.

    The file sometimes stores `M/YY` dates where Excel re-interprets a
    two-digit year with the current year as the century pivot, producing a
    sequence like (read top to bottom):
        2026-03-26, 2026-02-26, 2026-01-26, 2026-12-25, 2026-11-25, ...
    The fourth row is actually Dec 2025, not Dec 2026. The pattern is that
    dates are expected to be strictly monotonically decreasing as we go
    down the column. Whenever we see a forward jump (date[i] >= date[i-1]),
    everything from that row onward belongs to an earlier year — we
    decrement the carried year-offset until monotonicity is restored.

    Returns a list of pandas Timestamps of the same length as input with
    NaT preserved in place.
    """
    import pandas as pd
    if not len(dates):
        return dates
    ds = [pd.Timestamp(d) if d is not None and not pd.isna(d) else pd.NaT for d in dates]
    repaired = []
    prev = None
    year_offset = 0
    for d in ds:
        if d is pd.NaT or pd.isna(d):
            repaired.append(pd.NaT)
            continue
        try:
            adjusted = d.replace(year=d.year + year_offset)
        except ValueError:
            adjusted = d
        if prev is not None and adjusted >= prev:
            # Step back one year at a time until monotonicity is restored
            guard = 0
            while adjusted >= prev and guard < 150:
                year_offset -= 1
                try:
                    adjusted = d.replace(year=d.year + year_offset)
                except ValueError:
                    adjusted = d
                    break
                guard += 1
        repaired.append(adjusted)
        prev = adjusted
    return repaired


def load_universe_returns_consolidated(filepath):
    """Load the new consolidated universe returns file.

    Format (one file, multiple sheets, one sheet per peer group):
      Sheet names: 'US LC', 'ISC', 'EAFE', 'Global' (→ 'ACWI' peer tab)
      Row 1: manager names (col A blank, col B onward are names)
      Row 2: 'Return, %' (indicates percentage units — we divide by 100)
      Row 3: 'Total' (return type)
      Row 4+: col A is the month-end date, col B+ are monthly returns
              expressed as percentages (e.g. -3.65 means -3.65%).
              Missing values are the string 'NA'.
      Dates are in descending order (newest first).

    Returns {peer_tab: DataFrame} with rows indexed by date (descending),
    manager names as columns, returns stored as DECIMALS (divided by 100) to
    match the buy-list convention downstream.

    Uses pandas.read_excel for speed — the wide sheets (hundreds-to-thousands
    of manager columns) are much faster to vectorize than to iterate cell by
    cell in Python.
    """
    import pandas as pd
    xl = pd.ExcelFile(filepath, engine='openpyxl')
    out = {}

    for sn in xl.sheet_names:
        tab = _UNIVERSE_SHEET_TO_TAB.get(sn.strip())
        if tab is None:
            continue

        # Row 1 → column headers (manager names). Rows 2-3 (Return,%, Total) are
        # metadata we skip via `skiprows`. First column (blank in row 1) becomes
        # the unnamed index.
        df = xl.parse(sn, header=0, skiprows=[1, 2], index_col=0)

        if df.empty or df.shape[0] == 0:
            continue

        # Drop any columns with no name or an 'Unnamed' label (blank header cells)
        good_cols = [c for c in df.columns
                     if c is not None and not str(c).startswith('Unnamed:')
                     and str(c).strip() != '']
        df = df[good_cols]

        # Ensure index is datetime
        df.index = pd.to_datetime(df.index, errors='coerce')
        df = df[df.index.notna()]
        if df.empty:
            continue

        # Repair year-corruption (see _repair_descending_dates). Input file
        # has rows in descending date order; we walk top-to-bottom and enforce
        # strict monotonic decrease, stepping years back when the stored year
        # has been over-advanced by a two-digit-year pivot.
        df.index = pd.DatetimeIndex(_repair_descending_dates(list(df.index)))
        df = df[df.index.notna()]

        # Normalize every date to end-of-month so universe returns align with
        # the factor returns file (which is strictly month-end-indexed).
        df.index = df.index + pd.offsets.MonthEnd(0)

        # Replace 'NA' strings and any non-numeric cells with NaN, then cast.
        # Using pd.to_numeric(errors='coerce') is vectorized and handles 'NA',
        # 'N/A', '---', empty strings — anything unparseable becomes NaN.
        df = df.apply(pd.to_numeric, errors='coerce')

        # Divide by 100 to convert percentage to decimal (matching buy-list convention)
        df = df / 100.0

        # Collapse duplicate column names by keeping the one with fewest NaNs
        if df.columns.duplicated().any():
            keep = {}
            for c in df.columns:
                if c in keep:
                    # compare non-null counts of the existing vs new; keep better
                    existing_good = df[keep[c]].notna().sum() if isinstance(keep[c], int) else None
                    continue
                # Find all positional indices for this name
                positions = [i for i, col in enumerate(df.columns) if col == c]
                if len(positions) == 1:
                    keep[c] = positions[0]
                else:
                    best = max(positions, key=lambda p: df.iloc[:, p].notna().sum())
                    keep[c] = best
            df = df.iloc[:, [keep[c] for c in sorted(set(df.columns), key=lambda x: list(df.columns).index(x))]]

        # Sort descending and drop all-NaN columns
        df = df.sort_index(ascending=False)
        df = df.dropna(axis=1, how='all')
        if df.shape[1] > 0:
            out[tab] = df

    return out


def load_weights(filepath):
    """Parse the client-weights workbook.

    Layout (as of Q2-26): one Sheet with multiple client blocks. Each block:
      Row N:   [client name]   [benchmark display name]   (benchmark optional
               for backwards compat with files that don't carry it)
      Row N+1: blank
      Row N+2: 'Manager', 'Weight'
      Row N+3..: [manager name] [weight as fraction]
      Row N+k: blank separator before the next client block.

    Returns a 2-tuple:
      - clients:    {client_name: {manager_name: weight}}
      - benchmarks: {client_name: benchmark_str or None}

    The benchmark string is taken verbatim from the workbook and is
    intentionally NOT resolved here — downstream consumers (risk engine,
    market-cycle chart, FactSet exposures) each translate it into their own
    representation with their own fallbacks, so a non-standard name like
    'MSCI EAFE+CANADA' can still display as-is while risk math falls back
    to the closest peer group.
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    clients = {}
    benchmarks = {}
    current_client = None
    in_data = False
    for row in rows:
        vals = [v for v in row if v is not None]
        if not vals:
            in_data = False
            continue
        first = str(vals[0]).strip()
        if first.lower() in ['manager', 'weight']:
            in_data = (first.lower() == 'manager')
            continue
        # Client-header row: first cell is the client name, optional second
        # cell is the benchmark. Distinguish from manager rows by the absence
        # of a numeric second cell.
        second_is_numeric = False
        if len(vals) >= 2:
            try:
                float(vals[1]); second_is_numeric = True
            except (ValueError, TypeError):
                pass
        if not second_is_numeric:
            current_client = first
            clients[current_client] = {}
            # Second cell, if present and non-empty, is the benchmark
            bench = None
            if len(vals) >= 2:
                b = str(vals[1]).strip()
                if b and b.lower() not in ('none', ''):
                    bench = b
            benchmarks[current_client] = bench
            in_data = False
        elif in_data and current_client and len(vals) >= 2:
            try:
                clients[current_client][first] = float(vals[1])
            except:
                pass
    return clients, benchmarks


def resolve_peer_group(bench_str):
    """Map a user-supplied benchmark name from the weights file to a
    peer-group tag in risk_engine.PEER_BENCHMARKS. Returns None when no
    confident match exists; callers should fall back to their prior logic
    (typically dominant_peer_group or a hardcoded override).

    Matching is token-based so variants like 'MSCI EAFE SC' / 'MSCI EAFE
    Small Cap' / 'MSCI EAFE Small' all resolve identically, and non-standard
    composites ('MSCI EAFE+Canada') still resolve to their closest peer
    group by the primary region token."""
    if not bench_str:
        return None
    # Normalise: upper-case, strip plus/hyphens, split on whitespace
    u = str(bench_str).upper()
    for ch in ['+', '-', '/', ',', '.']:
        u = u.replace(ch, ' ')
    toks = set(u.split())

    is_small = bool(toks & {'SC', 'SMALL', 'SMALLCAP'})
    # 'ex-US' was previously requiring the token 'USA' specifically — but
    # FactSet labels almost always say 'ex-US' (token 'US' after split).
    # Also recognise 'World ex US' alongside 'ACWI ex US'. Without these,
    # 'MSCI ACWI ex-US' / 'MSCI World ex US SC' fell through to the broad
    # ACWI bucket and the risk tables read against the wrong benchmark.
    is_xus = (
        bool(toks & {'XUS', 'ACWIXUS'})
        or (('ACWI' in toks or 'WORLD' in toks)
            and 'EX' in toks
            and ('US' in toks or 'USA' in toks))
    )
    has_eafe = 'EAFE' in toks
    has_world = 'WORLD' in toks or 'GLOBAL' in toks
    has_acwi  = 'ACWI' in toks
    has_acwi_or_world = has_acwi or has_world
    has_em   = 'EM' in toks or 'EMERGING' in toks
    has_russell_2000 = 'RUSSELL' in toks and '2000' in toks
    has_russell_1000 = 'RUSSELL' in toks and '1000' in toks

    # Most-specific first
    if has_russell_2000:
        return 'USSC'
    if has_russell_1000:
        return 'US'
    # Small-cap routing
    if is_small:
        # 'World ex US SC' = developed ex-US small cap → EAFE Small Cap peer
        # 'ACWI ex-US SC'  = includes EM small cap        → ISC peer
        if has_world and is_xus and not has_acwi:
            return 'EAFE_SC'
        if is_xus:
            return 'ISC'
        if has_eafe:
            return 'EAFE_SC'
        if has_em:
            return 'EM'   # no dedicated EM SC peer group yet
    if is_xus:
        # 'ACWI ex-US' includes EM → its own peer group (ACWI_xUS).
        # 'World ex US' / 'EAFE+Canada' → EAFE (developed-only).
        if has_acwi:
            return 'ACWI_xUS'
        return 'EAFE'
    if has_eafe:
        return 'EAFE'     # includes 'MSCI EAFE+Canada' and similar composites
    if has_acwi_or_world:
        return 'ACWI'
    if has_em:
        return 'EM'
    return None


XUS_MARKERS  = ['xus', 'acwixus', 'acwi ex', 'ex us', 'ex-us']
EAFE_MARKERS = ['eafe', 'isc', 'arga', 'algert', 'ballina', 'hillsdale', 'runde', 'ativo',
                 'cedar street eafe', 'lizard eafe']


def _is_xus(name):
    """True if a manager name signals an ACWIxUS / ex-US strategy."""
    nl = name.lower()
    return any(m in nl for m in XUS_MARKERS)


# Regional-signal gating for fuzzy_match. The goal is to prevent false-positive
# cross-tab matches after normalize_name() strips regional tokens: e.g. a
# weight named "Paradigm Global" (ACWI) normalises to "Paradigm" and would
# otherwise exact-match an EAFE manager literally named "Paradigm". If the
# weight name carries an explicit regional marker, we refuse to look in
# incompatible peer tabs at all.
def _regional_signal(name):
    """Return a coarse region tag extracted from the raw name: one of
    'ACWI', 'ISC_xUS', 'EAFE', 'EM', 'ISC', 'US', or None.
    Order matters — check the most specific markers first."""
    u = str(name).upper()
    # ACWI ex-US / xUS variants — ambiguous between ISC (if small) and EAFE
    if 'XUS' in u or 'ACWIXUS' in u or 'ACWI EX' in u or 'EX US' in u or 'EX-US' in u:
        return 'ISC_xUS'
    if 'ISC' in u:
        return 'ISC'
    if 'GLOBAL' in u or 'ACWI' in u or 'WORLD' in u:
        return 'ACWI'
    if 'EAFE' in u:
        return 'EAFE'
    if ' EM' in u or u.startswith('EM ') or 'EMERGING' in u:
        return 'EM'
    if ' US' in u or u.startswith('US ') or u.endswith(' US') or 'U.S.' in u:
        return 'US'
    return None


def _region_compatible(signal, tab):
    """Is a candidate found in `tab` compatible with the weight's region
    signal? When signal is None there is no regional constraint."""
    if signal is None:
        return True
    if signal == 'ACWI':
        return tab == 'ACWI'
    if signal == 'EAFE':
        # Let EAFE-flagged names find EAFE SC managers too (some weight
        # files tag SC strategies simply as "Acme EAFE Small").
        return tab in ('EAFE', 'ISC')
    if signal == 'EM':
        return tab == 'EM'
    if signal == 'ISC':
        return tab == 'ISC'
    if signal == 'ISC_xUS':
        return tab in ('ISC', 'EAFE', 'ACWI')
    if signal == 'US':
        return tab in ('US', 'USSC')
    return True


def normalize_name(name):
    name = str(name).strip()
    suffixes = [
        r'\bEAFE SC\b', r'\bEAFE Small Cap\b', r'\bISC\b', r'\bEM SC\b',
        r'\bACWI SC\b', r'\bACWIxUS SC\b', r'\bxUS SC\b', r'\bEAFE\b',
        r'\bACWI\b', r'\bACWIxUS\b', r'\bxUS\b', r'\bEM\b',
        r'\bSC\b', r'\bUS\b', r'\bGlobal\b'
    ]
    for s in suffixes:
        name = re.sub(s, '', name, flags=re.IGNORECASE).strip()
    return re.sub(r'\s+', ' ', name).strip()


def infer_tab(weight_name):
    u = weight_name.upper()
    is_small = (' SC' in u or u.endswith(' SC') or u.startswith('SC ')
                or ' SMALL' in u or 'SMALLCAP' in u)
    has_global = 'GLOBAL' in u
    has_acwi   = any(x in u for x in ['ACWI', 'ACWIXUS', 'XUS'])
    has_eafe   = 'EAFE' in u
    has_isc    = 'ISC' in u
    has_em     = ' EM' in u or u.startswith('EM ') or 'EMERGING' in u
    has_us     = ' US' in u or u.startswith('US ') or u.endswith(' US') or 'U.S.' in u

    # Small-cap international: EAFE SC / ISC / GLOBAL SC / ACWI SC all → ISC tab
    if has_isc:
        return 'ISC'
    if is_small and (has_eafe or has_acwi or has_global):
        return 'ISC'
    if is_small and has_em:
        return 'EM'
    # Small-cap US — route to the dedicated USSC peer tab. This fires whether
    # the name has an explicit "US" marker ("US Small Cap Core") or not
    # ("Acme Small Cap Growth"): once we've ruled out international/EM
    # small-cap above, anything still flagged is_small is domestic.
    if is_small:
        return 'USSC'
    # Global / ACWI variants (both treated as ACWI peer group, non-small-cap)
    if has_acwi or has_global:
        return 'ACWI'
    if has_eafe:
        return 'EAFE'
    if has_em:
        return 'EM'
    if has_us:
        return 'US'
    return 'EAFE'


_INDEX_PREFIXES = ('msci ', 'russell ', 'ftse ', 's&p ', 'bloomberg ',
                   'stoxx ', 'nikkei ', 'topix ')
_INDEX_KEYWORDS = ('benchmark', 'index')


def _is_index_name(s):
    """True if a buy-list manager name looks like an index passthrough
    (e.g. 'MSCI EAFE + Canada', 'Russell 1000', 'Benchmark Foo'). Used
    by fuzzy_match to gate cross-token matching: a real manager weight
    like 'Haven EAFE + Canada' must NOT silently fall onto an index-
    named buy-list entry just because they share regional suffixes."""
    if not s:
        return False
    low = s.lower()
    return low.startswith(_INDEX_PREFIXES) or any(k in low for k in _INDEX_KEYWORDS)


def _shares_index_token(weight_name, candidate):
    """When `candidate` is an index-style name, only allow a fuzzy
    match if `weight_name` also references the same index family — i.e.
    it shares an MSCI/Russell/FTSE/etc. or Benchmark/Index token. So
    'Benchmark MSCI EAFE + Canada' (Client 5's index passthrough weight)
    can still resolve to 'MSCI EAFE + Canada' in the buy-list, but
    'Haven EAFE + Canada' (a real manager) cannot."""
    if not _is_index_name(candidate):
        return True
    wl = (weight_name or '').lower()
    for tok in _INDEX_PREFIXES:
        if tok in candidate.lower() and tok in wl:
            return True
    for tok in _INDEX_KEYWORDS:
        if tok in candidate.lower() and tok in wl:
            return True
    # Also accept if weight name contains the candidate's first index token
    # as a standalone word (e.g., weight 'MSCI EAFE + Canada' against
    # candidate 'MSCI EAFE + Canada'): the prefix-startswith check above
    # already covers this, but keep symmetric handling for safety.
    return False


def fuzzy_match(weight_name, manager_dfs, threshold=60):
    """
    Match a weight-file manager name to an actual manager in the returns file.

    Key rule: if the weight name contains 'EAFE' (but NOT xUS/ACWIxUS markers),
    restrict ISC-tab candidates to non-xUS managers only, and vice-versa.
    This prevents 'CastleArk EAFE SC' from matching 'Castleark xUS'.

    When the weight name carries an explicit regional marker (GLOBAL/ACWI/
    WORLD/EAFE/EM/ISC/xUS/US), tabs that are incompatible with that marker
    are skipped entirely — so 'Paradigm Global' can never resolve to an
    EAFE manager named 'Paradigm', even after normalize_name() has stripped
    the 'Global' token.

    manager_dfs can be either:
      - {tab: DataFrame} (the normal case, when manager returns are loaded)
      - {tab: list_of_manager_names} (fallback, when only clone_results is
        available — e.g., after a cache reload whose file paths no longer
        resolve)
    Either is accepted; only the .columns vs direct-iteration branch differs.
    """
    if manager_dfs is None:
        return None, None

    inferred     = infer_tab(weight_name)
    core         = normalize_name(weight_name)
    weight_xus   = _is_xus(weight_name)
    weight_region = _regional_signal(weight_name)

    search_order = [inferred] + [t for t in PEER_TABS if t != inferred]

    # First pass: exact-core match, then fuzzy@threshold. Capture the
    # region-compatible candidate pool along the way so the second-pass
    # fallback (below) doesn't have to recompute it.
    region_pool = []   # list of (tab, candidates, cores)

    for tab in search_order:
        if tab not in manager_dfs:
            continue
        # Regional gating: if the weight name declares a region, skip tabs
        # that can't possibly host the right manager.
        if not _region_compatible(weight_region, tab):
            continue
        tab_obj = manager_dfs[tab]
        # Support both DataFrame (has .columns) and plain list/iterable
        all_candidates = list(tab_obj.columns) if hasattr(tab_obj, 'columns') else list(tab_obj)

        # For ISC (and EAFE) tabs, filter by xUS vs EAFE side
        if tab in ('ISC', 'EAFE'):
            if weight_xus:
                # Weight name is xUS-style — only consider xUS managers
                candidates = [c for c in all_candidates if _is_xus(c)]
            else:
                # Weight name is EAFE-style — exclude xUS managers
                candidates = [c for c in all_candidates if not _is_xus(c)]
            # If filtering leaves nothing, fall back to all (avoids dead ends)
            if not candidates:
                candidates = all_candidates
        else:
            candidates = all_candidates

        cores = [normalize_name(c) for c in candidates]
        region_pool.append((tab, candidates, cores))

        # Exact core-name match first
        for orig, c in zip(candidates, cores):
            if c.lower() == core.lower():
                return tab, orig

        # Fuzzy match within filtered set, but gate index-style candidates
        # ('MSCI EAFE + Canada', 'Russell 1000', etc.) so real-manager
        # weight names like 'Haven EAFE + Canada' don't silently fall
        # onto them via the shared regional suffix.
        fuzzy_pool = [(c, k) for c, k in zip(candidates, cores)
                       if _shares_index_token(weight_name, c)]
        if fuzzy_pool:
            fuzzy_cores = [k for _, k in fuzzy_pool]
            fuzzy_origs = [c for c, _ in fuzzy_pool]
            match = process.extractOne(core, fuzzy_cores,
                                        scorer=fuzz.token_sort_ratio,
                                        score_cutoff=threshold)
            if match:
                return tab, fuzzy_origs[fuzzy_cores.index(match[0])]

    # Second pass — only relevant when the weight name carried an explicit
    # regional signal (so region_pool is meaningfully narrowed). Inside a
    # region-gated pool we trust that any same-first-token candidate is the
    # intended manager, even when the fuzz score undershoots the default
    # threshold. Example: weight "Paradigm Global" (ACWI-gated) matches the
    # ACWI sheet's "Paradigm World Value" — both share the "paradigm" first
    # token, and ACWI was the only compatible tab, so no false-positive risk.
    if weight_region is not None and core:
        weight_first = core.split()[0].lower()
        best = None   # (score, tab, orig)
        for tab, candidates, cores in region_pool:
            for orig, c in zip(candidates, cores):
                # Same index-token gate as the first pass — keeps an
                # 'MSCI EAFE + Canada' buy-list entry from absorbing real
                # managers when the first-token fallback fires.
                if not _shares_index_token(weight_name, orig):
                    continue
                toks = c.split()
                if not toks:
                    continue
                if toks[0].lower() != weight_first:
                    continue
                score = fuzz.token_sort_ratio(core, c)
                if best is None or score > best[0]:
                    best = (score, tab, orig)
        if best is not None:
            return best[1], best[2]

    return None, None


def compute_style_buckets(betas):
    buckets = {}
    for f, w in betas.items():
        b = STYLE_BUCKET_MAP.get(f, 'Other')
        buckets[b] = buckets.get(b, 0) + w
    return buckets


def compute_pct_small(betas):
    return sum(w for f, w in betas.items() if f in SMALL_CAP_FACTORS)


def compute_pct_em(betas):
    return sum(w for f, w in betas.items() if f in EM_FACTORS)


def compute_vg_full(betas):
    """Full-model V−G: (Value bucket + Yield bucket) − Growth bucket.

    Earlier versions used a hand-curated VALUE_FACTORS set that lumped Pure
    Value, Quality, Yield/Dividend, and the EAFE small-cap variants together
    into one "value" side. That overstated the value loading for managers
    with Quality or Yield exposure and didn't match the bucket definitions
    shown elsewhere in the tool. We now key off STYLE_BUCKET_MAP so the V−G
    calculation is consistent with the same buckets the UI surfaces.

    Quality is intentionally NOT included on the value side here — quality
    is its own bucket (a stand-alone style with both growth and value
    flavours) and rolling it into V−G again would conflate signals.
    """
    v = sum(w for f, w in betas.items() if STYLE_BUCKET_MAP.get(f) == 'Value')
    y = sum(w for f, w in betas.items() if STYLE_BUCKET_MAP.get(f) == 'Yield')
    g = sum(w for f, w in betas.items() if STYLE_BUCKET_MAP.get(f) == 'Growth')
    return (v + y) - g


def compute_vg_3factor(betas):
    """3-factor V−G: same V+Y vs G logic, applied to the 3-factor model.
    The 3-factor model only contains Core/Value/Growth factors today (no
    Yield bucket), so the Y term is zero in practice — but keeping the
    same formula here makes the two metrics directly comparable, and
    means the 3-factor V−G will Just Work if we ever add a yield-style
    factor to the 3-factor candidate list."""
    v = sum(w for f, w in betas.items() if STYLE_BUCKET_MAP.get(f) == 'Value')
    y = sum(w for f, w in betas.items() if STYLE_BUCKET_MAP.get(f) == 'Yield')
    g = sum(w for f, w in betas.items() if STYLE_BUCKET_MAP.get(f) == 'Growth')
    return (v + y) - g


def run_cloning(manager_returns_path, factor_returns_path, progress_callback=None, tabs_to_run=None):
    if progress_callback: progress_callback("Loading factor returns...")
    fdf = load_factor_returns(factor_returns_path)
    if progress_callback: progress_callback("Loading manager returns...")
    mdf = load_manager_returns(manager_returns_path)

    tabs = tabs_to_run if tabs_to_run else PEER_TABS
    results = {}
    total = sum(len(mdf[t].columns) for t in tabs if t in mdf)
    done = 0

    for tab in tabs:
        if tab not in mdf:
            continue
        mgr_df = mdf[tab]
        results[tab] = {}
        full_cols  = _resolve_factors_ci(FACTOR_CATEGORIES[f'{tab}_full'],    fdf.columns)
        three_cols = _resolve_factors_ci(FACTOR_CATEGORIES[f'{tab}_3factor'], fdf.columns)
        common = mgr_df.index.intersection(fdf.index)

        # Pre-compute factor matrices once per tab
        X_full_mat  = fdf.loc[common, full_cols].values  if full_cols  else None
        X_three_mat = fdf.loc[common, three_cols].values if three_cols else None
        dates_list  = [d.strftime('%Y-%m-%d') for d in common.tolist()]

        def _clone_manager(mgr_name):
            Y = mgr_df[mgr_name].reindex(common).values
            if np.sum(~np.isnan(Y)) < 24:
                return mgr_name, None
            mgr_result = {
                'tab': tab, 'peer_group': tab,
                'dates': dates_list,
                'manager_returns': [None if np.isnan(v) else round(float(v),6) for v in Y],
                '_input_hash': _input_hash(dates_list, full_cols, three_cols,
                                           X_full_mat, X_three_mat, Y),
            }
            if X_full_mat is not None:
                r = clone_fun(X_full_mat, Y, is_dynamic=False)
                mgr_result['static_clone_full'] = [None if np.isnan(v) else round(float(v),6) for v in r['outTS']]
                hr2 = _headline_r2(r['R2'])
                mgr_result['r2_full'] = round(hr2, 4) if hr2 is not None else None
                b_row = r['beta_all'][0]
                mgr_result['betas_full'] = {full_cols[i]: round(float(b),6)
                    for i,b in enumerate(b_row) if not np.isnan(b) and b > 1e-10}
            if X_three_mat is not None:
                r3 = clone_fun(X_three_mat, Y, is_dynamic=False)
                mgr_result['static_clone_3factor'] = [None if np.isnan(v) else round(float(v),6) for v in r3['outTS']]
                hr2_3 = _headline_r2(r3['R2'])
                mgr_result['r2_3factor'] = round(hr2_3, 4) if hr2_3 is not None else None
                b3 = r3['beta_all'][0]
                mgr_result['betas_3factor'] = {three_cols[i]: round(float(b),6)
                    for i,b in enumerate(b3) if not np.isnan(b) and b > 1e-10}
            mgr_result['style_buckets'] = compute_style_buckets(mgr_result.get('betas_full', {}))
            mgr_result['pct_small']     = round(compute_pct_small(mgr_result.get('betas_full', {})),4)
            mgr_result['pct_em']        = round(compute_pct_em(mgr_result.get('betas_full', {})),4)
            mgr_result['vg_full']       = round(compute_vg_full(mgr_result.get('betas_full', {})),4)
            mgr_result['vg_3factor']    = round(compute_vg_3factor(mgr_result.get('betas_3factor', {})),4)
            return mgr_name, mgr_result

        # Parallel execution — use min(cpu_count, 8) workers
        n_workers = min(os.cpu_count() or 2, 8)
        mgr_names = list(mgr_df.columns)
        with concurrent.futures.ThreadPoolExecutor(max_workers=n_workers) as executor:
            futures = {executor.submit(_clone_manager, m): m for m in mgr_names}
            for future in concurrent.futures.as_completed(futures):
                mgr_name, mgr_result = future.result()
                done += 1
                if progress_callback:
                    progress_callback(f"[{done}/{total}] {tab} - {mgr_name}")
                if mgr_result is not None:
                    results[tab][mgr_name] = mgr_result

    return results


def run_universe_cloning(universe_df_map, factor_df, progress_callback=None,
                         checkpoint_callback=None, checkpoint_every=50,
                         existing_results=None):
    """Clone every manager in a universe DataFrame map.

    universe_df_map      : {peer_tab: DataFrame}
    factor_df            : already loaded via load_factor_returns
    progress_callback    : callable(msg) for the progress log / bar
    checkpoint_callback  : callable(tab, partial_results_for_tab) — fired every
                           `checkpoint_every` managers AND at the end of each tab.
                           Caller should merge into persistent state and flush to
                           disk so an interruption doesn't lose all progress.
    checkpoint_every     : how often to checkpoint within a tab (default 50)
    existing_results     : {peer_tab: {mgr_name: result}} of already-completed
                           clones; matching managers are skipped so a re-run
                           resumes mid-tab rather than restarting from scratch.

    Returns {peer_tab: {mgr_name: result_dict}}.
    """
    existing = existing_results or {}

    # Grand total counts only managers that still need to be cloned
    total = sum(
        max(0, len(df.columns) - len(existing.get(tab) or {}))
        for tab, df in universe_df_map.items()
        if tab in PEER_TABS
    )
    done = 0
    results = {}

    for tab, mgr_df in universe_df_map.items():
        if tab not in PEER_TABS:
            continue

        # Seed this tab's results with whatever was already done so the
        # checkpoint_callback always receives the complete picture for the tab.
        results[tab] = dict(existing.get(tab) or {})
        already_done = set(results[tab].keys())

        full_cols  = _resolve_factors_ci(FACTOR_CATEGORIES[f'{tab}_full'],    factor_df.columns)
        three_cols = _resolve_factors_ci(FACTOR_CATEGORIES[f'{tab}_3factor'], factor_df.columns)
        common = mgr_df.index.intersection(factor_df.index)
        all_cols = set(full_cols) | set(three_cols)
        if all_cols:
            factor_slice = factor_df.loc[common, list(all_cols)]
            common = common[~factor_slice.isna().any(axis=1)]
        if len(common) < 24:
            continue

        X_full_mat  = factor_df.loc[common, full_cols].values  if full_cols  else None
        X_three_mat = factor_df.loc[common, three_cols].values if three_cols else None
        dates_list  = [d.strftime('%Y-%m-%d') for d in common.tolist()]

        def _clone_manager(mgr_name):
            Y = mgr_df[mgr_name].reindex(common).values
            if np.sum(~np.isnan(Y)) < 24:
                return mgr_name, None
            mgr_result = {
                'tab': tab, 'peer_group': tab,
                'dates': dates_list,
                'manager_returns': [None if np.isnan(v) else round(float(v),6) for v in Y],
            }
            if X_full_mat is not None:
                r = clone_fun(X_full_mat, Y, is_dynamic=False)
                mgr_result['static_clone_full'] = [None if np.isnan(v) else round(float(v),6) for v in r['outTS']]
                hr2 = _headline_r2(r['R2'])
                mgr_result['r2_full'] = round(hr2, 4) if hr2 is not None else None
                b_row = r['beta_all'][0]
                mgr_result['betas_full'] = {full_cols[i]: round(float(b),6)
                    for i,b in enumerate(b_row) if not np.isnan(b) and b > 1e-10}
            if X_three_mat is not None:
                r3 = clone_fun(X_three_mat, Y, is_dynamic=False)
                mgr_result['static_clone_3factor'] = [None if np.isnan(v) else round(float(v),6) for v in r3['outTS']]
                hr2_3 = _headline_r2(r3['R2'])
                mgr_result['r2_3factor'] = round(hr2_3, 4) if hr2_3 is not None else None
                b3 = r3['beta_all'][0]
                mgr_result['betas_3factor'] = {three_cols[i]: round(float(b),6)
                    for i,b in enumerate(b3) if not np.isnan(b) and b > 1e-10}
            mgr_result['style_buckets'] = compute_style_buckets(mgr_result.get('betas_full', {}))
            mgr_result['pct_small']     = round(compute_pct_small(mgr_result.get('betas_full', {})),4)
            mgr_result['pct_em']        = round(compute_pct_em(mgr_result.get('betas_full', {})),4)
            mgr_result['vg_full']       = round(compute_vg_full(mgr_result.get('betas_full', {})),4)
            mgr_result['vg_3factor']    = round(compute_vg_3factor(mgr_result.get('betas_3factor', {})),4)
            return mgr_name, mgr_result

        # Skip managers already present in existing_results
        pending = [m for m in mgr_df.columns if m not in already_done]
        n_workers = min(os.cpu_count() or 2, 8)
        since_last_checkpoint = 0

        with concurrent.futures.ThreadPoolExecutor(max_workers=n_workers) as executor:
            futures = {executor.submit(_clone_manager, m): m for m in pending}
            for future in concurrent.futures.as_completed(futures):
                mgr_name, mgr_result = future.result()
                done += 1
                since_last_checkpoint += 1
                if progress_callback:
                    progress_callback(f"[universe {done}/{total}] {tab} - {mgr_name}")
                if mgr_result is not None:
                    results[tab][mgr_name] = mgr_result
                # Mid-tab checkpoint — saves progress so an interruption
                # (laptop sleep, crash, etc.) doesn't lose the whole tab.
                if checkpoint_callback and since_last_checkpoint >= checkpoint_every:
                    checkpoint_callback(tab, results[tab])
                    since_last_checkpoint = 0

        # Final checkpoint for any remainder below the interval threshold
        if checkpoint_callback:
            checkpoint_callback(tab, results[tab])

    return results


def build_portfolio_view(client_name, weights_dict, clone_results, manager_dfs,
                         universe_clone_results=None, placeholder_buckets=None):
    # Defensive fallback: if manager_dfs didn't get reloaded from disk (cache
    # paths no longer resolve after a zip deploy), synthesize a lookup from
    # clone_results keys so fuzzy_match still works. Every manager in
    # clone_results was originally loaded from manager_dfs, so the name space
    # is identical. Also merge universe names so recently-added managers that
    # exist only in universe_clone_results can still be looked up.
    ucr = universe_clone_results or {}
    if not manager_dfs:
        merged_keys = {}
        for tab, td in (clone_results or {}).items():
            merged_keys.setdefault(tab, set()).update(td.keys())
        for tab, td in ucr.items():
            merged_keys.setdefault(tab, set()).update(td.keys())
        manager_dfs = {tab: list(names) for tab, names in merged_keys.items()}

    managers = []
    unmatched = []
    for wt_name, weight in weights_dict.items():
        tab, mgr_name = fuzzy_match(wt_name, manager_dfs)

        # Look for clone data in buy-list clones first, then universe clones.
        # IMPORTANT: use the exact key as stored in clone_results / ucr, not
        # the (possibly slightly different) string returned by fuzzy_match.
        # This guards against whitespace or casing discrepancies between the
        # Excel column header and whatever fuzzy_match reconstructed from the
        # normalised-key lookup — those discrepancies would cause
        # _norm_skill_for to silently return None even when the manager IS in
        # norm_skill_by_tab under its exact stored key.
        d        = None
        exact_key = mgr_name   # may be overridden below
        if tab is not None:
            # Try buy-list clones (exact, then case-insensitive)
            cr_tab = (clone_results or {}).get(tab, {})
            if mgr_name in cr_tab:
                d = cr_tab[mgr_name]
                exact_key = mgr_name
            else:
                key_norm = str(mgr_name or '').strip().lower()
                for k, v in cr_tab.items():
                    if str(k or '').strip().lower() == key_norm:
                        d = v
                        exact_key = k
                        break
            # Fall back to universe clones if not in buy-list
            if d is None:
                ucr_tab = ucr.get(tab, {})
                if mgr_name in ucr_tab:
                    d = ucr_tab[mgr_name]
                    exact_key = mgr_name
                else:
                    key_norm = str(mgr_name or '').strip().lower()
                    for k, v in ucr_tab.items():
                        if str(k or '').strip().lower() == key_norm:
                            d = v
                            exact_key = k
                            break
        if d is None:
            # Manager exists in the weights file but has no clone data yet
            # — typically a newly-funded manager with <3 years of returns.
            # Synthesise a placeholder entry so the manager still appears
            # in the portfolio table and counts toward FactSet exposure
            # aggregations (which match against the exposures/risk files,
            # not clone_results). Style buckets default to 100% Core; the
            # user can override per manager via the Placeholder peer-group
            # tab and those overrides flow through here.
            ph_buckets = (placeholder_buckets or {}).get(wt_name)
            sb = dict(ph_buckets) if ph_buckets else {'Core': 1.0}
            # Derived metrics from the (possibly user-edited) buckets so
            # V-G, %SC, %EM in the portfolio table reflect overrides.
            v = sum(w for b, w in sb.items() if b == 'Value')
            y = sum(w for b, w in sb.items() if b == 'Yield')
            g = sum(w for b, w in sb.items() if b == 'Growth')
            vg = round((v + y) - g, 6)
            managers.append({
                'weight_file_name': wt_name,
                'matched_name':     wt_name,  # use the weights name verbatim
                'tab':              'Placeholder',
                'current_weight':   weight,
                'proposed_weight':  weight,
                'vg_full':          vg,
                'vg_3factor':       vg,
                'style_buckets':    sb,
                'r2_full':          None,
                'r2_3factor':       None,
                'betas_full':       {},
                'betas_3factor':    {},
                'pct_small':        sb.get('Small Cap', 0) or 0,
                'pct_em':           0,
                'is_placeholder':   True,
            })
            continue
        managers.append({
            'weight_file_name': wt_name,
            'matched_name': exact_key,   # exact key from clone_results/ucr
            'tab': tab,
            'current_weight': weight,
            'proposed_weight': weight,
            'vg_full': d.get('vg_full', 0),
            'vg_3factor': d.get('vg_3factor', 0),
            'style_buckets': d.get('style_buckets', {}),
            'r2_full': d.get('r2_full'),
            'r2_3factor': d.get('r2_3factor'),
            'betas_full': d.get('betas_full', {}),
            'betas_3factor': d.get('betas_3factor', {}),
            'pct_small': d.get('pct_small', 0),
            'pct_em': d.get('pct_em', 0),
        })
    return {'client': client_name, 'managers': managers, 'unmatched': unmatched}


# ── Benchmark map by tab ───────────────────────────────────────────────────
TAB_BENCHMARK = {
    'EAFE':  'MSCI EAFE NR USD',
    'ACWI':  'MSCI ACWI NR USD',
    'ISC':   'MSCI EAFE Small Cap NR USD',
    'EM':    'MSCI EM NR USD',
    'US':    'Russell 1000 TR USD',
    'USSC':  'Russell 2000 TR USD',
}

US_SMALLCAP_KEYWORDS = ['sc', 'micro', 'small']
XUS_KEYWORDS = ['xus', 'acwixus', 'acwi ex']

def get_benchmark(tab, mgr_name):
    name_lower = mgr_name.lower()
    if tab == 'USSC':
        return 'Russell 2000 TR USD'
    if tab == 'US':
        if any(k in name_lower for k in US_SMALLCAP_KEYWORDS):
            return 'Russell 2000 TR USD'
        return 'Russell 1000 TR USD'
    if tab == 'ISC':
        # xUS SC managers get ACWIxUS Small; others get EAFE SC
        if any(k in name_lower for k in XUS_KEYWORDS):
            return 'MSCI ACWI Ex USA Small NR USD'
        return 'MSCI EAFE Small Cap NR USD'
    if tab == 'EAFE':
        if any(k in name_lower for k in XUS_KEYWORDS):
            return 'MSCI ACWI Ex USA NR USD'
        return 'MSCI EAFE NR USD'
    return TAB_BENCHMARK.get(tab, 'MSCI EAFE NR USD')


def compound_return(values):
    """Compound a list of decimal monthly returns. None/nan safe."""
    r = np.array([v for v in values if v is not None and not np.isnan(float(v))], dtype=float)
    if len(r) == 0:
        return None
    return float(np.prod(1 + r) - 1)


def annualize(total_ret, n_months):
    if total_ret is None or n_months < 1:
        return None
    # (1 + total_ret) ** (fractional) goes complex when (1 + total_ret) <= 0,
    # which would mean the portfolio lost all its value. Treat that as -100%.
    base = 1.0 + float(total_ret)
    if base <= 0:
        return -1.0
    return float(base ** (12.0 / n_months) - 1.0)


def compute_period_returns(returns_list, n_months=None):
    """
    returns_list: list of decimal monthly returns, index 0 = most recent.
    n_months: how many months to use. None = all available.
    Returns annualized return or None.
    """
    if n_months is not None:
        vals = returns_list[:n_months]
    else:
        vals = returns_list
    # Need at least 75% of requested period to compute
    min_needed = max(1, int((n_months or len(vals)) * 0.75)) if n_months else 1
    valid = [v for v in vals if v is not None and not np.isnan(float(v))]
    if len(valid) < min_needed:
        return None
    total = compound_return(valid)
    actual_months = len(valid)
    if n_months and actual_months >= 12:
        return annualize(total, actual_months)
    elif actual_months >= 12:
        return annualize(total, actual_months)
    else:
        return total  # < 1 year: just return total, not annualized


def compute_skill_periods(mgr_returns, clone_returns):
    """
    Compute skill (manager - clone) for standard periods.
    Returns dict with keys: qtd, ytd, t1, t3, t5, si (all annualized where applicable).
    Index 0 = most recent month.
    """
    # Most recent date is March 2026 (end of Q1)
    # QTD = 3 months (Jan-Mar), YTD = same as QTD for Q1
    # Determine QTD dynamically based on position in year
    # We'll always compute QTD as 3 months, YTD as months since year start
    # Since data ends March 2026 (month 3), YTD = 3 months = same as QTD

    def skill(n):
        mr = [v for v in mgr_returns[:n] if v is not None and not np.isnan(float(v))]
        cr = [v for v in clone_returns[:n] if v is not None and not np.isnan(float(v))]
        if len(mr) < max(1, n//2) or len(cr) < max(1, n//2):
            return None
        # Use only overlapping valid indices
        avail = min(n, len(mgr_returns), len(clone_returns))
        pairs = [(mgr_returns[i], clone_returns[i]) for i in range(avail)
                 if mgr_returns[i] is not None and clone_returns[i] is not None
                 and not np.isnan(float(mgr_returns[i])) and not np.isnan(float(clone_returns[i]))]
        if not pairs:
            return None
        m_tot = float(np.prod([1+p[0] for p in pairs]) - 1)
        c_tot = float(np.prod([1+p[1] for p in pairs]) - 1)
        skill_total = m_tot - c_tot
        nm = len(pairs)
        if nm >= 12:
            # Annualize each separately then subtract
            m_ann = annualize(m_tot, nm)
            c_ann = annualize(c_tot, nm)
            return m_ann - c_ann if m_ann is not None and c_ann is not None else None
        return skill_total

    # Since inception: use all valid paired months
    all_pairs = [(mgr_returns[i], clone_returns[i]) for i in range(min(len(mgr_returns), len(clone_returns)))
                 if mgr_returns[i] is not None and clone_returns[i] is not None
                 and not np.isnan(float(mgr_returns[i])) and not np.isnan(float(clone_returns[i]))]
    if all_pairs:
        m_tot = float(np.prod([1+p[0] for p in all_pairs]) - 1)
        c_tot = float(np.prod([1+p[1] for p in all_pairs]) - 1)
        nm = len(all_pairs)
        if nm >= 12:
            si_skill = annualize(m_tot, nm) - annualize(c_tot, nm)
        else:
            si_skill = m_tot - c_tot
    else:
        si_skill = None

    return {
        't1':  skill(12),
        't3':  skill(36),
        't5':  skill(60),
        'si':  si_skill,
    }


def compute_cumulative_skill(mgr_returns, clone_returns):
    """
    Build cumulative skill series (like growth of $100 chart for excess return).
    Returns list of cumulative skill values aligned with dates, most recent first.
    Starts at 0 and compounds (1+mgr)/(1+clone) - 1 each month.
    """
    n = min(len(mgr_returns), len(clone_returns))
    # Work oldest to newest (reverse)
    pairs = []
    for i in range(n-1, -1, -1):
        m = mgr_returns[i]
        c = clone_returns[i]
        if m is not None and c is not None and not np.isnan(float(m)) and not np.isnan(float(c)):
            pairs.append((float(m), float(c), i))
        else:
            pairs.append((None, None, i))

    # Build cumulative excess (oldest to newest)
    cum_excess = []
    running = 0.0
    for m, c, _ in pairs:
        if m is not None and c is not None:
            # Monthly excess return: (1+m)/(1+c) - 1
            monthly_excess = (1 + m) / (1 + c) - 1
            running += monthly_excess  # simple additive for display purposes
        cum_excess.append(running)

    # Reverse back to most-recent-first
    cum_excess.reverse()
    return [round(v * 100, 4) for v in cum_excess]  # in percentage points


def compute_manager_period_returns(mgr_returns, clone_returns, benchmark_returns):
    """
    Compute standard period returns for manager, clone, and benchmark.
    Returns dict with qtd, ytd, t1, t3, t5, si for each.
    Most recent index = 0. All annualized where >= 12 months.
    """
    def period(rets, n):
        vals = [v for v in rets[:n] if v is not None and not np.isnan(float(v))]
        if len(vals) < max(1, n // 2): return None
        tot = compound_return(vals)
        return annualize(tot, len(vals)) if len(vals) >= 12 else tot

    def si_ret(rets):
        vals = [v for v in rets if v is not None and not np.isnan(float(v))]
        if not vals: return None
        tot = compound_return(vals)
        return annualize(tot, len(vals)) if len(vals) >= 12 else tot

    result = {}
    for label, rets in [('mgr', mgr_returns), ('clone', clone_returns), ('bench', benchmark_returns)]:
        result[label] = {
            'qtd': period(rets, 3),
            'ytd': period(rets, 3),   # Mar 2026 = end of Q1, so YTD = QTD
            't1':  period(rets, 12),
            't3':  period(rets, 36),
            't5':  period(rets, 60),
            'si':  si_ret(rets),
        }
    return result
