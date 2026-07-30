"""
Qualitative-data loader — parses the firm-level qualitative workbook the user
uploads alongside the buy-list returns.

Firm-only format (single sheet), a header row then one row per firm:

    Firm                | Firm AUM ($mm) | Ownership            | Diverse/Female Ownership %
    Arga                | 12500          | 100% employee-owned  | 0
    Ativo               | 3100           | 60% minority-owned   | 60
    CastleArk           | 8700           | Majority woman-owned | 55

There are no strategy rows and no strategy AUM. Each buy-list strategy is
mapped to a firm at lookup time by PREFIX MATCH: a strategy maps to the firm
whose name the strategy name starts with (case/space-insensitive). So firm
"Arga" captures "Arga", "Arga ISC", "Arga xUS", "ARGA Global". When several
firms could match, the longest firm name wins (so "Ativo Global" firm would
beat "Ativo" for strategy "Ativo Global SC", if both were listed).

Returns
-------
{
    'firms': {
        firm_name: {'firm_aum': float|None, 'ownership': str|None,
                    'diverse_pct': float|None}
    },
    'firm_order': [firm_name, ...],   # longest-name-first, for prefix matching
    'n_firms': int,
    'warnings': [str, ...],
}
"""

import re
import openpyxl


_FIELD_SYNONYMS = {
    'firm_name':    ('firm', 'firm name', 'firm / strategy', 'firm/strategy',
                     'manager', 'name'),
    'firm_aum':     ('firm aum', 'firm aum ($mm)', 'firm aum $mm', 'total firm aum',
                     'firm assets', 'aum'),
    'ownership':    ('ownership', 'ownership structure', 'firm ownership'),
    'diverse_pct':  ('diverse/female ownership', 'diverse female ownership',
                     'diverse/female ownership %', 'diverse female ownership pct',
                     'diverse ownership', 'female ownership', 'diverse/female',
                     'diverse female', 'women/minority ownership',
                     'minority/women ownership', 'diverse pct',
                     'diverse/woman ownership', 'diverse woman ownership',
                     'diverse/woman owned', 'diverse woman owned'),
}


def _norm_hdr(s):
    s = str(s or '').lower().strip()
    s = re.sub(r'\([^)]*\)', '', s)          # drop unit parens
    s = re.sub(r'[^a-z0-9 /%]+', ' ', s)     # keep slash + percent
    s = s.replace('%', '')                   # then drop the % sign itself
    return re.sub(r'\s+', ' ', s).strip()


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('$', '')
    if s == '' or s == '-':
        return None
    # Explicit percent text, e.g. "55%" — already a whole percent, keep as 55.
    if s.endswith('%'):
        try:
            return float(s[:-1].strip())
        except ValueError:
            return None
    m = re.match(r'^(-?\d+(?:\.\d+)?)\s*([bmk])?$', s, re.I)
    if m:
        val = float(m.group(1))
        mult = {'b': 1000.0, 'm': 1.0, 'k': 0.001}.get((m.group(2) or 'm').lower(), 1.0)
        return val * mult
    try:
        return float(s)
    except ValueError:
        return None


def _norm_name(s):
    """Lowercase, collapse whitespace, strip — for prefix comparison."""
    return re.sub(r'\s+', ' ', str(s or '').lower()).strip()


def parse_qualitative_file(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # Optional second sheet of firm write-ups (Firm | Description) used to
    # populate the "New Managers" section of the Word rebalance memo.
    desc_rows = None
    for sn in wb.sheetnames:
        low = sn.lower()
        if sn != ws.title and ('descrip' in low or 'write' in low or 'narrative' in low):
            desc_rows = list(wb[sn].iter_rows(values_only=True))
            break
    wb.close()

    warnings = []
    if not rows:
        return {'firms': {}, 'firm_order': [], 'n_firms': 0,
                'warnings': ['Sheet is empty.']}

    # Locate the header row (name column + at least one firm attribute).
    header_idx = None
    col_map = {}
    for r_idx in range(min(10, len(rows))):
        cand = {}
        for c_idx, val in enumerate(rows[r_idx]):
            h = _norm_hdr(val)
            if not h:
                continue
            for field, syns in _FIELD_SYNONYMS.items():
                if h in syns and field not in cand.values():
                    cand[c_idx] = field
                    break
        if 'firm_name' in cand.values() and (
                'firm_aum' in cand.values() or 'ownership' in cand.values()
                or 'diverse_pct' in cand.values()):
            header_idx = r_idx
            col_map = cand
            break

    if header_idx is None:
        return {'firms': {}, 'firm_order': [], 'n_firms': 0,
                'warnings': ['Could not find a header row. Expected columns like '
                             '"Firm", "Firm AUM", "Ownership", '
                             '"Diverse/Female Ownership %" within the first 10 rows.']}

    idx = {field: c for c, field in col_map.items()}
    name_c = idx.get('firm_name')
    faum_c = idx.get('firm_aum')
    own_c  = idx.get('ownership')
    div_c  = idx.get('diverse_pct')

    firms = {}
    for row in rows[header_idx + 1:]:
        name = row[name_c] if name_c is not None and name_c < len(row) else None
        name = str(name).strip() if name is not None else ''
        faum = _to_float(row[faum_c]) if faum_c is not None and faum_c < len(row) else None
        own  = row[own_c] if own_c is not None and own_c < len(row) else None
        own  = str(own).strip() if own not in (None, '') else None
        div  = _to_float(row[div_c]) if div_c is not None and div_c < len(row) else None

        if not name and faum is None and own is None and div is None:
            continue  # blank row
        if not name:
            continue  # attribute values with no firm name — skip
        if name in firms:
            warnings.append(f'Firm "{name}" listed more than once — last wins.')
        firms[name] = {'firm_aum': faum, 'ownership': own, 'diverse_pct': div,
                       'description': None}

    # Diverse/woman-ownership scale normalization. This column is meant to be a
    # whole percent (0–100): the /diverse_ownership threshold (default 50) and
    # the UI both assume that scale. But Excel cells formatted AS a percentage
    # are read back by openpyxl (data_only) as fractions — "55%" → 0.55 — which
    # makes every figure 100× too small. If every non-null diverse value is a
    # fraction in (0, 1], treat the column as percent-formatted and rescale to
    # whole percent. Whole-number files (e.g. 55, 100) have a max > 1 and are
    # left untouched.
    div_vals = [f['diverse_pct'] for f in firms.values() if f['diverse_pct'] is not None]
    if div_vals and 0 < max(div_vals) <= 1.0 + 1e-9:
        for f in firms.values():
            if f['diverse_pct'] is not None:
                f['diverse_pct'] = round(f['diverse_pct'] * 100.0, 4)
        warnings.append('Diverse/woman ownership values looked like fractions '
                        '(all ≤ 1) — rescaled ×100 to whole percent.')

    # Merge in firm write-ups from the optional Descriptions sheet.
    if desc_rows:
        _attach_descriptions(firms, desc_rows, warnings)

    # Longest firm name first so prefix matching prefers the most specific firm.
    firm_order = sorted(firms.keys(), key=lambda f: len(_norm_name(f)), reverse=True)

    return {
        'firms':      firms,
        'firm_order': firm_order,
        'n_firms':    len(firms),
        'warnings':   warnings,
    }


def _attach_descriptions(firms, desc_rows, warnings):
    """Merge a Firm | Description sheet into `firms` (adds 'description').
    Firms present only in the descriptions sheet are added with null attrs."""
    _FIRM_HDRS = ('firm', 'firm name', 'manager', 'name',
                  'firm / strategy', 'firm/strategy')
    _DESC_HDRS = ('description', 'descriptions', 'write up', 'write-up',
                  'writeup', 'narrative', 'summary', 'blurb', 'bio', 'overview')
    hdr_idx = fc = dc = None
    for i, r in enumerate(desc_rows[:10]):
        cols = {}
        for ci, v in enumerate(r):
            h = _norm_hdr(v)
            if h in _FIRM_HDRS:
                cols.setdefault('firm', ci)
            elif h in _DESC_HDRS:
                cols.setdefault('desc', ci)
        if 'firm' in cols and 'desc' in cols:
            hdr_idx, fc, dc = i, cols['firm'], cols['desc']
            break
    if hdr_idx is None:
        warnings.append('Descriptions sheet found but no "Firm" + "Description" '
                        'header row detected in the first 10 rows.')
        return
    for r in desc_rows[hdr_idx + 1:]:
        nm = str(r[fc]).strip() if fc < len(r) and r[fc] is not None else ''
        txt = r[dc] if dc < len(r) else None
        txt = str(txt).strip() if txt not in (None, '') else None
        if not nm or not txt:
            continue
        if nm in firms:
            firms[nm]['description'] = txt
        else:
            firms[nm] = {'firm_aum': None, 'ownership': None,
                         'diverse_pct': None, 'description': txt}


def match_firm(strategy_name, qd):
    """Return (firm_name, firm_record) for a strategy via prefix match, or
    (None, None). A strategy matches a firm if its normalised name starts with
    the firm's normalised name (whole-token boundary, so 'Arga' matches
    'Arga ISC' but not 'Argatroban')."""
    if not qd or not qd.get('firms'):
        return None, None
    sn = _norm_name(strategy_name)
    if not sn:
        return None, None
    for firm in qd.get('firm_order', list(qd['firms'].keys())):
        fn = _norm_name(firm)
        if not fn:
            continue
        # exact, or prefix followed by a space (token boundary)
        if sn == fn or sn.startswith(fn + ' '):
            return firm, qd['firms'][firm]
    return None, None
